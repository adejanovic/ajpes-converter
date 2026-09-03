from flask import Flask, request, send_file, render_template_string, jsonify
import os, re, shutil, tempfile, uuid, subprocess, calendar, datetime, unicodedata
from collections import defaultdict
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

try:
    import pdfplumber
    import openpyxl
    try:
        import xlrd
    except ImportError:
        xlrd = None
    try:
        import pytesseract
    except ImportError:
        pytesseract = None
    try:
        from PIL import Image, ImageOps, ImageFilter
    except ImportError:
        Image = None
        ImageOps = None
        ImageFilter = None
except ImportError:
    pass

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

TEMPLATE_PATH    = Path(__file__).parent / "template.xlsx"
TEMPLATE_SP_PATH = Path(__file__).parent / "template_sp.xlsx"
TEMPLATE_DRUSTVO_PATH = Path(__file__).parent / "template_drustvo.xlsx"
UPLOAD_DIR      = Path(tempfile.gettempdir()) / "ajpes_uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
ACCESS_PASSWORD = os.environ.get("ACCESS_PASSWORD", "bilance2025")

# ── JOLP → AOP mapping ────────────────────────────────────────────────────────
JOLP_TO_AOP = {
    "SREDSTVA": "001",
    "A. DOLGOROČNA SREDSTVA": "002",
    "I. Neopredmetena sredstva in dolgoročne aktivne časovne razmejitve": "003",
    "1. Neopredmetena sredstva": "004",
    "2. Dolgoročne aktivne časovne razmejitve": "009",
    "II. Opredmetena osnovna sredstva": "010",
    "III. Naložbene nepremičnine": "018",
    "IV. Dolgoročne finančne naložbe": "019",
    "1. Dolgoročne finančne naložbe, razen posojil": "020",
    "2. Dolgoročna posojila": "024",
    "V. Dolgoročne poslovne terjatve": "027",
    "VI. Odložene terjatve za davek": "031",
    "B. KRATKOROČNA SREDSTVA": "032",
    "I. Sredstva (skupine za odtujitev) za prodajo": "033",
    "II. Zaloge": "034",
    "III. Kratkoročne finančne naložbe": "040",
    "1. Kratkoročne finančne naložbe, razen posojil": "041",
    "2. Kratkoročna posojila": "045",
    "IV. Kratkoročne poslovne terjatve": "048",
    "V. Denarna sredstva": "052",
    "C. KRATKOROČNE AKTIVNE ČASOVNE RAZMEJITVE": "053",
    "Zunajbilančna sredstva": "054",
    "OBVEZNOSTI DO VIROV SREDSTEV": "055",
    "A. KAPITAL": "056",
    "I. Vpoklicani kapital": "057",
    "1. Osnovni kapital": "058",
    "2. Nevpoklicani kapital (kot odbitna postavka)": "059",
    "II. Kapitalske rezerve": "060",
    "III. Rezerve iz dobička": "061",
    "IV. Revalorizacijske rezerve": "067",
    "V. Rezerve, nastale zaradi vrednotenja po pošteni vrednosti": "301",
    "VI. Preneseni čisti poslovni izid (preneseni čisti dobiček/izguba)": "068",
    "VII. Čisti poslovni izid poslovnega leta (čisti dobiček/čista izguba poslovnega leta)": "070",
    "B. REZERVACIJE IN DOLGOROČNE PASIVNE ČASOVNE RAZMEJITVE": "072",
    "1. Rezervacije": "073",
    "2. Dolgoročne pasivne časovne razmejitve": "074",
    "C. DOLGOROČNE OBVEZNOSTI": "075",
    "I. Dolgoročne finančne obveznosti": "076",
    "II. Dolgoročne poslovne obveznosti": "080",
    "III. Odložene obveznosti za davek": "084",
    "Č. KRATKOROČNE OBVEZNOSTI": "085",
    "I. Obveznosti, vključene v skupine za odtujitev": "086",
    "II. Kratkoročne finančne obveznosti": "087",
    "III. Kratkoročne poslovne obveznosti": "091",
    "D. KRATKOROČNE PASIVNE ČASOVNE RAZMEJITVE": "095",
    "Zunajbilančne obveznosti": "096",
    "1. ČISTI PRIHODKI OD PRODAJE": "110",
    "A. ČISTI PRIHODKI OD PRODAJE": "110",
    "2. SPREMEMBA VREDNOSTI ZALOG PROIZVODOV IN NEDOKONČANE PROIZVODNJE": "121",
    "B. POVEČANJE VREDNOSTI ZALOG PROIZVODOV IN NEDOKONČANE PROIZVODNJE": "121",
    "3. USREDSTVENI LASTNI PROIZVODI IN LASTNE STORITVE": "123",
    "Č. USREDSTVENI LASTNI PROIZVODI IN LASTNE STORITVE": "123",
    "4. DRUGI POSLOVNI PRIHODKI": "125",
    "5. Stroški blaga, materiala in storitev": "128",
    "a) Nabavna vrednost prodanega blaga in materiala ter stroški porabljenega materiala": "129",
    "b) Stroški storitev": "134",
    "6. Stroški dela": "139",
    "a) Stroški plač": "140",
    "b) Stroški pokojninskih zavarovanj": "141",
    "c) Stroški drugih socialnih zavarovanj": "142",
    "č) Drugi stroški dela": "143",
    "7. Odpisi vrednosti": "144",
    "a) Amortizacija": "145",
    "b) Prevrednotovalni poslovni odhodki pri neopredmetenih sredstvih in opredmetenih osnovnih sredstvih": "146",
    "c) Prevrednotovalni poslovni odhodki pri obratnih sredstvih": "147",
    "8. Drugi poslovni odhodki": "148",
    "9. Finančni prihodki iz deležev": "155",
    "10. Finančni prihodki iz danih posojil": "160",
    "11. Finančni prihodki iz poslovnih terjatev": "163",
    "12. Finančni odhodki iz oslabitve in odpisov finančnih naložb": "168",
    "13. Finančni odhodki iz finančnih obveznosti": "169",
    "14. Finančni odhodki iz poslovnih obveznosti": "174",
    "15. DRUGI PRIHODKI": "178",
    "16. DRUGI ODHODKI": "181",
    "17. DAVEK IZ DOBIČKA": "184",
    "18. ODLOŽENI DAVKI": "185",
    "19. ČISTI POSLOVNI IZID OBRAČUNSKEGA OBDOBJA (ČISTI DOBIČEK/IZGUBA OBRAČUNSKEGA OBDOBJA)": "186",
    "21. PRENESENI DOBIČEK/IZGUBA": "202",
    "25. BILANČNI DOBIČEK/IZGUBA": "215",

    # --- Dodatne (podrobnejše) postavke, dodane naknadno, da pokrivajo celoten uradni obrazec ---
    "a) Dolgoročne premoženjske pravice": "005",
    "b) Dobro ime": "006",
    "c) Dolgoročno odloženi stroški razvijanja": "007",
    "č) Druga neopredmetena sredstva": "008",
    "1. Zemljišča": "011",
    "2. Zgradbe": "012",
    "3. Proizvajalne naprave in stroji": "013",
    "4. Druge naprave in oprema, drobni inventar in druga opredmetena osnovna sredstva": "014",
    "5. Biološka sredstva": "015",
    "6. Opredmetena osnovna sredstva v gradnji in izdelavi": "016",
    "7. Predujmi za pridobitev opredmetenih osnovnih sredstev": "017",
    "c) Druge dolgoročne finančne naložbe": "023",
    "a) Dolgoročna posojila družbam v skupini": "025",
    "b) Druga dolgoročna posojila": "026",
    "1. Dolgoročne poslovne terjatve do družb v skupini": "028",
    "2. Dolgoročne poslovne terjatve do kupcev": "029",
    "3. Dolgoročne poslovne terjatve do drugih": "030",
    "1. Material": "035",
    "2. Nedokončana proizvodnja": "036",
    "3. Proizvodi": "037",
    "4. Trgovsko blago": "038",
    "5. Predujmi za zaloge": "039",
    "c) Druge kratkoročne finančne naložbe": "044",
    "a) Kratkoročna posojila družbam v skupini": "046",
    "b) Druga kratkoročna posojila": "047",
    "1. Kratkoročne poslovne terjatve do družb v skupini": "049",
    "2. Kratkoročne poslovne terjatve do kupcev": "050",
    "3. Kratkoročne poslovne terjatve do drugih": "051",
    "1. Zakonske rezerve": "062",
    "2. Rezerve za lastne delnice in lastne poslovne deleže": "063",
    "3. Lastne delnice in lastni poslovni deleži (kot odbitna postavka)": "064",
    "4. Statutarne rezerve": "065",
    "5. Druge rezerve iz dobička": "066",
    "VII. Prenesena čista izguba": "069",
    "IX. Čista izguba poslovnega leta": "071",
    "1. Dolgoročne finančne obveznosti do družb v skupini": "077",
    "2. Dolgoročne finančne obveznosti do bank": "078",
    "3. Druge dolgoročne finančne obveznosti": "079",
    "1. Dolgoročne poslovne obveznosti do družb v skupini": "081",
    "2. Dolgoročne poslovne obveznosti do dobaviteljev": "082",
    "3. Druge dolgoročne poslovne obveznosti": "083",
    "1. Kratkoročne finančne obveznosti do družb v skupini": "088",
    "2. Kratkoročne finančne obveznosti do bank": "089",
    "3. Druge kratkoročne finančne obveznosti": "090",
    "1. Kratkoročne poslovne obveznosti do družb v skupini": "092",
    "2. Kratkoročne poslovne obveznosti do dobaviteljev": "093",
    "3. Druge kratkoročne poslovne obveznosti": "094",
    "I. Čisti prihodki od prodaje na domačem trgu": "111",
    "1. Čisti prihodki od prodaje proizvodov in storitev razen najemnin": "112",
    "2. Čisti prihodki od najemnin": "113",
    "3. Čisti prihodki od prodaje blaga in materiala": "114",
    "II. Čisti prihodki od prodaje na trgu EU": "115",
    "III. Čisti prihodki od prodaje na trgu izven EU": "118",
    "C. ZMANJŠANJE VREDNOSTI ZALOG PROIZVODOV IN NEDOKONČANE PROIZVODNJE": "122",
    "D. SUBVENCIJE, DOTACIJE, REGRESI, KOMPENZACIJE IN DRUGI PRIHODKI, KI SO POVEZANI S POSLOVNIMI UČINKI": "124",
    "F. KOSMATI DONOS OD POSLOVANJA": "126",
    "G. POSLOVNI ODHODKI": "127",
    "2. Stroški porabljenega materiala": "130",
    "a) stroški materiala": "131",
    "b) stroški energije": "132",
    "c) drugi stroški materiala": "133",
    "a) transportne storitve": "135",
    "b) najemnine": "136",
    "c) povračila stroškov zaposlenim v zvezi z delom": "137",
    "č) drugi stroški storitev": "138",
    "1. Rezervacije": "149",
    "2. Drugi stroški": "150",
    "H. DOBIČEK IZ POSLOVANJA": "151",
    "I. IZGUBA IZ POSLOVANJA": "152",
    "J. FINANČNI PRIHODKI": "153",
    "Finančni prihodki od obresti (upoštevano že v II. in III.)": "154",
    "1. Finančni prihodki iz deležev v družbah v skupini": "156",
    "2. Finančni prihodki iz deležev v pridruženih družbah": "157",
    "3. Finančni prihodki iz deležev v drugih družbah": "158",
    "4. Finančni prihodki iz drugih naložb": "159",
    "1. Finančni prihodki iz posojil, danih družbam v skupini": "161",
    "2. Finančni prihodki iz posojil, danih drugim": "162",
    "1. Finančni prihodki iz poslovnih terjatev do družb v skupini": "164",
    "2. Finančni prihodki iz poslovnih terjatev do drugih": "165",
    "K. FINANČNI ODHODKI": "166",
    "Finančni odhodki za obresti (upoštevano že v II. in III.)": "167",
    "1. Finančni odhodki iz posojil, prejetih od družb v skupini": "170",
    "2. Finančni odhodki iz posojil, prejetih od bank": "171",
    "3. Finančni odhodki iz izdanih obveznic": "172",
    "4. Finančni odhodki iz drugih finančnih obveznosti": "173",
    "1. Finančni odhodki iz poslovnih obveznosti do družb v skupini": "175",
    "2. Finančni odhodki iz obveznosti do dobaviteljev in meničnih obveznosti": "176",
    "3. Finančni odhodki iz drugih poslovnih obveznosti": "177",
    "I. Subvencije, dotacije in podobni prihodki, ki niso povezani s poslovnimi učinki": "179",
    "II. Ostali prihodki": "180",
    "N. CELOTNI DOBIČEK": "182",
    "O. CELOTNA IZGUBA": "183",
    "Š. ČISTA IZGUBA OBRAČUNSKEGA OBDOBJA": "187",
    "*POVPREČNO ŠTEVILO ZAPOSLENIH NA PODLAGI DELOVNIH UR V OBRAČUNSKEM OBDOBJU (na dve decimalki)": "188",

    # --- Alias-i za besedilne variante, ki se pojavljajo v nekaterih internih (ne-AJPES) poročilih ---
    "B. Rezervacije": "072",                          # skrajšano ime za "B. REZERVACIJE IN DOLGOROČNE PASIVNE ČASOVNE RAZMEJITVE"
    "b) Dolgoročna posojila drugim": "026",           # = "b) Druga dolgoročna posojila"
    "b) Kratkoročna posojila drugim": "047",          # = "b) Druga kratkoročna posojila"
    "KOSMATI DONOS IZ POSLOVANJA": "126",             # = "F. KOSMATI DONOS OD POSLOVANJA"
    "I. Neopred. dolgoročna sredstva in dolgoročne aktivnečasovne razmejitve": "003",  # okrajšava "Vizija" programa
    # --- Alias-i za "AVTO ELITE" / PLAIN_MST izvoz (drugačno besedilo/okrajšave/tipkarske napake) ---
    "I. NEOPREDMETENA SREDSTVA IN DOLG. AKTIVNE ČASOVNE RAZMEJITVE": "003",
    "IV. KRATOROČNE POSLOVNE TERJATVE": "048",              # tipkarska napaka izvora: "Kratoročne" namesto "Kratkoročne"
    "II. Dogoročne poslovne obveznosti": "080",             # tipkarska napaka izvora: "Dogoročne" namesto "Dolgoročne"
    "1. Čisti prihodki iz prodaje": "110",                  # "iz prodaje" namesto uradnega "od prodaje"
    "a) Čisti prihodki od prodaje doseženi na domačem trgu": "111",
    "a) Nabavna vrednost prodanih blaga in materiala ter stroški porabljenega materiala": "129",
    "- od tega stroški pokojninskih zavarovanj": "141",     # informativna podvrstica pod "Stroški socialnih zavarovanj"
    "b) Prevrednotovalni poslovni odhodki pri neopredmetenih sredstvih in opr. os. sr.": "146",
    "c) Prevrednotevalni poslovni odhodki pri obratnih sredstvih": "147",  # tipkarska napaka: "Prevrednotevalni"
    "12. FIN. ODHODKI IZ OSLABITVE IN ODPISOV FIN. NALOŽB": "168",
    "2. SPREMEMBA VREDNOSTI ZALOG PROIZVODOV IN NEDOK. PROIZ.": "121",
    "7. Predujmi za pridobitev opredmetena OS": "017",  # = "7. Predujmi za pridobitev opredmetenih osnovnih sredstev"
    "6. Opredmetena OS v gradnji in izdelavi": "016",   # = "6. Opredmetena osnovna sredstva v gradnji in izdelavi"
    "VI. Preneseni čisti dobiček": "068",
    "VIII. Čisti dobiček poslovnega leta": "070",
    "1. Neopredmetena osnovna sredstva": "004",         # = "1. Neopredmetena sredstva"
    "S. ČISTI DOBIČEK OBRAČUNSKEGA OBDOBJA": "186",
}

SIZE_MAP = {
    "Mikro podjetje":"1","Majhno podjetje":"2",
    "Srednje podjetje":"3","Veliko podjetje":"4",
}

TIP_BILANCE_MAP = {
    "3": "3",   # Revidirana
    "4": "4",   # Zaključena nerevidirana
    "5": "5",   # Konsolidirana revidirana
    "6": "6",   # Konsolidirana nerevidirana
    "7": "7",   # Preliminarna
    "8": "8",   # Zaključena
}

TIP_BILANCE_LABELS = {
    "3": "REVIDIRANA",
    "4": "ZAKLJUCENA_NEREVIDIRANA",
    "5": "KONSOLIDIRANA_REVIDIRANA",
    "6": "KONSOLIDIRANA_NEREVIDIRANA",
    "7": "PRELIMINARNA",
    "8": "ZAKLJUCENA",
}

TIP_SUBJEKTA_MAP = {
    "1": "Gospodarske družbe",
    "2": "Samostojni podjetniki",
    "3": "Zadruge",
    "4": "Društva",
    "5": "Pravne osebe javnega prava",
    "6": "Pravne osebe zasebnega prava",
    "7": "Banke",
    "8": "Zavarovalnice",
    "9": "Druge osebe javnega prava",
}

DRUSTVO_KEYWORDS = [
    "SKLAD", "Društveni sklad", "PRESEŽEK POSLOVNIH PRIHODKOV",
    "PRESEŽEK POSLOVNIH ODHODKOV", "društvo", "Društvo", "DRUŠTVO"
]

# ── Podatkovne strukture ──────────────────────────────────────────────────────
@dataclass
class CompanyInfo:
    name:str=""; registration_number:str=""; tax_number:str=""
    period_from:str=""; period_to:str=""
    tip_bilance:str=""; tip_subjekta:str=""; obdobje_bilance:str=""

@dataclass
class AopEntry:
    aop:str; current_year:Optional[float]=None; previous_year:Optional[float]=None

@dataclass
class ParseResult:
    company:CompanyInfo = field(default_factory=CompanyInfo)
    aop_data:dict       = field(default_factory=dict)
    gvaop_data:dict      = field(default_factory=dict)  # za postavke brez AOP kode (npr. društva) - kljuc je GVAOP (brez oklepajev)
    warnings:list       = field(default_factory=list)
    errors:list         = field(default_factory=list)
    pdf_format:str      = ""
    subject_type:str    = "GD"   # "GD" = Gospodarska družba, "DR" = Društvo

# ── Skupne funkcije ───────────────────────────────────────────────────────────
def parse_si(raw):
    if not raw: return None
    try: return float(str(raw).strip().replace(".","").replace(",","."))
    except: return None

def _detect_subject_type_from_text(text: str) -> str:
    """Zazna ali gre za društvo ali gospodarsko družbo iz vsebine Excel/PDF."""
    for kw in DRUSTVO_KEYWORDS:
        if kw in text:
            return "DR"
    return "GD"

def _infer_period_from(period_to: str) -> str:
    """
    Izpelje začetek obdobja iz konca obdobja.
    31.12.2025 → 1.1.2025  (standardno leto)
    31.03.2026 → 1.4.2025  (nestandardno: april-mart, prejšnje leto)
    31.03.2025 → 1.4.2024
    Splošno: vzame dan+1 in mesec+1, leto prilagodi.
    """
    try:
        parts = period_to.split('.')
        day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
        # Standardno leto: konec = 31.12
        if month == 12:
            return f"1.1.{year}"
        # Nestandardno: začetek = 1.(month+1).(year-1)
        # npr. konec 31.03.2026 → začetek 1.4.2025
        return f"1.{month+1}.{year-1}"
    except Exception:
        year = period_to.split('.')[-1] if '.' in period_to else period_to[-4:]
        return f"1.1.{year}"

def finalize_company(c, tip_override=None, tip_subjekta_override=None):
    if tip_subjekta_override and tip_subjekta_override in TIP_SUBJEKTA_MAP:
        c.tip_subjekta = tip_subjekta_override
    elif not c.tip_subjekta:
        c.tip_subjekta = "1"

    # Matična številka za slovenske subjekte mora biti 10-mestna (CRMB/banka).
    # Uradna AJPES matična št. je običajno 7-mestna - dopolnimo jo s tremi ničlami na koncu.
    if c.registration_number:
        reg = str(c.registration_number).strip()
        if reg.isdigit() and len(reg) < 10:
            reg = reg.ljust(10, '0')
        c.registration_number = reg

    if c.period_to and (not c.period_from or len(str(c.period_from)) < 6):
        c.period_from = _infer_period_from(c.period_to)

    if c.period_to:
        ye = c.period_to.startswith('31.12')
        # Tip bilance: uporabi override če ga imamo, drugače avtomatsko
        if tip_override and tip_override in TIP_BILANCE_MAP:
            c.tip_bilance = tip_override
        else:
            c.tip_bilance = "8" if ye else "7"
        # 7 = preliminarna/medletna, ostali tipi iz template-a so zaključni računi
        c.obdobje_bilance = "ML" if c.tip_bilance == "7" else "ZR"
    return c

# ── PARSER: UPR format (interno mesečno poročilo - "BILANCA STANJA/IZKAZ POSLOVNEGA IZIDA družbe X") ──
# Ta format nima AOP oznak, samo besedilne postavke + 2 (BS) ali 3 (IPI) stolpce zneskov.
# Nekateri PDFji tega formata "razcepijo" vodilno števko zneska od preostanka (font/kerning napaka
# izvornega poročevalskega orodja), zato zneskov ne beremo iz linearnega besedila (extract_text),
# ampak iz pozicij posameznih besed (extract_words) in jih rekonstruiramo po x/y koordinatah.
UPR_VALUE_X_MIN     = 340   # ločnica med stolpcem z opisom postavke in stolpci z zneski
UPR_BS_BOUNDARIES   = [415]        # BS: 1 ločnica -> 2 stolpca (tekoče, primerjalno)
UPR_IPI_BOUNDARIES  = [415, 484]   # IPI: 2 ločnici -> 3 stolpci (mesec, kumulativa tekoče, kumulativa primerjalno)

def _upr_is_valueish(text):
    t = text.strip()
    return bool(re.fullmatch(r'\(?-?[\d.,]+\)?', t)) and any(ch.isdigit() for ch in t)

def _upr_cluster_numbers(val_words, gap_thresh=15):
    """Združi besedne fragmente istega zneska (npr. '9' + '.339.166,81') v celoto,
    glede na majhno x-razdaljo med njimi; večja razdalja pomeni nov stolpec."""
    groups, cur, prev_x1 = [], [], None
    for w in sorted(val_words, key=lambda w: w['x0']):
        if prev_x1 is not None and (w['x0'] - prev_x1) > gap_thresh:
            groups.append(cur); cur = []
        cur.append(w); prev_x1 = w['x1']
    if cur: groups.append(cur)
    return [(''.join(x['text'] for x in g), g[0]['x0']) for g in groups]

def _upr_reconstruct_rows(words):
    """
    Vrne seznam (label, [(besedilo_zneska, x0), ...]) — po eno postavko za vsako "vrstico z zneski".
    Opis postavke se lahko razteza čez več fizičnih vrstic (prelom besedila); vse besede opisa se
    pripišejo najbližji vrstici z zneski (po vertikalni legi), tako da se ohrani celoten opis tudi
    če znesek vizualno "pade" med dve vrstici preloma.
    """
    from collections import defaultdict as _dd
    rows = _dd(list)
    for w in words:
        rows[round(w['top'], 1)].append(w)
    row_tops = sorted(rows.keys())
    value_rows, label_rows = [], []
    for top in row_tops:
        ws = sorted(rows[top], key=lambda w: w['x0'])
        val_ws = [w for w in ws if w['x0'] >= UPR_VALUE_X_MIN and _upr_is_valueish(w['text'])]
        lbl_ws = [w for w in ws if w['x0'] < UPR_VALUE_X_MIN]
        if val_ws: value_rows.append((top, val_ws))
        if lbl_ws: label_rows.append((top, lbl_ws))
    value_tops = [t for t, _ in value_rows]
    entries = []
    for i, (top, val_ws) in enumerate(value_rows):
        lower = (value_tops[i-1] + top) / 2 if i > 0 else -1e9
        upper = (value_tops[i+1] + top) / 2 if i < len(value_tops) - 1 else 1e9
        lbls = sorted([(lt, lw) for (lt, lw) in label_rows if lower < lt <= upper], key=lambda x: x[0])
        label_text = ' '.join(' '.join(w['text'] for w in lw) for _, lw in lbls).strip()
        entries.append((label_text, _upr_cluster_numbers(val_ws)))
    return entries

def _upr_bucket_columns(numgroups, boundaries):
    """Razporedi zneske v stolpce po x0 poziciji (ne po vrstnem redu v seznamu) —
    tako '-' (manjkajoč znesek, ni besede) ne zamakne preostalih stolpcev."""
    n = len(boundaries) + 1
    result = [None] * n
    for text, x0 in numgroups:
        col = sum(1 for b in boundaries if x0 >= b)
        if col < n and result[col] is None:
            result[col] = text
    return result

def _upr_parse_number(s):
    if s is None: return None
    neg = '(' in s or ')' in s
    s = s.replace('(', '').replace(')', '')
    v = parse_si(s)
    if v is None: return None
    return -v if neg else v

_JOLP_TO_AOP_CI = {k.strip().lower(): v for k, v in JOLP_TO_AOP.items()}
_JOLP_TO_AOP_CI_NOPAREN = {re.sub(r'\s*\(.*$', '', k).strip().lower(): v for k, v in JOLP_TO_AOP.items()}

def _classify_and_strip_item(s):
    """Vrne (nivo_oznake, besedilo_brez_oznake_in_oklepaja).
    'top'  = skupinski seštevek najvišjega nivoja (A./B./C./Č./D. ...) — nikoli se ne sme
             pomešati s posamezno postavko znotraj skupine (npr. 'B. Rezervacije' proti
             '1. Rezervacije' — to sta različni AOP kodi).
    'item' = katerakoli posamezna postavka, ne glede na globino gnezdenja (I., 1., a) ...) —
             različni PDFji isto postavko gnezdijo na različnih globinah (npr. AJPES obrazec ima
             ločeni '1. Zemljišča' in '2. Zgradbe', nekateri interni PDFji pa to združijo v
             '1. Zemljišča in zgradbe' s podpostavkama 'a) Zemljišča' / 'b) Zgradbe') — te je
             treba obravnavati kot ujemajoče, saj gre za isto dejansko postavko."""
    s = s.strip()
    s = re.sub(r'\s*\(.*$', '', s)
    # Varne enocrkovne oznake najvišjega nivoja (skupinski seštevki A./B./C.../D., pa tudi IPI-jeve
    # G./H./K./N./O./Š. ipd.) — IZKLJUČNO 'I' in 'V' sta izključena, ker se v teh obrazcih uporabljata
    # SAMO kot rimski številki (I., V., in kombinacije), nikoli kot oznaka najvišjega nivoja.
    m = re.match(r'^([ABCČDFGHJKNOPRSŠŽ])\.\s*', s)
    if m: return ('top', s[m.end():].strip().lower())
    m = re.match(r'^([IVXLCDM]{1,6})\.\s*', s, flags=re.I)
    if m: return ('item', s[m.end():].strip().lower())
    m = re.match(r'^(\d{1,2}[a-zčžš]?)\.\s*', s, flags=re.I)
    if m: return ('item', s[m.end():].strip().lower())
    m = re.match(r'^([a-zčžš])\)\s*', s, flags=re.I)
    if m: return ('item', s[m.end():].strip().lower())
    m = re.match(r'^([a-zčžš])\.\s*', s, flags=re.I)
    if m: return ('item', s[m.end():].strip().lower())
    return ('item', s.strip().lower())

def _build_normalized_aop_map():
    groups = {}
    ambiguous = set()
    for k, v in JOLP_TO_AOP.items():
        nk = _classify_and_strip_item(k)
        if not nk[1]:
            continue
        if nk in groups and groups[nk] != v:
            ambiguous.add(nk)
        groups[nk] = v
    for nk in ambiguous:
        groups.pop(nk, None)
    return groups

_JOLP_TO_AOP_NORM = _build_normalized_aop_map()

def _match_aop_label_exact(label):
    """Ujemanje postavke na AOP kodo: najprej natančno (case-insensitive), nato brez oklepaja
    na koncu, nato brez vodilne oznake ob upoštevanju njenega nivoja (za primere preštevilčenja
    rimskih/arabskih oznak, kadar vmesna postavka v poročilu manjka). Namerno BREZ 'fuzzy'/
    startswith ujemanja na celotnem besedilu — to bi lahko napačno združilo dve sosednji
    vrstici, kadar je uradna postavka prazna ('-')."""
    key = label.strip().lower()
    if not key: return None
    if key in _JOLP_TO_AOP_CI: return _JOLP_TO_AOP_CI[key]
    if key in _JOLP_TO_AOP_CI_NOPAREN: return _JOLP_TO_AOP_CI_NOPAREN[key]
    key_noparen = re.sub(r'\s*\(.*$', '', key).strip()
    if key_noparen in _JOLP_TO_AOP_CI_NOPAREN: return _JOLP_TO_AOP_CI_NOPAREN[key_noparen]
    nkey = _classify_and_strip_item(label)
    if nkey in _JOLP_TO_AOP_NORM: return _JOLP_TO_AOP_NORM[nkey]
    return None

def _match_aop_label_merged(label):
    """Nekatere vrstice v izvornem PDF-ju nimajo nobenega zneska (same '-' v vseh stolpcih),
    zato pri rekonstrukciji vrstic nimajo svoje 'vrstice z zneski' in se njihov opis pripiše
    PREJŠNJI (sosednji) postavki — npr. '2. Kratkoročne finančne obveznosti do bank
    4. Druge kratkoročne finančne obveznosti' (drugo je '-' povsod). To poskusi razrešiti:
    če se label začne z znanim, natančnim nazivom postavke, ki mu takoj sledi presledek in
    NOVA številčna/črkovna oznaka, uporabi znesek za PRVI (ujemajoči) del."""
    low = label.strip()
    low_lc = low.lower()
    best = None
    for k, v in JOLP_TO_AOP.items():
        kk = k.strip()
        klc = kk.lower()
        if len(low_lc) <= len(klc) or not low_lc.startswith(klc + ' '):
            continue
        rest = low[len(kk):].strip()
        if re.match(r'^(?:[IVXLCDM]{1,6}\.|[A-ZČŽŠ]\.|\d{1,2}[a-zčžš]?\.|[a-zčžš]\))', rest, flags=re.I):
            if best is None or len(kk) > len(best[0]):
                best = (kk, v)
    return best[1] if best else None

# Postavke, ki se v uradnem obrazcu z istim besedilom pojavijo večkrat (npr. pri dolgoročnih IN
# kratkoročnih naložbah), zato jih ni mogoče razločiti brez konteksta nadrejene postavke. Ključ je
# besedilo (brez oznake), vrednost pa slovar {kontekst: aop_koda}. Kontekst nastavijo postavke,
# navedene v _CONTEXT_TRIGGERS spodaj, glede na trenutno zaporedje branja dokumenta.
_AMBIGUOUS_ITEMS = {
    "delnice in deleži v družbah v skupini": {"dolg_fin_nalozbe": "021", "kratk_fin_nalozbe": "042"},
    "druge delnice in deleži":               {"dolg_fin_nalozbe": "022", "kratk_fin_nalozbe": "043"},
    "čisti prihodki od prodaje proizvodov in storitev": {"eu": "116", "izven_eu": "119"},
    "čisti prihodki od prodaje blaga in materiala":      {"eu": "117", "izven_eu": "120"},
}
_CONTEXT_TRIGGERS = {
    "019": "dolg_fin_nalozbe",   # IV. Dolgoročne finančne naložbe
    "040": "kratk_fin_nalozbe",  # III. Kratkoročne finančne naložbe
    "115": "eu",                 # II. Čisti prihodki od prodaje na trgu EU
    "118": "izven_eu",           # III. Čisti prihodki od prodaje na trgu izven EU
}

def _last_day_of_month(yyyy_mm: str):
    """
    Iz vnosa tipa 'YYYY-MM' (HTML input type=month) izračuna zadnji dan v mesecu
    in ga vrne v formatu 'D.M.YYYY' (brez vodilnih ničel, konsistentno z ostalim parsanjem).
    Vrne None če vnos ni veljaven.
    """
    try:
        year_s, month_s = yyyy_mm.strip().split("-")
        year, month = int(year_s), int(month_s)
        last_day = calendar.monthrange(year, month)[1]
        return f"{last_day:02d}.{month:02d}.{year}"
    except Exception:
        return None

def _subject_type_from_tip_subjekta(tip_subjekta: str) -> str:
    s = str(tip_subjekta).strip()
    if s == "2": return "SP"
    if s == "4": return "DR"
    return "GD"

def is_cid_encoded(pages_text):
    """Zazna PDFje z pokvarjenim font encodingom (cid:XX) ki jih ne moremo brati."""
    if not pages_text: return False
    sample = " ".join(pages_text[:3])
    cid_count = sample.count("(cid:")
    total_chars = len(sample.replace(" ",""))
    return cid_count > 10 and (cid_count * 6) > (total_chars * 0.3)

def is_scanned_no_text(pages_text):
    """Zazna skenirane PDFje brez kakršnegakoli besedilnega sloja (samo slika strani)."""
    if not pages_text: return False
    total_chars = sum(len(t.strip()) for t in pages_text)
    return total_chars < 20 * len(pages_text)

def validate(result):
    if not result.company.name and result.pdf_format not in ("GDZADRUGE", "OCR", "OCR_SCAN", "AJPES_NATIVE_XLS", "OBRACUNI_MEDLETNI"):
        result.errors.append("Ime podjetja ni najdeno.")
    if not result.company.registration_number and result.pdf_format not in ("UPR", "GDZADRUGE", "OCR", "OCR_SCAN", "POROCILO_GD", "VIZIJA", "ZAVOD_MEDLETNI", "AJPES_NATIVE_XLS", "URADNI", "OBRACUNI_MEDLETNI", "BS_IPI_AJPES", "AJPES_1COL", "VASCO_AJPES_1COL", "AJPES_ZAP"):
        result.errors.append("Matična številka ni najdena.")
    # Preveri 001/055 samo če je to BS datoteka (IPI-only je ok brez njiju)
    def _safe_aop_int(k):
        try: return int(str(k).rstrip('abcdefghijklmnopqrstuvwxyz'))
        except: return 0
    has_ipi = any(110 <= _safe_aop_int(k) <= 199 for k in result.aop_data)
    has_bs  = any(1   <= _safe_aop_int(k) <= 109 for k in result.aop_data)
    if (has_bs or not has_ipi) and result.pdf_format not in ("OCR", "OCR_SCAN"):
        for code in ["001","055"]:
            if code not in result.aop_data:
                result.errors.append(f"Manjka AOP {code}.")
    a1 = result.aop_data.get("001"); a5 = result.aop_data.get("055")
    if a1 and a5:
        v1 = a1.current_year or 0; v2 = a5.current_year or 0
        if abs(v1-v2) > 0.02:
            result.warnings.append("AOP 001 ≠ AOP 055 — preverite bilanco.")
    return result

# ── FORMAT DETEKCIJA ──────────────────────────────────────────────────────────
def _is_ajpes_1col_format(pages_text):
    """Prepozna PDF izvoz oblike 'TIGI TRADE' (in podobnih računovodskih programov, ki
    generirajo isto obliko) - naslov v prvih vrsticah prve strani je oblike
    'BILANCA STANJA - AJPES/DD.MM.YYYY[-DD.MM.YYYY]' ali 'IZKAZ POSLOVNEGA IZIDA -
    AJPES/...'. Format nima ločenih stolpcev za tekoče/prejšnje leto - vsaka vrstica
    ima samo EN znesek na koncu (glej _AJPES_1COL_NUM_RE), zato je previous_year pri
    tem formatu vedno prazen - to je lastnost izvora, ne napaka pri pretvorbi."""
    first = pages_text[0] if pages_text else ""
    for line in first.split('\n')[:6]:
        if "AJPES/" in line and ("BILANCA STANJA" in line or "IZKAZ POSLOVNEGA IZIDA" in line):
            return True
    return False

def _is_vasco_ajpes_1col_format(pages_text):
    """Prepozna medletni PDF izvoz programa Vasco/Rave Reports.

    BS ima naslov ``BS za družbe- Ajpes LETO/OD-DO``, IPI pa
    ``Izkaz poslovnega izida LETO- PIZ na ločenih kontih/OD-DO``. V obeh
    datotekah je stolpec ``Zap.`` (zaporedna številka postavke), na desni pa
    en sam znesek za tekoče leto. Zaporedna številka ni vedno enaka AOP kodi,
    zato tega formata ne smemo obravnavati kot navaden AOP PDF.
    """
    first = pages_text[0] if pages_text else ""
    compact = re.sub(r'\s+', ' ', first)
    has_bs_title = bool(re.search(r'\bBS\s+za\s+družbe\s*-\s*Ajpes\b', compact, re.I))
    has_ipi_title = bool(re.search(
        r'\bIzkaz\s+poslovnega\s+izida\s+\d{4}\s*-\s*PIZ\s+na\s+ločenih\s+kontih\s*/',
        compact, re.I
    ))
    has_table_header = bool(re.search(r'\bZap\.\s+', first)) and bool(
        re.search(r'\btekoče\s+leto\b', first, re.I)
    )
    return has_table_header and (has_bs_title or has_ipi_title)

def _is_plain_mst_format(pages_text):
    """Prepozna izvoz oblike npr. 'AVTO ELITE D.O.O.' — glava ima ime podjetja neposredno
    (brez oznake 'Ime poslovnega subjekta') skupaj z 'Matična številka: NNNNNNN' v isti
    vrstici, naslov je gol 'BILANCA STANJA' / 'IZKAZ POSLOVNEGA IZIDA' (brez besede
    'družbe', ki bi to sicer uvrstila med UPR), postavke pa nimajo AOP kod, temveč samo
    besedilne oznake (A./I./1./a) ...), tako kot pri UPR-ju."""
    first = pages_text[0] if pages_text else ""
    has_header = bool(re.search(r'Matična številka:\s*\d{6,8}\b', first))
    has_title = bool(re.search(r'\bBILANCA STANJA\b', first) or re.search(r'\bIZKAZ POSLOVNEGA IZIDA\b', first))
    has_druzbe_title = bool(re.search(r'(?:BILANCA STANJA|IZKAZ POSLOVNEGA IZIDA)\s+družbe', first))
    return has_header and has_title and not has_druzbe_title and "AOP" not in first

def _is_besedilo_2col_format(pages_text):
    """Prepozna dvostolpčni interni izvoz BS/IPI (npr. ŠUŠTAR TRANS).

    Značilnosti tega izvoza:
      * ime podjetja je samostojno v prvi vrstici,
      * ``davčna številka`` in ``matična številka`` sta v ločenih vrsticah,
      * naslov je gol ``BILANCA STANJA`` ali ``IZKAZ POSLOVNEGA IZIDA``,
      * ni AOP stolpca, na desni pa sta tekoči in primerjalni znesek.

    Ta format je namenoma ločen od PLAIN_MST, ker ima drugačno horizontalno
    postavitev stolpcev in nekaj vsebinsko drugačnih nazivov postavk.
    """
    first = pages_text[0] if pages_text else ""
    compact = re.sub(r'\s+', ' ', first)
    lines = [ln.strip() for ln in first.split('\n') if ln.strip()]
    has_title = bool(re.search(r'\bBILANCA\s+STANJA\b|\bIZKAZ\s+POSLOVNEGA\s+IZIDA\b', compact, re.I))
    # Pomembno: ID-ja morata biti samostojni vrstici. Tako ne prestrežemo obstoječega
    # PLAIN_MST formata, kjer je "Matična številka" na isti vrstici kot ime podjetja.
    has_tax = any(re.match(r'^dav[cč]na\s+[sš]tevilka\s*:\s*\d{8}\b', ln, re.I) for ln in lines[:8])
    has_mst = any(re.match(r'^mati[cč]na\s+[sš]tevilka\s*:\s*\d{7,10}\b', ln, re.I) for ln in lines[:8])
    has_two_periods = len(re.findall(r'\b\d{1,2}\.\d{1,2}\.\d{4}\b', first)) >= 2
    return has_title and has_tax and has_mst and has_two_periods and "AOP" not in first

def _is_ajpes_zap_format(pages_text):
    """Varianta istega izvoznega programa kot VASCO_AJPES_1COL (isti naslov 'BS za
    družbe- Ajpes/...' oz. 'Izkaz poslovnega izida- Ajpes/...' s stolpcem 'Zap.'), a
    glava tabele nima besede 'tekoče leto', temveč samo 'Leto NNNN' - zato je ne ujame
    _is_vasco_ajpes_1col_format. Ker ta varianta nima zanesljivega 1:1 ujemanja med
    zaporedno številko vrstice in AOP kodo (nekateri izvozi izpustijo vrstico "V.
    Rezerve, nastale zaradi vrednotenja po pošteni vrednosti", če je vedno 0, kar
    zamakne vse nadaljnje zaporedne številke), se BS pri tej varianti bere po
    besedilu postavke (glej parse_ajpes_zap_format), IPI pa po zaporedni številki
    (ta del uradnega obrazca nima podobnih izpustov)."""
    first = pages_text[0] if pages_text else ""
    compact = re.sub(r'\s+', ' ', first)
    has_title = bool(re.search(r'\bBS\s+za\s+družbe\s*-\s*Ajpes\b', compact, re.I)) or \
                bool(re.search(r'\bIzkaz\s+poslovnega\s+izida\s*-\s*Ajpes\b', compact, re.I))
    has_zap = bool(re.search(r'\bZap\.\s+', first))
    return has_title and has_zap

def detect_pdf_format(pages_text):
    first = pages_text[0] if pages_text else ""
    lines = first.split('\n')
    # VIZIJA: "Vizija računovodstvo" izvoz ("DENAR | OBRAČUNI" v glavi, brez AOP kode)
    if "DENAR" in first and "OBRAČUNI" in first:
        return "VIZIJA"
    # NAPOVED: ima "AOP Konto POSTAVKA" v headerju
    if any("AOP Konto" in l or "AOP" == l.strip()[:3] for l in lines[:8]):
        # Preverimo ali so vrstice ki začnejo z AOP kodo
        for line in lines[5:10]:
            if re.match(r'^\d{3}\s+', line.strip()):
                return "NAPOVED"
    # URADNI: čist uraden AJPES obrazec ("Oznaka za AOP" + zaporedje zneskov, podvojena ali enojna koda)
    if "Oznaka" in first and "za AOP" in first:
        return "URADNI"
    # AOPCOL: ima eksplicitno AOP kolono v besedilu ("Postavka AOP Tekoče leto Preteklo leto")
    if "Postavka AOP" in first and "Tekoče leto" in first:
        return "AOPCOL"
    # Vasco/Rave Reports: medletna BS in IPI v ločenih PDF-jih, z zaporednimi
    # številkami postavk in enim samim stolpcem zneskov.
    if _is_vasco_ajpes_1col_format(pages_text):
        return "VASCO_AJPES_1COL"
    if _is_ajpes_zap_format(pages_text):
        return "AJPES_ZAP"
    # AJPES_1COL: izvoz oblike "TIGI TRADE" ipd. - polno besedilo + EN sam znesek na
    # vrstico (brez ločenih stolpcev tekoče/prejšnje leto), naslov v obliki
    # "BILANCA STANJA - AJPES/DD.MM.YYYY[-DD.MM.YYYY]" ali "IZKAZ POSLOVNEGA IZIDA - AJPES/...".
    if _is_ajpes_1col_format(pages_text):
        return "AJPES_1COL"
    # UPR: interno mesečno poročilo ("BILANCA STANJA družbe X" / "IZKAZ POSLOVNEGA IZIDA družbe X")
    if re.search(r'BILANCA STANJA\s+družbe', first) or re.search(r'IZKAZ POSLOVNEGA IZIDA\s+družbe', first):
        return "UPR"
    # Dvostolpčni interni izvoz z ločenima vrsticama za davčno/matično številko
    # (npr. ŠUŠTAR TRANS). Preverimo ga pred PLAIN_MST, ker je naslov prav tako gol.
    if _is_besedilo_2col_format(pages_text):
        return "BESEDILO_2COL"
    # PLAIN_MST: izvoz z imenom podjetja in matično št. v glavi, brez AOP kod (npr. AVTO ELITE)
    if _is_plain_mst_format(pages_text):
        return "PLAIN_MST"
    # AOP (lp2025): bold tekst — 4x ponavljanje
    first_line = lines[0] if lines else ""
    sample = first_line[:20].replace(" ","")
    chunks = [sample[i:i+4] for i in range(0,len(sample)-3,4)]
    if sum(1 for c in chunks if len(c)==4 and len(set(c))==1) >= 2:
        return "AOP"
    # JOLP: ime podjetja prva vrstica, "Bilanca stanja na dan" v prvih vrsticah
    if "Bilanca stanja na dan" in first or "Podatki so v EUR" in first:
        return "JOLP"
    return "AOP"

def detect_napoved_type(pages_text):
    """Vrne 'BS' ali 'IPI' za NAPOVED format."""
    first = pages_text[0] if pages_text else ""
    if "BILANCA STANJA" in first: return "BS"
    if "IZKAZ POSLOVNEGA IZIDA" in first: return "IPI"
    return "BS"

# ── PARSER: AOP format (lp2025) ──────────────────────────────────────────────
def decode_bold(line):
    result=[]; i=0
    while i < len(line):
        ch=line[i]
        if ch==' ':
            while i<len(line) and line[i]==' ': i+=1
            result.append(' ')
        elif i+3<len(line) and line[i]==line[i+1]==line[i+2]==line[i+3]:
            result.append(ch); i+=4
        else:
            result.append(ch); i+=1
    return ''.join(result).strip()

def is_bold(line):
    s=line.strip()[:32].replace(" ","")
    if len(s)<8: return False
    chunks=[s[i:i+4] for i in range(0,len(s)-3,4)]
    return sum(1 for c in chunks if len(c)==4 and len(set(c))==1)>=2

AOP_RE = re.compile(r'(\d{3})\s+([\d.]+,\d{2})(?:\s+([\d.]+,\d{2}))?')
NAPOVED_RE = re.compile(r'^(\d{3})\s+.+?\s+([\d.]+,\d{2})\s+([\d.]+,\d{2})\s*$')

def parse_aop_format(pages_text, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="AOP")
    seen = set()
    full = "\n".join(pages_text)
    c = CompanyInfo()
    m = re.search(r'Ime poslovnega subjekta[:\s]+([A-ZŠĐŽČĆ][A-ZŠĐŽČĆ\s\.\,\-]{2,60}?)(?=\n|\s{2,}|$)', full)
    if m: c.name = m.group(1).strip()
    m = re.search(r'Mati[cč]na [sš]tevilka\s+(\d{10})', full)
    if m: c.registration_number = m.group(1).strip()
    m = re.search(r'Dav[cč]na [sš]tevilka\s+(\d{8})', full)
    if m: c.tax_number = m.group(1).strip()
    m = re.search(r'\bod\s+([\d]{1,2}\.[\d]{1,2}\.[\d]{4})', full)
    if m: c.period_from = m.group(1).strip()
    m = re.search(r'\bdo\s+([\d]{1,2}\.[\d]{1,2}\.[\d]{4})', full)
    if m: c.period_to = m.group(1).strip()
    # Tip subjekta je vedno 1 (gospodarska družba) — banka to nastavi posebej
    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    for page_text in pages_text[1:]:
        for line in page_text.split('\n'):
            decoded = decode_bold(line) if is_bold(line) else line
            m = AOP_RE.search(decoded)
            if not m: continue
            aop = m.group(1)
            if not (1 <= int(aop) <= 310): continue
            if aop in seen: continue
            result.aop_data[aop] = AopEntry(aop=aop,
                current_year=parse_si(m.group(2)),
                previous_year=parse_si(m.group(3)) if m.group(3) else None)
            seen.add(aop)
    return result

def parse_jolp_format(pages_text, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="JOLP")
    full = "\n".join(pages_text)
    lines0 = pages_text[0].split('\n') if pages_text else []
    c = CompanyInfo()
    if lines0: c.name = lines0[0].strip()
    m = re.search(r'Mati[cč]na [sš]tevilka:\s*(\d{10})', full)
    if m: c.registration_number = m.group(1).strip()
    m = re.search(r'na dan\s+([\d]{1,2}\.[\d]{1,2}\.[\d]{4})', full)
    if m: c.period_to = m.group(1).strip()
    m = re.search(r'od\s+([\d]{1,2}\.[01]?\d\.[\d]{4})', full)
    if m: c.period_from = m.group(1).strip()
    if (not c.period_from or len(c.period_from) < 6) and c.period_to:

        c.period_from = _infer_period_from(c.period_to)
    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    all_lines = []
    for pt in pages_text:
        all_lines.extend(pt.split('\n'))
    NUM_RE = re.compile(r'^(.+?)\s+([\d.]+,\d{2})\s+([\d.]+,\d{2})\s*$')
    joined = []
    for line in all_lines:
        line = line.strip()
        if not line or line.startswith('Stran') or line in ('2025 2024','Podatki so v EUR s centi'): continue
        if NUM_RE.match(line):
            joined.append(line)
        else:
            if joined and not NUM_RE.match(joined[-1]):
                joined[-1] = joined[-1] + " " + line
            else:
                joined.append(line)
    for line in joined:
        m = NUM_RE.match(line)
        if not m: continue
        label = m.group(1).strip()
        aop = JOLP_TO_AOP.get(label)
        if not aop:
            for key, code in JOLP_TO_AOP.items():
                if label.startswith(key) or key.startswith(label[:50]):
                    aop = code; break
        if aop and aop not in result.aop_data:
            result.aop_data[aop] = AopEntry(aop=aop,
                current_year=parse_si(m.group(2)),
                previous_year=parse_si(m.group(3)))
    return result

def parse_napoved_format(pages_text, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="NAPOVED")
    full = "\n".join(pages_text)
    lines = pages_text[0].split('\n') if pages_text else []
    c = CompanyInfo()
    # Ime in matična sta v prvi vrstici skupaj
    if lines:
        m = re.search(r'^(.+?)\s+Matična številka:\s*(\d+)', lines[0])
        if m:
            c.name = m.group(1).strip()
            reg = m.group(2).strip()
            # NAPOVED format ima 7-mestno matično — dodaj 000 do 10 mest
            c.registration_number = reg.ljust(10, '0') if len(reg) < 10 else reg
    m = re.search(r'Davčna številka:\s*(\d+)', full)
    if m: c.tax_number = m.group(1).strip()
    m = re.search(r'na dan\s+([\d]{1,2}\.[\d]{1,2}\.[\d]{4})', full)
    if m: c.period_to = m.group(1).strip()
    # IPI ima 'v obdobju od DD.MM.YYYY do DD.MM.YYYY'
    m = re.search(r'do\s+([\d]{1,2}\.[\d]{1,2}\.[\d]{4})', full)
    if m and not c.period_to: c.period_to = m.group(1).strip()
    m = re.search(r'od\s+([\d]{1,2}\.[01]?\d\.[\d]{4})\s+do', full)
    if m: c.period_from = m.group(1).strip()
    if (not c.period_from or len(c.period_from) < 6) and c.period_to:

        c.period_from = _infer_period_from(c.period_to)
    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    seen = set()
    for page_text in pages_text:
        for line in page_text.split('\n'):
            m = NAPOVED_RE.match(line.strip())
            if not m: continue
            aop = m.group(1)
            if not (1 <= int(aop) <= 310): continue
            if aop in seen: continue
            result.aop_data[aop] = AopEntry(aop=aop,
                current_year=parse_si(m.group(2)),
                previous_year=parse_si(m.group(3)))
            seen.add(aop)
    return result

def parse_upr_format(pages_text, pages_words, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="UPR")
    full = "\n".join(pages_text)
    first = pages_text[0] if pages_text else ""
    c = CompanyInfo()

    m = re.search(r'(?:BILANCA STANJA|IZKAZ POSLOVNEGA IZIDA)\s+družbe\s+(.+)', first)
    if m: c.name = m.group(1).strip()

    is_ipi = "IZKAZ POSLOVNEGA IZIDA" in first
    if is_ipi:
        m = re.search(r'od\s+(\d{1,2}\.\d{1,2}\.\d{4})\s+do\s+(\d{1,2}\.\d{1,2}\.\d{4})\s+in\s+od', full)
        if m:
            c.period_from = m.group(1).strip()
            c.period_to   = m.group(2).strip()
        else:
            m2 = re.search(r'do\s+(\d{1,2}\.\d{1,2}\.\d{4})', full)
            if m2: c.period_to = m2.group(1).strip()
    else:
        m = re.search(r'na dan\s+(\d{1,2}\.\d{1,2}\.\d{4})', full)
        if m: c.period_to = m.group(1).strip()

    if (not c.period_from or len(c.period_from) < 6) and c.period_to:
        # Mesečna interna poročila so vedno na standardno koledarsko leto (1.1.-31.12.),
        # zato tu NE uporabimo splošne _infer_period_from (ta predvideva nestandardno
        # poslovno leto za AJPES letna poročila, kar tu ni relevantno).
        leto = c.period_to.split('.')[-1]
        c.period_from = f"1.1.{leto}"

    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    if not result.company.registration_number:
        result.warnings.append(
            "Ta format (interno mesečno poročilo) ne vsebuje matične številke — vpiši jo ročno v celico D4."
        )

    boundaries = UPR_IPI_BOUNDARIES if is_ipi else UPR_BS_BOUNDARIES
    context = None
    for words in pages_words:
        for label, numgroups in _upr_reconstruct_rows(words):
            aop = _match_aop_label_exact(label)
            if not aop:
                # Postavke z enakim besedilom v različnih sklopih (npr. dolgoročne/kratkoročne
                # naložbe) razrešimo glede na kontekst zadnje prebrane nadrejene postavke.
                _, item_text = _classify_and_strip_item(label)
                opts = _AMBIGUOUS_ITEMS.get(item_text)
                if opts and context in opts:
                    aop = opts[context]
                else:
                    # Zadnji poskus: vrstica se je morda "zlepila" s sosednjo (prazno) postavko.
                    aop = _match_aop_label_merged(label)
            if aop and aop in _CONTEXT_TRIGGERS:
                context = _CONTEXT_TRIGGERS[aop]
            if not aop or aop in result.aop_data:
                continue
            # Varovalka: BS in IPI imata ločena obmocja AOP kod (BS: 001-096/301, IPI: 110-189+).
            # Nekatere postavke imajo v BS in IPI zelo podobno besedilo (npr. 'Čista izguba
            # poslovnega leta' v BS proti 'ČISTA IZGUBA POSLOVNEGA LETA' v IPI), zato brez te
            # varovalke lahko pride do napačnega križnega ujemanja med dokumentoma.
            try:
                aop_num = int(re.sub(r'\D', '', aop) or 0)
            except Exception:
                aop_num = 0
            if is_ipi and aop_num < 110:
                continue
            if not is_ipi and aop_num >= 110:
                continue
            cols = _upr_bucket_columns(numgroups, boundaries)
            if is_ipi:
                # stolpec 0 = samo mesec (ne uporabimo), 1 = kumulativa tekoče leto, 2 = kumulativa primerjalno leto
                current  = _upr_parse_number(cols[1]) if len(cols) > 1 else None
                previous = _upr_parse_number(cols[2]) if len(cols) > 2 else None
            else:
                current  = _upr_parse_number(cols[0]) if len(cols) > 0 else None
                previous = _upr_parse_number(cols[1]) if len(cols) > 1 else None
            if current is not None or previous is not None:
                result.aop_data[aop] = AopEntry(aop=aop, current_year=current, previous_year=previous)
    return result


# ── PARSER: PLAIN_MST format (npr. "AVTO ELITE D.O.O." izvoz) ─────────────────
# Glava: "<IME PODJETJA> Matična številka: NNNNNNN" v prvi vrstici, naslov gol
# "BILANCA STANJA" / "IZKAZ POSLOVNEGA IZIDA" (brez "družbe"), postavke brez AOP kod.
# Zneski so — tako kot pri UPR — bran po x/y pozicijah besed, ne po linearnem besedilu,
# ker pdfplumber pri teh PDF-jih včasih razcepi/premeša fragmente zneskov ali glave.
_PLAINMST_VALUE_X_MIN_BS  = 260   # ločnica opis/zneski za BS (postavke segajo do x0≈220)
_PLAINMST_BS_BOUNDARIES   = [365] # BS: 1 ločnica med stolpcema (tekoče ~300-350, primerjalno ~375-425)
_PLAINMST_VALUE_X_MIN_IPI = 340   # ločnica opis/zneski za IPI (postavke segajo do x0≈200)
_PLAINMST_IPI_BOUNDARIES  = [430] # IPI: 1 ločnica med stolpcema (tekoče ~390-420, primerjalno ~460-490)

def _plainmst_is_valueish(text):
    t = text.strip()
    # Datumi v glavi tabele (npr. '31.07.2026', '01.01.2026') imajo isto obliko kot
    # znesek s pikami, zato jih tu izrecno izključimo, da se ne zamenjajo za postavko.
    if re.fullmatch(r'\d{1,2}\.\d{1,2}\.\d{4}', t):
        return False
    return bool(re.fullmatch(r'\(?-?[\d.,]+\)?', t)) and any(ch.isdigit() for ch in t)

def _plainmst_reconstruct_rows(words, value_x_min):
    from collections import defaultdict as _dd
    rows = _dd(list)
    for w in words:
        rows[round(w['top'], 1)].append(w)
    row_tops = sorted(rows.keys())
    value_rows, label_rows = [], []
    for top in row_tops:
        ws = sorted(rows[top], key=lambda w: w['x0'])
        val_ws = [w for w in ws if w['x0'] >= value_x_min and _plainmst_is_valueish(w['text'])]
        lbl_ws = [w for w in ws if w['x0'] < value_x_min]
        # 'POSTAVKA' je samo naslov stolpca z opisi (tabelska glava), ne del nobene
        # dejanske postavke - brez tega izloči bi se zlepil s prvo pravo vrstico (SREDSTVA).
        if len(lbl_ws) == 1 and lbl_ws[0]['text'].strip().upper() == 'POSTAVKA':
            lbl_ws = []
        if val_ws: value_rows.append((top, val_ws))
        if lbl_ws: label_rows.append((top, lbl_ws))
    value_tops = [t for t, _ in value_rows]
    entries = []
    for i, (top, val_ws) in enumerate(value_rows):
        lower = (value_tops[i-1] + top) / 2 if i > 0 else -1e9
        upper = (value_tops[i+1] + top) / 2 if i < len(value_tops) - 1 else 1e9
        lbls = sorted([(lt, lw) for (lt, lw) in label_rows if lower < lt <= upper], key=lambda x: x[0])
        label_text = ' '.join(' '.join(w['text'] for w in lw) for _, lw in lbls).strip()
        # Zadnja postavka na strani se lahko zlepi s podpisnim blokom na dnu obrazca
        # (nima lastne "vrstice z zneski", zato se pripiše zadnji pravi postavki).
        label_text = re.sub(
            r'\s*(Oseba odgovorna za sestavljanje bilance:|Vodja družbe:).*$',
            '', label_text, flags=re.I).strip()
        entries.append((label_text, _upr_cluster_numbers(val_ws)))
    return entries

def parse_plain_mst_format(pages_text, pages_words, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="PLAIN_MST")
    full = "\n".join(pages_text)
    first = pages_text[0] if pages_text else ""
    is_ipi = "IZKAZ POSLOVNEGA IZIDA" in first
    c = CompanyInfo()

    first_line = first.split('\n')[0] if first else ""
    m = re.search(r'^(.+?)\s+Matična številka:\s*(\d{6,8})', first_line)
    if m:
        c.name = m.group(1).strip()
        c.registration_number = m.group(2).strip()
    m = re.search(r'Davčna številka:\s*(\d+)', full)
    if m: c.tax_number = m.group(1).strip()

    if is_ipi:
        m = re.search(r'za obdobje od\s+(\d{1,2}\.\d{1,2}\.\d{4})\s+do\s+(\d{1,2}\.\d{1,2}\.\d{4})', full)
        if m:
            c.period_from = m.group(1).strip()
            c.period_to = m.group(2).strip()
    else:
        m = re.search(r'na dan\s+(\d{1,2}\.\d{1,2}\.\d{4})', full)
        if m: c.period_to = m.group(1).strip()

    if (not c.period_from or len(c.period_from) < 6) and c.period_to:
        leto = c.period_to.split('.')[-1]
        c.period_from = f"1.1.{leto}"

    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)

    boundaries   = _PLAINMST_IPI_BOUNDARIES if is_ipi else _PLAINMST_BS_BOUNDARIES
    value_x_min  = _PLAINMST_VALUE_X_MIN_IPI if is_ipi else _PLAINMST_VALUE_X_MIN_BS
    context = None
    for words in pages_words:
        for label, numgroups in _plainmst_reconstruct_rows(words, value_x_min):
            aop = _match_aop_label_exact(label)
            if not aop:
                _, item_text = _classify_and_strip_item(label)
                opts = _AMBIGUOUS_ITEMS.get(item_text)
                if opts and context in opts:
                    aop = opts[context]
                else:
                    aop = _match_aop_label_merged(label)
            if aop and aop in _CONTEXT_TRIGGERS:
                context = _CONTEXT_TRIGGERS[aop]
            if not aop or aop in result.aop_data:
                continue
            try:
                aop_num = int(re.sub(r'\D', '', aop) or 0)
            except Exception:
                aop_num = 0
            if is_ipi and aop_num < 110:
                continue
            if not is_ipi and aop_num >= 110:
                continue
            cols = _upr_bucket_columns(numgroups, boundaries)
            current  = _upr_parse_number(cols[0]) if len(cols) > 0 else None
            previous = _upr_parse_number(cols[1]) if len(cols) > 1 else None
            if current is not None or previous is not None:
                result.aop_data[aop] = AopEntry(aop=aop, current_year=current, previous_year=previous)
    return result


# ── PARSER: BESEDILO_2COL (interni BS/IPI, npr. ŠUŠTAR TRANS) ────────────────
#
# Ta družina PDF-jev nima AOP kod. Ima pa stabilno semantično strukturo in dva
# desno poravnana stolpca zneskov (tekoče + primerjalno obdobje). Ker se del
# postavk poimenovno razlikuje od uradnega AJPES obrazca (npr. "Stroški materiala"
# pomeni skupni AOP 130, ne podpostavke AOP 131), uporabljamo namensko preslikavo.

def _b2c_norm(text):
    s = re.sub(r'\s+', ' ', str(text or '')).strip()
    # Odreži podpisni blok, ki se na zadnji vrstici strani lahko zlepi z zadnjo postavko.
    s = re.split(r'\bSestavil\s*:', s, maxsplit=1, flags=re.I)[0].strip()
    # Številka strani se pri nekaterih izvozih zlepi z zadnjo postavko ("... zaloge 1").
    s = re.sub(r'(?<=\D)\s+[1-9]\s*$', '', s).strip()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower()

def _b2c_money_boundary(pages_words):
    """Poišče ločnico med tekočim in primerjalnim stolpcem iz x-koordinat zneskov.

    Zneski so desno poravnani, zato imajo začetki prvega stolpca praviloma x≈370–420,
    drugega pa x≈450–505. Največjo vrzel med opazovanimi začetki uporabimo kot ločnico.
    """
    xs = []
    money_re = re.compile(r'^-?\d[\d.]*,\d{2}$')
    for words in pages_words:
        for w in words:
            t = str(w.get('text', '')).strip().replace('(', '').replace(')', '')
            if w.get('x0', 0) >= 320 and money_re.fullmatch(t):
                xs.append(float(w['x0']))
    if len(xs) < 4:
        return 445.0
    vals = sorted(set(round(x, 1) for x in xs))
    gaps = [(vals[i+1] - vals[i], vals[i], vals[i+1]) for i in range(len(vals)-1)]
    gap, left, right = max(gaps, key=lambda x: x[0])
    if gap >= 18:
        return (left + right) / 2.0
    return 445.0

def _b2c_put(result, code, current, previous, add=False):
    """Vpiše AOP. ``add=True`` se uporablja, ko več izvornih vrstic pripada istemu CBK AOP."""
    if not code or (current is None and previous is None):
        return
    if add and code in result.aop_data:
        old = result.aop_data[code]
        cur = ((old.current_year or 0.0) + (current or 0.0)) if (old.current_year is not None or current is not None) else None
        prev = ((old.previous_year or 0.0) + (previous or 0.0)) if (old.previous_year is not None or previous is not None) else None
        result.aop_data[code] = AopEntry(aop=code, current_year=cur, previous_year=prev)
    elif code not in result.aop_data:
        result.aop_data[code] = AopEntry(aop=code, current_year=current, previous_year=previous)

def _b2c_derive(result, code, parts, signs=None):
    """Izpelje AOP iz že prebranih postavk, vendar samo če ga izvor ni podal neposredno."""
    if code in result.aop_data:
        return
    signs = signs or [1] * len(parts)
    entries = [result.aop_data.get(p) for p in parts]
    if not any(e is not None for e in entries):
        return
    def calc(attr):
        have = False; total = 0.0
        for e, sg in zip(entries, signs):
            if e is not None and getattr(e, attr) is not None:
                have = True; total += sg * getattr(e, attr)
        return total if have else None
    result.aop_data[code] = AopEntry(code, calc('current_year'), calc('previous_year'))

def _b2c_bs_map(label_norm, context):
    n = label_norm

    # Kontrolna seštevka najprej, ker vsebujeta besedilo drugih naslovov.
    if 'sredstva skupaj' in n: return '001', False, context
    if 'skupaj obveznosti do virov sredstev' in n: return '055', False, context

    # Aktiva
    if 'a. dolgorocna sredstva' in n: return '002', False, 'assets'
    if 'neopredmetena sredstva in dolgorocne aktivne casovne' in n: return '003', False, context
    if 'dolgorocne premozenjske pravice' in n: return '005', False, context
    if re.search(r'\bdobro ime\b', n): return '006', False, context
    if 'predujmi za neopredmetena osnovna sredstva' in n: return '008', True, context
    if 'dolgorocno odlozeni stroski' in n: return '007', False, context
    if 'druga neopredmetena osnovna sredstva' in n: return '008', True, context
    if 'dolgorocne aktivne casovne razmejitve' in n: return '009', False, context
    if 'ii. opredmetena osnovna sredstva' in n: return '010', False, context
    if re.search(r'\ba\)\s*zemljisc', n): return '011', False, context
    if re.search(r'\bb\)\s*zgradbe\b', n): return '012', False, context
    if 'proizvajalna oprema' in n: return '013', False, context
    if re.search(r'\bdruga oprema\b', n): return '014', False, context
    if 'v gradnji oz. izdelavi' in n or 'v gradnji oziroma izdelavi' in n: return '016', False, context
    if 'predujmi za opredmetena osnovna sredstva' in n: return '017', False, context
    if 'iii. nalozbene nepremicnine' in n: return '018', False, context
    if 'iv. dolgorocne financne nalozbe' in n: return '019', False, 'lt_fin_inv'
    if '1. dolgorocne financne nalozbe, razen posojil' in n: return '020', False, context
    if 'delnice in delezi v druzbah v skupini' in n:
        return ('021' if context == 'lt_fin_inv' else '042'), False, context
    if 'delnice in delezi v pridruzene druzbe' in n or 'druge delnice in delezi' in n:
        return ('022' if context == 'lt_fin_inv' else '043'), True, context
    if '2. dolgorocno dana posojila' in n: return '024', False, context
    if 'dolgorocno dana posojila druzbam v skupini' in n: return '025', False, context
    if 'dolgorocno dana posojila in depoziti drugim' in n or 'dolgorocno nevplacani vpoklicani kapital' in n:
        return '026', True, context
    if 'v. dolgorocne poslovne terjatve' in n: return '027', False, 'lt_receivables'
    if 'dolgorocne poslovne terjatve do druzb v skupini' in n: return '028', False, context
    if 'dolgorocne poslovne terjatve do kupcev' in n: return '029', False, context
    if 'druge dolgorocne poslovne terjatve' in n: return '030', False, context
    if 'vi. odlozene terjatve za davke' in n: return '031', False, context
    if 'b. kratkorocna sredstva' in n: return '032', False, 'short_assets'
    if 'i. sredstva za prodajo' in n: return '033', False, context
    if 'ii. zaloge' in n: return '034', False, context
    if 'material in surovine' in n: return '035', False, context
    if 'nedokoncana proizvodnja' in n: return '036', False, context
    if 'gotovi proizvodi' in n: return '037', False, context
    if 'trgovsko blago' in n: return '038', False, context
    if 'predujmi za zaloge' in n: return '039', False, context
    if 'iii. kratkorocne financne nalozbe' in n: return '040', False, 'st_fin_inv'
    if '1. kratkorocne financ' in n and 'nalozbe' in n: return '041', False, context
    if '2. kratkorocna posojila' in n: return '045', False, context
    if 'kratkorocno dana posojila druzbam v skupini' in n: return '046', False, context
    if 'kratkorocno dana posojila in depoziti drugim' in n or 'kratkorocno nevplacani vpoklicani kapital' in n:
        return '047', True, context
    if 'iv. kratkorocne poslovne terjatve' in n: return '048', False, 'st_receivables'
    if 'kratkorocne poslovne terjatve do druzb v skupini' in n: return '049', False, context
    if 'kratkorocne poslovne terjatve do kupcev' in n: return '050', False, context
    if 'kratkorocne poslovne terjatve do drugih' in n: return '051', False, context
    if 'v. denarna sredstva' in n: return '052', False, context
    if 'c. kratkorocne aktivne casovne razmejitve' in n: return '053', False, context

    # Pasiva
    if 'a. kapital' in n: return '056', False, 'capital'
    if 'i. vpoklicani kapital' in n: return '057', False, context
    if '1. osnovni kapital' in n: return '058', False, context
    if '2. nevpoklicani kapital' in n: return '059', False, context
    if 'ii. kapitalske rezerve' in n: return '060', False, context
    if 'iii. rezerve iz dobicka' in n: return '061', False, context
    if '1. zakonske rezerve' in n: return '062', False, context
    if '2. rezerve za lastne delnice' in n: return '063', False, context
    if '3. rezerve za lastne delnice' in n and 'odbitna postavka' in n: return '064', False, context
    if '4. statutarne rezerve' in n: return '065', False, context
    if '5. druge rezerve iz dobicka' in n: return '066', False, context
    if 'iv. revalorizacijske rezerve' in n: return '067', False, context
    if 'v. rezerve nastale zaradi vrednotenja po posteni vrednosti' in n: return '301', False, context
    if 'vi. preneseni cisti dobicki' in n: return '068', False, context
    if 'vii. prenesena cista izguba' in n: return '069', False, context
    if 'viii. cisti dobicek poslovnega leta' in n: return '070', False, context
    if 'ix. cista izguba poslovnega leta' in n: return '071', False, context
    if 'rezervacije in dolgorocne pasivne casovne' in n and ('b.' in n or n.startswith('b ')):
        return '072', False, 'provisions'
    if 'rezervacije za pokojnine in podobne obveznosti' in n: return '073', True, context
    if '2. druge rezervacije' in n: return '073', True, context
    if 'dolgorocne pasivne casovne razmejitve' in n: return '074', False, context
    if 'c. dolgorocne obveznosti' in n: return '075', False, 'lt_liab'
    if 'i. dolgorocne financne obveznosti' in n: return '076', False, 'lt_fin_liab'
    if 'dolgorocne financne obveznosti do druzb v skupini' in n: return '077', False, context
    if 'dolgorocne financne obveznosti do bank' in n: return '078', False, context
    if 'dolgorocne obveznosti na podlagi obveznic' in n or 'druge dolgorocne financne obveznosti' in n:
        return '079', True, context
    if 'ii. dolgorocne poslovne obveznosti' in n: return '080', False, 'lt_trade_liab'
    if 'dolgorocne poslovne obveznosti do druzb v skupini' in n: return '081', False, context
    if 'dolgorocne poslovne obveznosti do dobaviteljev' in n: return '082', False, context
    if 'dolgorocne menicne obveznosti' in n or 'dolgorocne poslovne obveznosti na podlagi predujmov' in n or 'druge dolgorocne poslovne obveznosti' in n:
        return '083', True, context
    if 'iii. odlozene obveznosti za davke' in n: return '084', False, context
    if 'c. kratkorocne obveznosti' in n or 'č. kratkorocne obveznosti' in n: return '085', False, 'st_liab'
    if 'i. obveznosti, vkljucene v skupine za odtujitev' in n: return '086', False, context
    if 'ii. kratkorocne financne obveznosti' in n: return '087', False, 'st_fin_liab'
    if 'kratkorocne financne obveznosti do druzb v skupini' in n: return '088', False, context
    if 'kratkorocne financne obveznosti do bank' in n: return '089', False, context
    if 'kratkorocne obveznosti na podlagi obveznic' in n or 'druge kratkorocne financne obveznosti' in n:
        return '090', True, context
    if 'iii. kratkorocne poslovne obveznosti' in n: return '091', False, 'st_trade_liab'
    if 'kratkorocne poslovne obveznosti do druzb v skupini' in n: return '092', False, context
    if 'kratkorocne poslovne obveznosti do dobaviteljev' in n: return '093', False, context
    if 'kratkorocne menicne obveznosti' in n or 'kratkorocne poslovne obveznosti na podlagi predujmov' in n or 'druge kratkorocne poslovne obveznosti' in n:
        return '094', True, context
    if 'd. kratkorocne pasivne casovne razmejitve' in n: return '095', False, context
    return None, False, context

def _b2c_ipi_map(label_norm):
    n = label_norm
    # Prihodki od prodaje - najprej podpostavke, nato seštevek.
    if 'cisti prihodki od prodaje v sloveniji' in n: return '111'
    if 'cisti prihodki od prodaje v eu' in n: return '115'
    if 'cisti prihodki od prodaje v tujini' in n: return '118'
    if 'a cisti prihodki od prodaje' in n or 'a. cisti prihodki od prodaje' in n: return '110'
    if 'povecanje vrednosti zalog proizvodov' in n: return '121'
    if 'zmanjsanje vrednosti zalog proizvodov' in n or 'znamjsanje vrednosti zalog proizvodov' in n: return '122'
    if 'usredstveni lastni proizvodi in lastne storitve' in n: return '123'
    if 'drugi poslovni prihodki' in n: return '125'
    if re.search(r'\be\s+poslovni odhodki\b', n) or 'e. poslovni odhodki' in n: return '127'
    if 'stroski blaga, materiala in storitev' in n: return '128'
    if 'nabavna vrednost prodanega blaga' in n: return '129'
    if re.search(r'\bstroski materiala\b', n): return '130'
    if re.search(r'\bstroski storitev\b', n): return '134'
    if re.search(r'\b2\. stroski dela\b', n): return '139'
    if 'stroski plac' in n: return '140'
    if 'stroski pokojninskih zavarovanj' in n: return '141'
    if 'stroski drugih socialnih zavarovanj' in n: return '142'
    if 'drugi stroski dela' in n: return '143'
    if '3. odpisi vrednosti' in n: return '144'
    if 'amortizacija' in n: return '145'
    if 'prevrednotovalni poslovni odhodki pri neopredmetenih' in n: return '146'
    if 'prevrednotovalni poslovni odhodki pri obratnih sredstvih' in n: return '147'
    if '4. drugi poslovni odhodki' in n: return '148'
    if 'f dobicek iz poslovanja' in n or 'f. dobicek iz poslovanja' in n: return '151'
    if 'g izguba iz poslovanja' in n or 'g. izguba iz poslovanja' in n: return '152'
    if 'h financni prihodki' in n or 'h. financni prihodki' in n: return '153'
    if '1. financni prihodki iz delezev' in n: return '155'
    if re.search(r'\ba\)\s*financni prihodki iz delezev v druzbah v skupini', n): return '156'
    if re.search(r'\bb\)\s*financni prihodki iz delezev v druzbah v skupini', n) or 'financni prihodki iz delezev v pridruzenih druzbah' in n: return '157'
    if 'financni prihodki iz delezev v drugih druzbah' in n: return '158'
    if 'financni prihodki drugih nalozb' in n: return '159'
    if '2. financni prihodki iz danih posojil' in n: return '160'
    if 'financni prihodki iz danih posojil druzbam v skupini' in n: return '161'
    if 'financni prihodki iz danih posojil drugim' in n: return '162'
    if '3. financni prihodki iz poslovnih terjatev' in n: return '163'
    if 'financni prihodki iz poslovnih terjatev do druzb v skupini' in n: return '164'
    if 'financni prihodki iz poslovnih terjatev do drugih' in n: return '165'
    if 'i financni odhodki' in n or 'i. financni odhodki' in n: return '166'
    if 'financni odhodki iz oslabitve in odpisov financnih nalozb' in n: return '168'
    if '2. financni odhodki iz financnih obveznosti' in n: return '169'
    if 'financni odhodki iz posojil, prejetih od druzb v skupini' in n: return '170'
    if 'financni odhodki iz posojil, prejetih od bank' in n: return '171'
    if 'financni odhodki iz izdanih obveznic' in n: return '172'
    if 'financni odhodki iz drugih financnih obveznosti' in n: return '173'
    if '3. financni odhodki iz poslovnih obveznosti' in n: return '174'
    if 'financni odhodki iz poslovnih obveznosti do druzb v skupini' in n: return '175'
    if 'financni odhodki iz poslovnih obveznosti do dobaviteljev' in n: return '176'
    if 'financni odhodki iz drugih poslovnih obveznosti' in n: return '177'
    if 'j drugi prihodki' in n or 'j. drugi prihodki' in n: return '178'
    if 'k drugi odhodki' in n or 'k. drugi odhodki' in n: return '181'
    if 'l celotni dobicek' in n or 'l. celotni dobicek' in n: return '182'
    if 'm celotna izguba' in n or 'm. celotna izguba' in n: return '183'
    if 'n davek iz dobicka' in n or 'n. davek iz dobicka' in n: return '184'
    if 'o odlozeni davki' in n or 'o. odlozeni davki' in n: return '185'
    if 'p cisti dobicek poslovnega obdobja' in n or 'p. cisti dobicek poslovnega obdobja' in n: return '186'
    if 'r cista izguba poslovnega obdobja' in n or 'r. cista izguba poslovnega obdobja' in n: return '187'
    return None

def parse_besedilo_2col_format(pages_text, pages_words, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="BESEDILO_2COL")
    full = "\n".join(pages_text)
    first = pages_text[0] if pages_text else ""
    is_ipi = bool(re.search(r'IZKAZ\s+POSLOVNEGA\s+IZIDA', first, re.I))

    c = CompanyInfo()
    # Ime je prva neprazna vrstica; v tej družini izvoza naslov/ID-ji sledijo šele za njim.
    for ln in first.split('\n'):
        ln = ln.strip()
        if ln:
            c.name = ln
            break
    m = re.search(r'dav[cč]na\s+[sš]tevilka\s*:\s*(\d{8})\b', full, re.I)
    if m: c.tax_number = m.group(1)
    m = re.search(r'mati[cč]na\s+[sš]tevilka\s*:\s*(\d{7,10})\b', full, re.I)
    if m: c.registration_number = m.group(1)

    if is_ipi:
        m = re.search(r'za\s+obdobje\s+od\s+(\d{1,2}\.\d{1,2}\.\d{4})\s+do\s+(\d{1,2}\.\d{1,2}\.\d{4})', full, re.I)
        if m:
            c.period_from, c.period_to = m.group(1), m.group(2)
    else:
        m = re.search(r'BILANCA\s+STANJA\s+na\s+dan\s+(\d{1,2}\.\d{1,2}\.\d{4})', re.sub(r'\s+', ' ', full), re.I)
        if not m:
            m = re.search(r'na\s+dan\s+(\d{1,2}\.\d{1,2}\.\d{4})', full, re.I)
        if m:
            c.period_to = m.group(1)
            # Ta izvoz je medletni izkaz koledarskega poslovnega leta (IPI istega paketa
            # uporablja 01.01.YYYY -> datum bilance), zato ne uporabljamo splošnega
            # pravila za nestandardna poslovna leta iz finalize_company().
            c.period_from = f"1.1.{c.period_to.split('.')[-1]}"

    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)

    boundary = _b2c_money_boundary(pages_words)
    value_x_min = max(320.0, boundary - 100.0)
    context = None

    for words in pages_words:
        for label, numgroups in _plainmst_reconstruct_rows(words, value_x_min):
            # Tabelske glave vsebujejo fragmente datumov ("1.1.") - niso zneski/postavke.
            if not numgroups or all(re.fullmatch(r'\d{1,2}\.\d{1,2}\.', str(t).strip()) for t, _ in numgroups):
                continue
            cols = _upr_bucket_columns(numgroups, [boundary])
            current = _upr_parse_number(cols[0]) if len(cols) > 0 else None
            previous = _upr_parse_number(cols[1]) if len(cols) > 1 else None
            n = _b2c_norm(label)
            if not n:
                continue

            if is_ipi:
                # V CBK je neto sprememba zalog kombinirana GVAOP vrstica brez lastne AOP kode.
                if ('b spremembe vrednosti zalog' in n or 'b. spremembe vrednosti zalog' in n) \
                        and current is not None:
                    result.gvaop_data['051'] = current
                code = _b2c_ipi_map(n)
                if code:
                    _b2c_put(result, code, current, previous)
            else:
                code, add, context = _b2c_bs_map(n, context)
                if code:
                    _b2c_put(result, code, current, previous, add=add)

    if is_ipi:
        # V tem izvoru F. KOSMATI DONOS ni izpisan, CBK pa ga ima kot samostojen AOP 126.
        # Izpeljava je uradna struktura izkaza: prodaja + sprememba zalog + usredstveni
        # proizvodi + poslovni prihodki.
        _b2c_derive(result, '126', ['110','121','122','123','124','125'], [1,1,-1,1,1,1])
    else:
        # Izvor razbije neopredmetena sredstva drugače od CBK. AOP 004 je zato skupni
        # del AOP 003 brez dolgoročnih aktivnih časovnih razmejitev (AOP 009).
        _b2c_derive(result, '004', ['003','009'], [1,-1])
        # Varnostni izpeljavi za kontrolna seštevka, če ju kakšna različica izvoza izpusti.
        _b2c_derive(result, '001', ['002','032','053'])
        _b2c_derive(result, '055', ['056','072','075','085','095'])

    return result

def _cell_to_text(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

def _cell_to_number(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        return parse_si(v)
    return None

def _aop_from_cell(v):
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if float(v).is_integer():
            n = int(v)
            if 100 <= n <= 499:   # prepreči headerje 1,2,3,4,5
                return str(n).zfill(3)
        return None
    s = str(v).strip()
    m = re.fullmatch(r'(\d{3})([a-z])?', s, flags=re.I)
    if not m:
        return None
    n = int(m.group(1))
    if not (1 <= n <= 499):
        return None
    return m.group(1).zfill(3) + (m.group(2).lower() if m.group(2) else "")

def _find_value_by_label(rows_by_sheet, label_patterns):
    patterns = [re.compile(p, re.I) for p in label_patterns]
    for rows in rows_by_sheet.values():
        for r_idx, row in enumerate(rows):
            for c_idx, cell in enumerate(row):
                txt = _cell_to_text(cell)
                if not txt:
                    continue
                if any(p.search(txt) for p in patterns):
                    # 1) desno v isti vrstici
                    for j in range(c_idx + 1, min(len(row), c_idx + 5)):
                        val = _cell_to_text(row[j])
                        if val:
                            return val
                    # 2) spodaj v istem stolpcu ali blizu stolpca
                    for rr in range(r_idx + 1, min(len(rows), r_idx + 5)):
                        for cc in range(max(0, c_idx - 1), min(len(rows[rr]), c_idx + 4)):
                            val = _cell_to_text(rows[rr][cc])
                            if val:
                                return val
    return ""

def _extract_company_from_excel_rows(rows_by_sheet):
    c = CompanyInfo(); c.tip_subjekta = "1"
    c.name = _find_value_by_label(rows_by_sheet, [r'ime poslovnega subjekta', r'ime podjetnika'])
    c.registration_number = re.sub(r'\D', '', _find_value_by_label(rows_by_sheet, [r'mati[cč]na [sš]tevilka']))
    c.tax_number = re.sub(r'\D', '', _find_value_by_label(rows_by_sheet, [r'dav[cč]na [sš]tevilka']))

    # Če label-finder pri AJPES LP2025 vrne header namesto vrednosti, popravi z znanim vzorcem.
    for rows in rows_by_sheet.values():
        for i, row in enumerate(rows[:-1]):
            texts = [_cell_to_text(x) for x in row]
            low = [x.lower() for x in texts]
            if any('ime poslovnega subjekta' in x for x in low) and i + 1 < len(rows):
                candidate = _cell_to_text(rows[i+1][0]) if rows[i+1] else ""
                if candidate and len(candidate) > 2:
                    c.name = candidate
            for j, x in enumerate(low):
                if 'matična številka' in x or 'maticna stevilka' in x:
                    # desno ali spodaj
                    right = ""
                    if j + 1 < len(row):
                        right = re.sub(r'\D', '', _cell_to_text(row[j+1]))
                    down = ""
                    if i + 1 < len(rows) and j < len(rows[i+1]):
                        down = re.sub(r'\D', '', _cell_to_text(rows[i+1][j]))
                    c.registration_number = right or down or c.registration_number
                if 'davčna številka' in x or 'davcna stevilka' in x:
                    right = ""
                    if j + 1 < len(row):
                        right = re.sub(r'\D', '', _cell_to_text(row[j+1]))
                    down = ""
                    if i + 1 < len(rows) and j < len(rows[i+1]):
                        down = re.sub(r'\D', '', _cell_to_text(rows[i+1][j]))
                    c.tax_number = right or down or c.tax_number

    all_text = "\n".join(
        " ".join(_cell_to_text(v) for v in row if _cell_to_text(v))
        for rows in rows_by_sheet.values() for row in rows
    )
    m = re.search(r'na dan\s+([\d]{1,2}\.[\d]{1,2}\.[\d]{4})', all_text, re.I)
    if m: c.period_to = m.group(1)
    m = re.search(r'od\s+([\d]{1,2}\.[\d]{1,2}\.[\d]{4})\s*(?:-|do)\s*([\d]{1,2}\.[\d]{1,2}\.[\d]{4})', all_text, re.I)
    if m:
        c.period_from = m.group(1); c.period_to = m.group(2)
    # SP Excel pogosto vsebuje "v obdobju od 1.1. do 31.12.2025"; period_from se varno izpelje spodaj.
    return c, all_text

def parse_excel_rows(rows_by_sheet, tip_override=None, tip_subjekta_override=None, source_format="XLSX"):
    result = ParseResult(pdf_format=source_format)
    c, all_cell_text = _extract_company_from_excel_rows(rows_by_sheet)
    c = finalize_company(c, tip_override, tip_subjekta_override)
    result.company = c
    result.subject_type = _subject_type_from_tip_subjekta(c.tip_subjekta)

    seen = set()
    for rows in rows_by_sheet.values():
        for row in rows:
            for idx, val in enumerate(row):
                aop = _aop_from_cell(val)
                if not aop or aop in seen:
                    continue
                # prva numerična vrednost desno od AOP = tekoče leto, naslednja = prejšnje leto
                nums = []
                for j in range(idx + 1, min(len(row), idx + 5)):
                    num = _cell_to_number(row[j])
                    if num is not None:
                        nums.append(num)
                if not nums:
                    continue
                result.aop_data[aop] = AopEntry(
                    aop=aop,
                    current_year=nums[0],
                    previous_year=nums[1] if len(nums) > 1 else None
                )
                seen.add(aop)
    return validate(result)

# ── PARSER: "GD in ZADRUGE" zbirni Excel (BS+IPI na enem listu, stolpec C = AOP koda,
# stolpci D-H = ZR preteklih let / TEKOČI POD (medletni) / PLANSKI POD) ──────────────
def _find_gdzadruge_sheet(wb):
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in range(1, 16):
            for col in range(1, 4):
                v = ws.cell(row=row, column=col).value
                if v and "PODATKI IZ BILANCE STANJA IN IZKAZA POSLOVNEGA IZIDA" in str(v):
                    return sn
    return None

def parse_gdzadruge_excel(xlsx_path, tip_override=None, tip_subjekta_override=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = _find_gdzadruge_sheet(wb)
    ws = wb[sheet_name]
    result = ParseResult(pdf_format="GDZADRUGE")

    c = CompanyInfo()
    name = ws["B7"].value
    if name: c.name = str(name).strip()
    maticna = ws["B8"].value
    if maticna:
        c.registration_number = str(int(maticna)) if isinstance(maticna, (int, float)) else str(maticna).strip()

    # "TEKOČI POD" (stolpec G) je medletno/preliminarno obdobje - datum je v G14.
    g14 = ws["G14"].value
    if isinstance(g14, datetime.datetime):
        c.period_to = f"{g14.day}.{g14.month}.{g14.year}"
    if c.period_to:
        leto = c.period_to.split('.')[-1]
        c.period_from = f"1.1.{leto}"

    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    if not result.company.registration_number:
        result.warnings.append(
            "Ta Excel format nima izpolnjene matične številke — vpiši jo ročno v celico D4."
        )
    if not result.company.name:
        result.warnings.append(
            "Ta Excel format nima izpolnjenega naziva podjetja — vpiši ga ročno v celico D3."
        )

    for row in range(1, ws.max_row + 1):
        code_val = ws.cell(row=row, column=3).value  # C = AOP koda (navadno število, npr. 1, 93, 110)
        if code_val is None:
            continue
        try:
            code_num = int(code_val)
        except Exception:
            continue
        if not (1 <= code_num <= 320):
            continue
        code = f"{code_num:03d}"
        if code in result.aop_data:
            continue
        current  = ws.cell(row=row, column=7).value  # G = TEKOČI POD (medletni podatki, ki jih rabimo)
        previous = ws.cell(row=row, column=6).value  # F = zadnje zaključeno leto (za primerjavo)
        cur_num  = current  if isinstance(current,  (int, float)) else (parse_si(current)  if isinstance(current,  str) else None)
        prev_num = previous if isinstance(previous, (int, float)) else (parse_si(previous) if isinstance(previous, str) else None)
        if cur_num is not None or prev_num is not None:
            result.aop_data[code] = AopEntry(aop=code, current_year=cur_num, previous_year=prev_num)
    return validate(result)

# ── PARSER: JOLP Excel izvoz za DRUŠTVA (tip_subjekta=4) ──────────────────────────
# AJPES-jev javni izvoz za društva ima popolnoma drugo strukturo kot gospodarske družbe
# (SKLAD namesto KAPITAL, drugačna/krajša AOP shema, nekatere postavke v CBK templatu
# nimajo lastne AOP kode - samo GVAOP). Ker gre za uraden, standardiziran obrazec, so
# nazivi postavk pri vseh društvih enaki, zato uporabimo neposreden (exact-match) slovar.
DRUSTVO_BS_MAP = {
    "SREDSTVA": ("aop", "001"),
    "A. DOLGOROČNA SREDSTVA": ("aop", "002"),
    "I. Neopredmetena sredstva in dolgoročne aktivne časovne razmejitve": ("aop", "003"),
    "II. Opredmetena osnovna sredstva": ("aop", "010"),
    "III. Naložbene nepremičnine": ("aop", "018"),
    "IV. Dolgoročne finančne naložbe": ("aop", "019"),
    "V. Dolgoročne poslovne terjatve": ("aop", "027"),
    "B. KRATKOROČNA SREDSTVA": ("aop", "032"),
    "I. Sredstva (skupine za odtujitev) za prodajo": ("aop", "033"),
    "II. Zaloge": ("aop", "034"),
    "III. Kratkoročne finančne naložbe": ("aop", "040"),
    "IV. Kratkoročne poslovne terjatve": ("aop", "048"),
    "V. Denarna sredstva": ("aop", "052"),
    "C. KRATKOROČNE AKTIVNE ČASOVNE RAZMEJITVE": ("aop", "053"),
    "Zunajbilančna sredstva": ("aop", "054"),
    "OBVEZNOSTI DO VIROV SREDSTEV": ("aop", "055"),
    "A. SKLAD": ("aop", "056"),
    "I. Društveni sklad": ("gvaop", "720030101"),
    "II. Revalorizacijske rezerve": ("gvaop", "7200310"),
    "III. Rezerve, nastale zaradi vrednotenja po pošteni vrednosti": ("gvaop", "7200311"),
    "B. REZERVACIJE IN DOLGOROČNE PASIVNE ČASOVNE RAZMEJITVE": ("aop", "072"),
    "C. DOLGOROČNE OBVEZNOSTI": ("aop", "075"),
    "I. Dolgoročne finančne obveznosti": ("aop", "076"),
    "II. Dolgoročne poslovne obveznosti": ("aop", "077"),
    "Č. KRATKOROČNE OBVEZNOSTI": ("aop", "085"),
    "I. Obveznosti, vključene v skupine za odtujitev": ("aop", "086"),
    "II. Kratkoročne finančne obveznosti": ("aop", "087"),
    "III. Kratkoročne poslovne obveznosti": ("aop", "091"),
    "D. KRATKOROČNE PASIVNE ČASOVNE RAZMEJITVE": ("aop", "095"),
    "Zunajbilančne obveznosti": ("aop", "096"),
}
# Neposredno ujemljive IPI postavke (brez izračuna)
DRUSTVO_IPI_DIRECT = {
    "1. ČISTI PRIHODKI OD PRODAJE": ("gvaop", "72050"),
    "2. SPREMEMBA VREDNOSTI ZALOG PROIZVODOV IN NEDOKONČANE PROIZVODNJE": ("aop", "113"),
    "3. USREDSTVENI LASTNI PROIZVODI IN LASTNE STORITVE": ("aop", "114"),
    "4. DRUGI POSLOVNI PRIHODKI": ("gvaop", "720910102"),
    "5. Stroški blaga, materiala in storitev": ("gvaop", "72063"),
    "6. Stroški dela": ("gvaop", "72068"),
    "7. Odpisi vrednosti": ("gvaop", "72070"),
    "a) Amortizacija": ("gvaop", "720701001"),
    "b) Prevrednotovalni poslovni odhodki pri neopredmetenih sredstvih in opredmetenih osnovnih sredstvih": ("gvaop", "720701002"),
    "c) Prevrednotovalni poslovni odhodki pri obratnih sredstvih": ("gvaop", "720701003"),
    "8. Drugi poslovni odhodki": ("gvaop", "72071"),
    "15. DRUGI PRIHODKI": ("aop", "178"),
    "16. DRUGI ODHODKI": ("aop", "181"),
    "17. Davek od dohodkov": ("aop", "184"),
    "20. Kritje odhodkov  obravnavanega obračunskega obdobja iz presežka prihodkov iz prejšnjih obračunskih obdobij": ("aop", "186"),
    "21. POVPREČNO ŠTEVILO ZAPOSLENIH NA PODLAGI DELOVNIH UR V OBRAČUNSKEM OBDOBJU": ("aop", "189"),
}

def _find_drustvo_jolp_sheet(wb):
    return "Bilanca stanja" in wb.sheetnames or "Izkaz poslovnega izida" in wb.sheetnames

def _is_drustvo_workbook(wb, tip_subjekta_override=None):
    if str(tip_subjekta_override or "") == "4":
        return _find_drustvo_jolp_sheet(wb)
    if "Bilanca stanja" not in wb.sheetnames:
        return False
    ws = wb["Bilanca stanja"]
    text = "\n".join(str(c) for row in ws.iter_rows(values_only=True) for c in row if c)
    return _detect_subject_type_from_text(text) == "DR"

def parse_jolp_drustvo_excel(xlsx_path, tip_override=None, tip_subjekta_override=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result = ParseResult(pdf_format="JOLP_DRUSTVO")
    c = CompanyInfo()

    if "Bilanca stanja" in wb.sheetnames:
        ws = wb["Bilanca stanja"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[:15]:
            for cell in row:
                if isinstance(cell, str) and cell.strip().startswith("Matična številka:"):
                    c.registration_number = cell.split(":", 1)[1].strip()
            if not c.name:
                v = row[0] if row else None
                if isinstance(v, str) and v.strip() and not v.strip().startswith("Matična številka"):
                    c.name = v.strip()
            v0 = row[0] if row else None
            if isinstance(v0, str) and "na dan" in v0:
                m = re.search(r'(\d{1,2}\.\d{1,2}\.\d{4})', v0)
                if m: c.period_to = m.group(1)

        bs_raw = {}
        for row in rows:
            label = row[0] if row else None
            if not label or not isinstance(label, str):
                continue
            label = label.strip()
            val = row[1] if len(row) > 1 else None
            if isinstance(val, (int, float)):
                bs_raw[label] = val

        for label, (kind, key) in DRUSTVO_BS_MAP.items():
            if label in bs_raw:
                v = bs_raw[label]
                if kind == "aop":
                    result.aop_data[key] = AopEntry(aop=key, current_year=v, previous_year=None)
                else:
                    result.gvaop_data[key] = v

    if "Izkaz poslovnega izida" in wb.sheetnames:
        ws = wb["Izkaz poslovnega izida"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[:15]:
            v0 = row[0] if row else None
            if isinstance(v0, str) and "Izkaz poslovnega izida" in v0:
                m = re.search(r'do\s+(\d{1,2}\.\d{1,2}\.\d{4})', v0)
                if m: c.period_to = m.group(1)

        ipi_raw = {}
        for row in rows:
            label = row[0] if row else None
            if not label or not isinstance(label, str):
                continue
            label = label.strip()
            val = row[1] if len(row) > 1 else None
            if isinstance(val, (int, float)):
                ipi_raw[label] = val

        def g(lbl): return ipi_raw.get(lbl, 0) or 0

        for lbl, (kind, key) in DRUSTVO_IPI_DIRECT.items():
            if lbl in ipi_raw:
                if kind == "aop":
                    result.aop_data[key] = AopEntry(aop=key, current_year=ipi_raw[lbl], previous_year=None)
                else:
                    result.gvaop_data[key] = ipi_raw[lbl]

        if ipi_raw:
            # I. KOSMATI DONOS OD POSLOVANJA (110) = vsota postavk 1-4, ki JOLP izvoz ne izračuna eksplicitno
            prihodki = (g("1. ČISTI PRIHODKI OD PRODAJE") + g("2. SPREMEMBA VREDNOSTI ZALOG PROIZVODOV IN NEDOKONČANE PROIZVODNJE")
                        + g("3. USREDSTVENI LASTNI PROIZVODI IN LASTNE STORITVE") + g("4. DRUGI POSLOVNI PRIHODKI"))
            result.aop_data["110"] = AopEntry(aop="110", current_year=prihodki, previous_year=None)

            # II. POSLOVNI ODHODKI (121) = vsota postavk 5-8
            odhodki = (g("5. Stroški blaga, materiala in storitev") + g("6. Stroški dela")
                       + g("7. Odpisi vrednosti") + g("8. Drugi poslovni odhodki"))
            result.aop_data["121"] = AopEntry(aop="121", current_year=odhodki, previous_year=None)

            presezek_posl = prihodki - odhodki
            result.gvaop_data["7209401"] = presezek_posl  # III. PRESEŽEK POSL. PRIHODKOV/ODHODKOV

            # IV. FINANČNI PRIHODKI (153) = vsota postavk 9-11
            fin_prih = (g("9. Finančni prihodki iz deležev") + g("10. Finančni prihodki iz danih posojil")
                        + g("11. Finančni prihodki iz poslovnih terjatev"))
            result.aop_data["153"] = AopEntry(aop="153", current_year=fin_prih, previous_year=None)

            # V. FINANČNI ODHODKI (166) = vsota postavk 12-14
            fin_odh = (g("12. Finančni odhodki iz oslabitve in odpisov finančnih naložb") + g("13. Finančni odhodki iz finančnih obveznosti")
                       + g("14. Finančni odhodki iz poslovnih obveznosti"))
            result.aop_data["166"] = AopEntry(aop="166", current_year=fin_odh, previous_year=None)

            # VIII. PRESEŽEK PRIHODKOV/ODHODKOV = III + IV - V + VI(15) - VII(16)
            presezek_final = presezek_posl + fin_prih - fin_odh + g("15. DRUGI PRIHODKI") - g("16. DRUGI ODHODKI")
            result.gvaop_data["72095"] = presezek_final

            # X. ČISTI PRESEŽEK PRIH./ODH. OBR. OBD. (185) = presežek prihodkov - presežek odhodkov (18-19)
            cisti_presezek = g("18. ČISTI PRESEŽEK PRIHODKOV OBRAČUNSKEGA OBDOBJA") - g("19. ČISTI PRESEŽEK ODHODKOV OBRAČUNSKEGA OBDOBJA")
            result.aop_data["185"] = AopEntry(aop="185", current_year=cisti_presezek, previous_year=None)

    if (not c.period_from or len(c.period_from) < 6) and c.period_to:
        leto = c.period_to.split('.')[-1]
        c.period_from = f"1.1.{leto}"

    c.tip_subjekta = "4"
    result.company = finalize_company(c, tip_override, "4")
    result.subject_type = "DR"
    if not result.company.name:
        result.warnings.append("Ime društva ni bilo zanesljivo prepoznano — prosim preveri/vpiši ročno v D3.")
    return validate(result)

# ── PARSER: "Zavod / Inštitut" medletni izvoz (interno računovodsko poročilo,   ──
# dva lista "BS <datum>" / "IPI <datum>", zaporedne Zap. številke namesto AOP kod,  ──
# skrajšana besedila postavk, en sam stolpec z zneskom - brez primerjave s prejšnjim ──
# letom). Uporablja se za zavode (tip_subjekta 5 ali 6); ciljni CBK template je      ──
# STANDARDNI template.xlsx (ista struktura kot za gospodarske družbe), zato tu       ──
# vračamo navadne AOP kode kot pri ostalih parserjih - noben poseben template ni     ──
# potreben. ─────────────────────────────────────────────────────────────────────────

_ZAVOD_ABBR = [
    (r'\bdolg\.\s*', 'dolgoročne '),
    (r'\bkrat\.\s*', 'kratkoročne '),
    (r'\bkratk\.\s*', 'kratkoročne '),
    (r'\bposl\.\s*', 'poslovne '),
    (r'\bterj\.\s*', 'terjatve '),
    (r'\bčas\.\s*', 'časovne '),
    (r'\bOOS\b', 'opredmetena osnovna sredstva'),
    (r'\bDI\b', 'drobni inventar'),
    (r'\bprevred\.\s*', 'prevrednotovalni '),
    (r'\bopred\.\s*', 'opredmetenih '),
    (r'\bneopr\.\s*', 'neopredmetenih '),
    (r'\bmater\.\s*', 'materiala '),
    (r'\bzaposl\.\s*', 'zaposlenim '),
    (r'\bvred\.\s*', 'vrednotenja '),
    (r'\bobv\.\s*', 'obveznosti '),
    (r'\bnalož\.\s*', 'naložb '),
    (r'\bfin\.(?=obveznosti)', 'finančne '),
    (r'\bfin\.(?=naložbe)', 'finančne '),
    (r'\bfin\.\s*', 'finančni '),
]

def _zavod_expand(text):
    for pat, rep in _ZAVOD_ABBR:
        text = re.sub(pat, rep, text, flags=re.I)
    return re.sub(r'\s+', ' ', text).strip()

# Neposredni aliasi (natančen, mala črka, cel label vklj. z oznako) -> AOP koda,
# ali None za postavke, ki jih namerno preskočimo (kontrolne vrstice ipd.)
_ZAVOD_DIRECT_BS = {
    "sredstva": "001",
    "kontrola-mora biti 0": None,
    "kontrola sredstev": None,
    "kontrola obveznosti": None,
    "c) dolg.odloženi stroški razvijanja": "007",
    "6. opred.osnovna sredstva v gradnji in izdelavi": "016",
    "i. sredstva (skupine za odtujitev) za prodajo": "033",
    "3. lastne delnice in lastni posl.deleži(kot odbitna p": "064",
    "1. rezervacije": "073",  # JOLP_TO_AOP ima "1. Rezervacije" dvakrat (BS:073, IPI:149) -
                              # drugi zapis povozi prvega, zato ga tu eksplicitno popravimo za BS
}
_ZAVOD_DIRECT_IPI = {
    "b. povečanje vrednosti zalog proizv. in nedokon. proizvodnje": "121",
    "c. zmanjšanje vrednosti zalog proizv. in nedokon. proizvodnje": "122",
    "d. subvencije, dotacije, regresi, kompenzacije in drugi prih., ki so povezani s poslov. učinki": "124",
    "1. nabavna vrednost prodanega blaga in mater.": "129",
    "1: stroški plač": "140",
    "2. prevred.poslovni odhodki pri neopr. sredstvih": "146",
    "3. prevrednotovalni poslovni": "147",
    "i. fin.odhodki iz oslabitev in odpisov fin.naložb": "168",
    "i. subvencije, dotacije in podobni prihodki, ki": "179",
    "p. davek iz dobička": "184",
    "r. odloženi davki": "185",
}

# Postavke, ki se v tem formatu ponovijo z enakim (skrajšanim) besedilom v več
# sklopih (domači trg/EU/izven EU, dolgoročno/kratkoročno ...) - razrešimo jih
# glede na kontekst zadnje prepoznane nadrejene postavke (isti vzorec kot
# _AMBIGUOUS_ITEMS/_CONTEXT_TRIGGERS zgoraj za UPR format).
_ZAVOD_AMBIGUOUS_BS = {
    "delnice in deleži v družbah v skupini": {"dolg": "021", "krat": "042"},
    "druge delnice in deleži": {"dolg": "022", "krat": "043"},
}
_ZAVOD_CONTEXT_TRIGGERS_BS = {"019": "dolg", "040": "krat"}

_ZAVOD_AMBIGUOUS_IPI = {
    "čisti prihodki od prodaje proizv. in storitev": {"domaci": "112", "eu": "116", "izven_eu": "119"},
    "čisti prihodki od prodaje blaga in mater.": {"domaci": "114", "eu": "117", "izven_eu": "120"},
    "fin.prihodki iz deležev v družbah v skupini": {"delezi": "156"},
    "fin.prihodki iz deležev v pridruženih družbah": {"delezi": "157"},
    "fin.prihodki iz deležev v drugih družbah": {"delezi": "158"},
    "fin.prihodki iz drugih naložb": {"delezi": "159"},
    "fin.prihodki iz posojil, danih družbam v skupini": {"posojila": "161"},
    "fin.prihodki iz posojil, danih drugim": {"posojila": "162"},
    "fin.prihodki iz posl.terjatev do družb v skupini": {"posl_terj": "164"},
    "fin.prihodki iz posl.terjatev do drugih": {"posl_terj": "165"},
    "fin.odhodki iz posojil, prejetih od družb v skupini": {"fin_obv": "170"},
    "fin.odhodki iz posojil, prejetih od bank": {"fin_obv": "171"},
    "fin.odhodki iz izdanih obveznic": {"fin_obv": "172"},
    "fin.odhodki iz drugih fin.obveznosti": {"fin_obv": "173"},
    "fin.odhodki iz posl.obveznosti do družb v skupini": {"posl_obv": "175"},
    "fin.odhodki iz obv.do dobaviteljev in meničnih obv.": {"posl_obv": "176"},
    "fin.odhodki iz drugih posl.obveznosti": {"posl_obv": "177"},
}
_ZAVOD_CONTEXT_TRIGGERS_IPI = {
    "111": "domaci", "115": "eu", "118": "izven_eu",
    "156": "delezi", "160": "posojila", "163": "posl_terj",
    "169": "fin_obv", "174": "posl_obv",
}

def _match_zavod_label(label, direct_map, ambiguous_map, context):
    raw_key = label.strip().lower()
    if raw_key in direct_map:
        return direct_map[raw_key]
    kind, text = _classify_and_strip_item(label)
    text_expanded = _zavod_expand(text)
    aop = _JOLP_TO_AOP_NORM.get((kind, text_expanded))
    if aop:
        return aop
    opts = ambiguous_map.get(text_expanded) or ambiguous_map.get(text)
    if opts and context in opts:
        return opts[context]
    return None

def _is_zavod_medletni_sheet(ws):
    a2 = ws["A2"].value
    a4 = ws["A4"].value
    if not a2 or not a4:
        return False
    a2 = str(a2)
    return str(a4).strip().rstrip('.') == "Zap" and (
        "BS za družbe" in a2 or "Izkaz poslovnega izida" in a2
    )

def _find_zavod_medletni_sheets(wb):
    """Vrne (bs_sheet_name, ipi_sheet_name) ali (None, None) če format ni prepoznan.
    Imena listov se spreminjajo z datumom, zato NE iščemo po imenu, ampak po vsebini."""
    bs_sheet = ipi_sheet = None
    for sn in wb.sheetnames:
        ws = wb[sn]
        if not _is_zavod_medletni_sheet(ws):
            continue
        a2 = str(ws["A2"].value)
        if "BS za družbe" in a2:
            bs_sheet = sn
        elif "Izkaz poslovnega izida" in a2:
            ipi_sheet = sn
    return bs_sheet, ipi_sheet

def _parse_zavod_medletni_sheet(ws, direct_map, ambiguous_map, triggers):
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    items = []  # [zap, label, value]
    for row in rows:
        zap = row[0] if len(row) > 0 else None
        label = row[2] if len(row) > 2 else None
        value = row[3] if len(row) > 3 else None
        has_zap = zap is not None and str(zap).strip() != "" and str(zap).strip().isdigit()
        has_label = label is not None and str(label).strip() != ""
        if has_zap:
            items.append([zap, str(label).strip() if has_label else "", value])
        elif has_label and value is None and items:
            cont = str(label).strip()
            if cont.lower().startswith("kontrola"):
                continue  # naslov ločenega kontrolnega bloka, ni nadaljevanje opisa
            items[-1][1] = (items[-1][1] + " " + cont).strip()

    aop_data = {}
    context = None
    for zap, label, value in items:
        if label.strip().lower().startswith("kontrola"):
            continue
        aop = _match_zavod_label(label, direct_map, ambiguous_map, context)
        if aop and aop in triggers:
            context = triggers[aop]
        if not aop or value is None:
            continue
        try:
            aop_data[aop] = AopEntry(aop=aop, current_year=float(value), previous_year=None)
        except (TypeError, ValueError):
            continue
    return aop_data

def parse_zavod_medletni_excel(xlsx_path, tip_override=None, tip_subjekta_override=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    bs_sheet, ipi_sheet = _find_zavod_medletni_sheets(wb)
    result = ParseResult(pdf_format="ZAVOD_MEDLETNI")

    c = CompanyInfo()
    ref_sheet = wb[bs_sheet] if bs_sheet else wb[ipi_sheet]
    name = ref_sheet["A1"].value
    if name:
        c.name = str(name).strip()
    period_src = str(ref_sheet["A2"].value or "")
    m = re.search(r'(\d{1,2}\.\d{1,2}\.\d{4})\s*-\s*(\d{1,2}\.\d{1,2}\.\d{4})', period_src)
    if m:
        c.period_from, c.period_to = m.group(1), m.group(2)

    # tip_subjekta 5 (javni zavod) ali 6 (zasebni zavod) izbere uporabnik v obrazcu;
    # privzeto (če ni izbrano) pustimo splošno gospodarsko družbo, da ne "uganjujemo".
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)

    if not result.company.registration_number:
        result.warnings.append(
            "Ta format (medletni izvoz za zavod) ne vsebuje matične številke — vpiši jo ročno v celico D4."
        )

    if bs_sheet:
        bs_aop = _parse_zavod_medletni_sheet(wb[bs_sheet], _ZAVOD_DIRECT_BS, _ZAVOD_AMBIGUOUS_BS, _ZAVOD_CONTEXT_TRIGGERS_BS)
        result.aop_data.update(bs_aop)
    else:
        result.warnings.append("List z bilanco stanja (BS) ni bil najden v datoteki.")

    if ipi_sheet:
        ipi_aop = _parse_zavod_medletni_sheet(wb[ipi_sheet], _ZAVOD_DIRECT_IPI, _ZAVOD_AMBIGUOUS_IPI, _ZAVOD_CONTEXT_TRIGGERS_IPI)
        result.aop_data.update(ipi_aop)
    else:
        result.warnings.append("List z izkazom poslovnega izida (IPI) ni bil najden v datoteki.")

    return validate(result)


# ── PARSER: GVIN "izvoz bilance" (bonitetni izvoz z lastnimi GVAOP kodami, en list,   ──
# stolpci Gvaop|Kategorije|Vrednost). Uporablja se, ko stranka/AJPES nima uradnih         ──
# podatkov, pa jih uporabnik dobi iz GVIN-a. Vrednosti so brez primerjave s prejšnjim     ──
# letom (samo tekoče leto), format ustreza polnemu zaključenemu letu (01.01.-31.12.).     ──
# Ciljni CBK template je standardni template.xlsx - GVIN uporablja institucionalno        ──
# terminologijo (npr. "Lastni viri" namesto "Kapital", "Presežek prihodkov/odhodkov"      ──
# namesto "Dobiček/izguba"), ki jo preslikamo na uradne AOP kode prek GVIN_DIRECT_ALIASES. ──
# OPOMBA: GVIN nekatere uradne AOP postavke združuje v eno kategorijo (npr. "Nabavna       ──
# vrednost prodanega blaga in materiala ter stroški porabljenega materiala" je ena         ──
# vrednost namesto dveh ločenih AOP), zato ostane nekaj podrobnejših AOP celic prazna -   ──
# to ni napaka parserja, ampak omejitev GVIN-ovega povzetka (v izvorni bilanci ta          ──
# podrobnost sploh ni na voljo). ──────────────────────────────────────────────────────────

_GVIN_DIRECT_ALIASES = {
    "sredstva": "001",
    "sredstva (skupine za odtujitev) za prodajo": "033",
    "lastni viri": "056",
    "ustanovitveni vložek": "058",
    "čisti presežek prihodkov poslovnega leta": "070",
    "čisti presežek odhodkov poslovnega leta": "071",
    "poslovni prihodki": None,  # podvojena vrednost "Kosmati donos od poslovanja" - preskoči
    "presežek poslovnih prihodkov/odhodkov": "151",
    "čisti presežek prihodkov/odhodkov obračunskega obdobja": "186",
    "davek od dohodkov": "184",
}

def _gvin_match_label(label):
    key = label.strip().lower()
    if key in _GVIN_DIRECT_ALIASES:
        return _GVIN_DIRECT_ALIASES[key]
    for kind in ("top", "item"):
        aop = _JOLP_TO_AOP_NORM.get((kind, key))
        if aop:
            return aop
    return None

def _gvin_parse_value(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.lower() in ("n.p.", "", "-"):
        return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except (TypeError, ValueError):
        return None

def _is_gvin_izvoz_workbook(wb):
    if len(wb.sheetnames) != 1:
        return False
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=1, values_only=True):
        if row and row[0] and str(row[0]).strip() == "Gvaop":
            return True
    return False

def parse_gvin_izvoz_excel(xlsx_path, tip_override=None, tip_subjekta_override=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = {}
    header_end_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if row and row[0]:
            key = str(row[0]).strip()
            if key == "Gvaop":
                header_end_row = i
                break
            header[key] = row[1] if len(row) > 1 else None

    result = ParseResult(pdf_format="GVIN_IZVOZ")
    c = CompanyInfo()
    c.name = str(header.get("Subjekt", "") or "").strip()
    reg = str(header.get("Matična številka", "") or "").strip()
    if reg:
        c.registration_number = reg
    tax = str(header.get("Davčna številka", "") or "").strip()
    if tax:
        c.tax_number = re.sub(r'^SI', '', tax, flags=re.I)

    m1 = re.search(r'(\d{1,2})\.\s*(\d{1,2})\.\s*(\d{4})', str(header.get("Začetek poslovnega leta", "") or ""))
    m2 = re.search(r'(\d{1,2})\.\s*(\d{1,2})\.\s*(\d{4})', str(header.get("Zaključek poslovnega leta", "") or ""))
    if m1:
        c.period_from = f"{int(m1.group(1)):02d}.{int(m1.group(2)):02d}.{m1.group(3)}"
    if m2:
        c.period_to = f"{int(m2.group(1)):02d}.{int(m2.group(2)):02d}.{m2.group(3)}"

    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)

    aop_data = {}
    if header_end_row:
        for row in ws.iter_rows(min_row=header_end_row + 1, values_only=True):
            gvaop, label, val = (row + (None, None, None))[:3]
            if not gvaop or not str(gvaop).strip().startswith("["):
                continue
            code_num = str(gvaop).strip("[]")
            if code_num in ("090", "091"):
                continue
            if code_num.startswith("9") and len(code_num) == 5 and code_num[1] == "2":
                continue  # bonitetni kazalniki (92xxx) - niso AOP postavke
            if not label:
                continue
            aop = _gvin_match_label(str(label))
            if not aop:
                continue
            v = _gvin_parse_value(val)
            if v is not None:
                aop_data[aop] = AopEntry(aop=aop, current_year=v, previous_year=None)
    result.aop_data.update(aop_data)

    result.warnings.append(
        "Vir podatkov je GVIN (ne uradni AJPES/CRMB), zato nekatere podrobnejše AOP postavke, "
        "ki jih GVIN združuje v skupne kategorije (npr. nabavna vrednost blaga skupaj s stroški "
        "materiala), v templatu ostanejo prazne - preveri pred oddajo."
    )
    return validate(result)


# ── PARSER: "BS za AJPES" / "IPI za AJPES" izvoz (.xlsx, en list "Sheet1", polni uradni    ──
# besedilni opisi postavk s formulskimi referencami v oklepaju - zelo podoben mehanizem     ──
# ujemanja kot pri OBRACUNI_MEDLETNI (glej _match_obracuni_label zgoraj), le da je TOKRAT   ──
# bilanca stanja in izkaz poslovnega izida vsak v SVOJI datoteki (uporabnik ju naloži       ──
# skupaj, obstoječi mehanizem za več datotek jih združi). Format VSEBUJE ime podjetja in    ──
# obdobje (za razliko od OBRACUNI_MEDLETNI) - izluščimo ju avtomatsko namesto ročnega       ──
# vnosa. Matične številke NIMA. POMEMBNA POSEBNOST: stran "obveznosti do virov sredstev"    ──
# (od AOP 055 dalje) je v izvoru izpisana z NASPROTNIM (negativnim) predznakom              ──
# (knjigovodska konvencija kredit=minus) - vrednosti od te vrstice dalje obrnemo nazaj na   ──
# pozitivne, da se ujemajo z uradnim obrazcem (kjer SREDSTVA==OBVEZNOSTI DO VIROV SREDSTEV).──
_EXTAJPES_DIRECT_BS = {k.lower(): v for k, v in {
    "SREDSTVA (002+032+053)": "001",
    "č) Dolgoročna neopredmetena sredstva": "008",
    "4. Druge naprave in oprema, drobni inventar in druga opredmetena osnovna sredstv": "014",
    "I. Sredstva (skupine za odtujitev) za prodajo": "033",
    "2. Rezerve za lastne deleže in lastne poslovne deleže": "063",
    "1. Rezervacije": "073",  # JOLP_TO_AOP ima "1. Rezervacije" dvakrat (BS:073, IPI:149) -
                              # drugi zapis povozi prvega, zato ga tu eksplicitno popravimo za BS
    "2. Dolgoročne PČR": "074",
    "C. DOLGOROČNE OBVEZNOSTIi (076+080+084)": "075",
    "I. Obveznosti, vključene v skupino za odtujitev": "086",
}.items()}

_EXTAJPES_DIRECT_IU = {k.lower(): v for k, v in {
    "D. SUBVENCIJE,DOTACIJE,REGRESI,KOMP. IN DRUGI PRIHODKI,KI SO POV.S POSL.UČINKI": "124",
    "1. Nabavna vrednost prodanega blaga in materiala": "129",
    "2. Stroški prabljenega materiala (131 do 133)": "130",
    "c) povračila stroškov zaposlencev v zvezi z delom": "137",
    "Stroški pokojninskega zavarovanja": "141",
    "2. Prevrednotovalni poslovni odhodki pri neop.sr.in opr.sr.": "146",
    "3. Prevrednotovalni poslovni odhodki pri obratnih sredstih": "147",
    "1. Finančni prihodki iz danih posojil, danih družbam v skupini": "161",
    "2. Finančni prihodki iz danih posojil, danih drugim": "162",
    "P. DAVEK IZ DOBIČKA": "184",
    "R. ODLOŽENI DAVKI": "185",
}.items()}

_EXTAJPES_SKIP_PATTERNS = (
    "BS ZA AJPES", "IPI ZA AJPES", "OBDOBJE:", "ZAČETNI DATUM POSLOVNEGA LETA:",
    "VSI ZNESKI SO V EUR", "STRAN ", "OPIS",
)


def _is_extajpes_xlsx(rows):
    """Vrne 'BS', 'IPI' ali None glede na vsebino celice A1."""
    if not rows or not rows[0] or not rows[0][0]:
        return None
    a1 = str(rows[0][0]).strip()
    if a1 == "BS za AJPES":
        return "BS"
    if a1 == "IPI za AJPES":
        return "IPI"
    return None


def _extajpes_extract_company(rows):
    c = CompanyInfo()
    for row in rows[:25]:
        if not row or not row[0]:
            continue
        cell = str(row[0]).strip()
        if not cell:
            continue
        m = re.match(r'Obdobje:\s*(\d{1,2})\.(\d{1,2})\.(\d{2})\.\.(\d{1,2})\.(\d{1,2})\.(\d{2})', cell)
        if m:
            d1, mo1, y1, d2, mo2, y2 = m.groups()
            c.period_from = f"{int(d1)}.{int(mo1)}.20{y1}"
            c.period_to = f"{int(d2)}.{int(mo2)}.20{y2}"
            continue
        up = cell.upper()
        if any(up.startswith(p) for p in _EXTAJPES_SKIP_PATTERNS):
            continue
        if "\\" in cell or re.fullmatch(r'\d{1,2}\.\d{1,2}\.\d{2,4}', cell) or re.fullmatch(r'Stran\s*\d+\s*/\s*\d+', cell, re.IGNORECASE):
            continue
        if not c.name:
            c.name = cell
    return c


def parse_extajpes_xlsx(xlsx_path, tip_override=None, tip_subjekta_override=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    kind = _is_extajpes_xlsx(rows)

    result = ParseResult(pdf_format="BS_IPI_AJPES")
    c = _extajpes_extract_company(rows)
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    if not result.company.registration_number:
        result.warnings.append(
            "Ta format (BS/IPI za AJPES) ne vsebuje matične številke — vpiši jo ročno v celico D4."
        )

    is_bs = (kind == "BS")
    direct_map = _EXTAJPES_DIRECT_BS if is_bs else _EXTAJPES_DIRECT_IU
    ambiguous_map = _ZAVOD_AMBIGUOUS_BS if is_bs else _OBRACUNI_AMBIGUOUS_IU
    triggers = _ZAVOD_CONTEXT_TRIGGERS_BS if is_bs else _OBRACUNI_TRIGGERS_IU

    context = None
    liab_side = False  # od AOP 055 (OBVEZNOSTI DO VIROV SREDSTEV) dalje - predznak obrnemo
    for row in rows:
        if not row or not row[0] or not str(row[0]).strip():
            continue
        label = str(row[0])
        cur_raw = row[2] if len(row) > 2 else None
        prev_raw = row[4] if len(row) > 4 else None
        aop = _match_obracuni_label(label, direct_map, ambiguous_map, context)
        if not aop:
            continue
        if aop in triggers:
            context = triggers[aop]
        current = parse_si(cur_raw) if cur_raw not in (None, "") else None
        previous = parse_si(prev_raw) if prev_raw not in (None, "") else None
        if is_bs:
            if aop == "055":
                liab_side = True
            if liab_side:
                if current is not None:
                    current = -current
                if previous is not None:
                    previous = -previous
        if current is not None or previous is not None:
            result.aop_data[aop] = AopEntry(aop=aop, current_year=current, previous_year=previous)

    if not rows or kind is None:
        result.warnings.append("Format BS/IPI za AJPES ni bil prepoznan v tej datoteki.")

    return validate(result)


def parse_xlsx_file(xlsx_path, tip_override=None, tip_subjekta_override=None):
    wb_check = openpyxl.load_workbook(xlsx_path, data_only=True)
    if _is_gvin_izvoz_workbook(wb_check):
        return parse_gvin_izvoz_excel(xlsx_path, tip_override, tip_subjekta_override)
    zavod_bs, zavod_ipi = _find_zavod_medletni_sheets(wb_check)
    if zavod_bs or zavod_ipi:
        return parse_zavod_medletni_excel(xlsx_path, tip_override, tip_subjekta_override)
    if _is_drustvo_workbook(wb_check, tip_subjekta_override):
        return parse_jolp_drustvo_excel(xlsx_path, tip_override, tip_subjekta_override)
    if _find_gdzadruge_sheet(wb_check):
        return parse_gdzadruge_excel(xlsx_path, tip_override, tip_subjekta_override)
    first_ws = wb_check[wb_check.sheetnames[0]]
    first_rows = [tuple(row) for row in first_ws.iter_rows(values_only=True)]
    if _is_extajpes_xlsx(first_rows):
        return parse_extajpes_xlsx(xlsx_path, tip_override, tip_subjekta_override)
    wb = wb_check
    rows_by_sheet = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_by_sheet[sheet_name] = [tuple(row) for row in ws.iter_rows(values_only=True)]
    return parse_excel_rows(rows_by_sheet, tip_override, tip_subjekta_override, "XLSX")

# ── PARSER: "Poročilo GD" .xls izvoz (NAZIV | AOP | Znesek stolpci, ena vrednost na vrstico) ──
# Zelo čist format - AOP koda je vedno v stolpcu B kot celo število, zato lahko brez tveganja
# uporabimo širši razpon (1-320) kot pri splošnem _aop_from_cell (ki namenoma izloči kode <100,
# da se izogne napačnim ujemanjem pri skeniranju poljubnih stolpcev pri drugih formatih).
def _is_porocilo_gd_xls(rows):
    for row in rows[:5]:
        if row and isinstance(row[0], str) and "Poročilo GD" in row[0]:
            return True
    return False

def parse_porocilo_gd_xls(rows, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="POROCILO_GD")
    c = CompanyInfo()
    is_ipi = False
    for row in rows[:10]:
        if not row: continue
        v0 = row[0]
        if not isinstance(v0, str): continue
        v0s = v0.strip()
        if v0s.startswith("Poročilo GD") and "izida" in v0s.lower():
            is_ipi = True
        if v0s.startswith("za ") and not c.name:
            c.name = v0s[3:].strip()
        m = re.search(r'NA DAN\s+(\d{1,2})\.(\d{1,2})\.(\d{4})', v0s, flags=re.I)
        if m:
            d,mo,y = m.groups()
            c.period_to = f"{int(d)}.{int(mo)}.{y}"
        m2 = re.search(r'za obdobje od\s+(\d{1,2})\.(\d{4})\s+do\s+(\d{1,2})\.(\d{4})', v0s, flags=re.I)
        if m2:
            m1,y1,m2_,y2 = m2.groups()
            c.period_from = f"1.{int(m1)}.{y1}"
            last_day = calendar.monthrange(int(y2), int(m2_))[1]
            c.period_to = f"{last_day}.{int(m2_)}.{y2}"

    if (not c.period_from or len(c.period_from) < 6) and c.period_to:
        leto = c.period_to.split('.')[-1]
        c.period_from = f"1.1.{leto}"

    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    if not result.company.registration_number:
        result.warnings.append("Ta format ne vsebuje matične številke — vpiši jo ročno v celico D4.")

    for row in rows:
        if not row or len(row) < 3: continue
        aop_raw = row[1]
        if not isinstance(aop_raw, (int, float)) or isinstance(aop_raw, bool):
            continue
        if not float(aop_raw).is_integer():
            continue
        n = int(aop_raw)
        if not (1 <= n <= 320):
            continue
        code = str(n).zfill(3)
        if code in result.aop_data:
            continue
        val = row[2] if len(row) > 2 else None
        num = val if isinstance(val, (int, float)) and not isinstance(val, bool) else parse_si(val) if isinstance(val, str) else None
        if num is not None:
            result.aop_data[code] = AopEntry(aop=code, current_year=num, previous_year=None)
    return validate(result)

# ── PARSER: AJPES-native "Bilanca" .xls izvoz (listi BS / IPIZ, eksplicitna kolona    ──
# "Oznaka za AOP"). Zelo čist format, ker ima AOP kodo že eksplicitno navedeno v svoji  ──
# koloni (ni je treba ugibati iz besedila). Datum tekočega/primerjalnega obdobja je     ──
# zakodiran kot Excel serijski datum v glavi vsakega lista - dekodiramo ga prek         ──
# xlrd.xldate_as_datetime. Nekateri izvozi vsebujejo še dodaten list "podatki" s        ──
# starimi/nepovezanimi arhivskimi podatki (druga AOP-oštevilčena tabela) - tega lista   ──
# NAMENOMA ne beremo, ker bi se njegove vrednosti pomešale s pravimi in pokvarile       ──
# rezultat (SREDSTVA in OBVEZNOSTI DO VIROV SREDSTEV se morata ujemati na cent natančno,──
# kar se z branjem "podatki" lista pokvari). Ta format prav tako ne vsebuje imena       ──
# podjetja niti matične številke - uporabnik ju vpiše ročno (enako kot pri UPR/zavod    ──
# medletnih formatih).

# ── PARSER: "naravni" AJPES xls izvoz s polno glavo podjetja (Konto/Postavka/Oznaka za    ──
# AOP/Znesek, na obeh listih BS in IPI), kjer sta podstolpca "Tekočega leta (DATUM)" in     ──
# "Prejšnjega leta (DATUM)" pod glavnim stolpcem "Znesek" - POMEMBNO: vrstni red teh dveh   ──
# podstolpcev se med izvozi RAZLIKUJE (pri večini je D=tekoče/E=prejšnje, pri nekaterih pa   ──
# ravno obratno - D=prejšnje/E=tekoče) - zato ju NIKOLI ne beremo po fiksni poziciji, ampak  ──
# vedno preberemo BESEDILO glave in dinamično ugotovimo, kateri fizični stolpec je kateri.  ──
# Ta format (za razliko od AJPES_NATIVE_XLS/OBRACUNI_MEDLETNI) vsebuje polne podatke o       ──
# podjetju (ime, sedež, matična, davčna) - izluščimo jih namesto ročnega vnosa.              ──
# ── PARSER: "Finančna razkritja poslovanja" xls izvoz (banka/FURS obrazec, npr. Bankart/  ──
# GORE-vrste izvozi) - VELIKO listov (Seznami, Osnovni podatki, Prva stran, BS, IPI, ter    ──
# vrsta podrobnostnih razčlenitvenih listov po posameznih AOP kodah, npr. "003 Neopred.OS", ──
# "010 Opred.OS" ipd.) - od tega naju zanimata SAMO lista "BS" in "IPI", preostale          ──
# razčlenitvene liste namenoma IGNORIRAMO (razčlenjujejo že zajete zbirne AOP vrednosti,    ──
# ne prinašajo novih). Glava obeh listov ima stolpce Upor./Konto/X/Postavka/(Oznaka za)     ──
# AOP/Tekoče obdobje/Prejšnje obdobje - stolpca s trenutnimi/preteklimi vrednostmi spet     ──
# poiščemo dinamično po besedilu (enako kot pri URADNI_XLS), ne po fiksni poziciji. Ime     ──
# podjetja, matična, davčna številka in obdobje poročanja se, če je prisoten, raje         ──
# preberejo iz lista "Osnovni podatki" (bolj zanesljivo, structured), sicer iz glave BS/IPI.──
def _find_letinvest_xls_sheets(rows_by_sheet):
    found = {}
    for name, rows in rows_by_sheet.items():
        if name not in ("BS", "IPI"):
            continue
        for i, row in enumerate(rows[:15]):
            if not row or not isinstance(row[0], str) or row[0].strip() != "Upor.":
                continue
            col_aop = col_cur = col_prev = None
            for j, v in enumerate(row):
                if not isinstance(v, str) or not v:
                    continue
                # Nekateri izvozi imajo v isti glavi ŠE EN par "Tekoče/Prejšnje obdobje"
                # bolj desno (stranski "Kontrola usklajenosti" panel s kontrolnimi seštevki,
                # ne dejanski podatkovni stolpci) - vzamemo PRVI zadetek, ne zadnjega.
                if col_aop is None and "AOP" in v:
                    col_aop = j
                elif col_cur is None and "Tekoče obdobje" in v:
                    col_cur = j
                elif col_prev is None and "Prejšnje obdobje" in v:
                    col_prev = j
            if col_aop is not None and col_cur is not None and col_prev is not None:
                found[name] = {"header_row": i, "col_aop": col_aop, "col_cur": col_cur, "col_prev": col_prev}
            break
    return found


def _aop_code_from_any(v):
    """AOP koda se med listi pojavi včasih kot niz ('001'), včasih kot število (110.0)."""
    if isinstance(v, bool):
        return None
    if isinstance(v, str):
        s = v.strip()
        return s.zfill(3) if s.isdigit() else None
    if isinstance(v, (int, float)):
        iv = int(v)
        if float(iv) == v and 1 <= iv <= 320:
            return str(iv).zfill(3)
    return None


def _letinvest_osnovni_podatki(book):
    if "Osnovni podatki" not in book.sheet_names():
        return None
    sh = book.sheet_by_name("Osnovni podatki")
    dm = book.datemode
    data = {}
    for r in range(sh.nrows):
        if sh.ncols <= 3:
            continue
        label = sh.cell_value(r, 1)
        if not isinstance(label, str) or not label.strip():
            continue
        data[label.strip()] = sh.cell_value(r, 3)

    c = CompanyInfo()
    name = data.get("Naziv posl.subjekta")
    if name:
        c.name = str(name).strip()
    maticna = data.get("Matična številka")
    if maticna not in (None, ""):
        c.registration_number = str(int(maticna)) if isinstance(maticna, (int, float)) else str(maticna).strip()
    davcna = data.get("Davčna številka")
    if davcna not in (None, ""):
        c.tax_number = str(int(davcna)) if isinstance(davcna, (int, float)) else str(davcna).strip()

    def _conv(v):
        if isinstance(v, (int, float)) and v > 0:
            try:
                d = xlrd.xldate_as_datetime(v, dm)
                return f"{d.day}.{d.month}.{d.year}"
            except Exception:
                return None
        return None

    d_od = _conv(data.get("Obdobje Od"))
    d_do = _conv(data.get("Obdobje Do"))
    if d_od:
        c.period_from = d_od
    if d_do:
        c.period_to = d_do
    return c


def _letinvest_company_from_headers(rows_by_sheet, sheets):
    c = CompanyInfo()
    for name in sheets:
        for row in rows_by_sheet[name][:6]:
            if not row:
                continue
            for cell in row:
                if not isinstance(cell, str):
                    continue
                cell_s = cell.strip()
                if not cell_s:
                    continue
                up = cell_s.upper()
                if not c.name and re.search(r'\b(d\.?\s*o\.?\s*o\.?|d\.?\s*d\.?)\b', cell_s, re.IGNORECASE) \
                        and "BILANCA" not in up and "IZKAZ" not in up:
                    c.name = cell_s
                m = re.search(r'na dan\s*(\d{1,2})\.(\d{1,2})\.(\d{4})', cell_s)
                if m and not c.period_to:
                    d, mo, y = m.groups()
                    c.period_to = f"{int(d)}.{int(mo)}.{y}"
                    c.period_from = f"1.1.{y}"
                m2 = re.search(r'od\s*(\d{1,2})\.(\d{1,2})\.(\d{4})\s*do\s*(\d{1,2})\.(\d{1,2})\.(\d{4})', cell_s)
                if m2 and not c.period_to:
                    d1, mo1, y1, d2, mo2, y2 = m2.groups()
                    c.period_from = f"{int(d1)}.{int(mo1)}.{y1}"
                    c.period_to = f"{int(d2)}.{int(mo2)}.{y2}"
    return c


def parse_letinvest_xls(xls_path, tip_override=None, tip_subjekta_override=None):
    book = xlrd.open_workbook(str(xls_path))
    rows_by_sheet = {}
    for sheet in book.sheets():
        rows_by_sheet[sheet.name] = [tuple(sheet.cell_value(r, c) for c in range(sheet.ncols)) for r in range(sheet.nrows)]

    sheets = _find_letinvest_xls_sheets(rows_by_sheet)
    result = ParseResult(pdf_format="LETINVEST_XLS")

    c = _letinvest_osnovni_podatki(book) or CompanyInfo()
    if not c.name or not c.period_to:
        hc = _letinvest_company_from_headers(rows_by_sheet, sheets)
        c.name = c.name or hc.name
        c.period_to = c.period_to or hc.period_to
        c.period_from = c.period_from or hc.period_from

    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)

    if not sheets:
        result.warnings.append("Lista BS/IPI (finančna razkritja poslovanja) nista bila prepoznana v tej datoteki.")

    for name, info in sheets.items():
        rows = rows_by_sheet[name]
        col_aop, col_cur, col_prev = info["col_aop"], info["col_cur"], info["col_prev"]
        for row in rows[info["header_row"] + 2:]:
            if len(row) <= max(col_aop, col_cur, col_prev):
                continue
            code = _aop_code_from_any(row[col_aop])
            if not code:
                continue
            cur = row[col_cur] if isinstance(row[col_cur], (int, float)) and not isinstance(row[col_cur], bool) else None
            prev = row[col_prev] if isinstance(row[col_prev], (int, float)) and not isinstance(row[col_prev], bool) else None
            if cur is None and prev is None:
                continue
            result.aop_data[code] = AopEntry(aop=code, current_year=cur, previous_year=prev)

    return validate(result)


def _find_native_uradni_xls_sheets(rows_by_sheet):
    """Vrne {sheet_name: {'kind': 'BS'|'IPI', 'header_row': int, 'col_aop': int,
    'col_cur': int, 'col_prev': int}} za vsak prepoznan list."""
    found = {}
    for name, rows in rows_by_sheet.items():
        for i, row in enumerate(rows[:25]):
            if not row or len(row) < 4:
                continue
            c0 = str(row[0]).strip() if isinstance(row[0], str) else ""
            c2 = str(row[2]).strip() if isinstance(row[2], str) else ""
            c3 = str(row[3]).strip() if isinstance(row[3], str) else ""
            if c0 not in ("Konto", "Postavka") or "AOP" not in c2 or c3 != "Znesek":
                continue
            sub = rows[i + 1] if i + 1 < len(rows) else None
            col_cur = col_prev = None
            if sub:
                for j, v in enumerate(sub):
                    if not isinstance(v, str) or not v:
                        continue
                    if "Tekočega leta" in v:
                        col_cur = j
                    elif "Prejšnjega leta" in v:
                        col_prev = j
            if col_cur is None or col_prev is None:
                col_cur, col_prev = 3, 4  # privzeti (najpogostejši) vrstni red
            full_text = " ".join(str(r[0]) for r in rows[:i] if r and isinstance(r[0], str)).lower()
            if "bilance stanja" in full_text:
                kind = "BS"
            elif "izkaza poslovnega izida" in full_text:
                kind = "IPI"
            else:
                continue
            found[name] = {"kind": kind, "header_row": i, "col_aop": 2, "col_cur": col_cur, "col_prev": col_prev}
            break
    return found


def _native_uradni_xls_company(rows):
    c = CompanyInfo()
    for i, row in enumerate(rows[:20]):
        if not row:
            continue
        nxt_row = rows[i + 1] if i + 1 < len(rows) else None

        if isinstance(row[0], str):
            label = row[0].strip()
            nxt = nxt_row[0] if nxt_row else None
            nxt = str(nxt).strip() if nxt is not None else ""
            if label == "Ime poslovnega subjekta" and nxt:
                c.name = nxt
            elif label == "Matična številka" and nxt:
                c.registration_number = nxt
            elif label == "Davčna številka" and nxt:
                c.tax_number = nxt
            m = re.search(r'na dan\s*(\d{1,2})\.(\d{1,2})\.(\d{4})', label)
            if m:
                d, mo, y = m.groups()
                c.period_to = f"{int(d):02d}.{int(mo):02d}.{y}"
                c.period_from = f"01.01.{y}"
                continue
            m = re.search(r'v obdobju\s*(\d{1,2})\.(\d{1,2})\.(\d{4})\s*-\s*(\d{1,2})\.(\d{1,2})\.(\d{4})', label)
            if m:
                d1, mo1, y1, d2, mo2, y2 = m.groups()
                c.period_from = f"{int(d1):02d}.{int(mo1):02d}.{y1}"
                c.period_to = f"{int(d2):02d}.{int(mo2):02d}.{y2}"

        # Nekateri izvozi (npr. s.p.) imajo "Ime poslovnega subjekta"/"Sedež..." v stolpcu A,
        # a "Matična številka"/"Davčna številka" v stolpcu D ISTE vrstice (vrednost je v D
        # naslednje vrstice) - ista dvovrstična label/vrednost oblika, le v drugem stolpcu.
        if len(row) > 3 and isinstance(row[3], str):
            label3 = row[3].strip()
            nxt3 = nxt_row[3] if nxt_row and len(nxt_row) > 3 else None
            if nxt3 not in (None, ""):
                val3 = str(int(nxt3)) if isinstance(nxt3, float) and nxt3.is_integer() else str(nxt3).strip()
                if label3 == "Matična številka" and not c.registration_number:
                    c.registration_number = val3
                elif label3 == "Davčna številka" and not c.tax_number:
                    c.tax_number = val3
    return c



def _native_uradni_xls_is_sp(name, raw_codes):
    """Zazna, ali gre za s.p. (samostojni podjetnik) izvoz tega formata - bodisi po
    imenu podjetja (konča se na 's.p.') bodisi po prisotnosti s.p.-specifičnih
    črkovno-obarvanih AOP kod (npr. '060a'/'060b' na BS, '148a'/'148b' na IPI), ki jih
    gospodarske družbe nikoli nimajo."""
    if name and re.search(r'\bS\.\s*P\.?\s*$', name.strip(), re.I):
        return True
    return any(re.fullmatch(r'\d{2,3}[a-zA-Z]', code) for code in raw_codes)

# S.p. bilanca stanja/izkaz poslovnega izida uporablja UPRAVIČENO DRUGAČNO uradno AOP
# oštevilčenje kot gospodarske družbe (drugačna struktura kapitala/dohodka, brez davka
# iz dobička ipd.). Ker CBK template za s.p. (template_sp.xlsx) marsikje ISTO številko
# AOP kot pri GD uporabi za POVSEM DRUG pojem (npr. njegov "110" = "kosmati donos", ne
# "čisti prihodki od prodaje" kot pri GD!), neposredno prepisovanje po AOP kodi bi na
# nekaterih vrsticah zapisalo NAPAČNO vrednost. Zato za s.p. namesto neposrednega AOP
# ujemanja pišemo v gvaop_data, indeksiramo pa po unikatni GVAOP oznaki iz stolpca B
# templata (glej build_idx_by_gvaop) - ta je vedno nedvoumna, ne glede na AOP kolizije.
# Format: gvaop_tag -> izvorna AOP koda (neposreden prepis) ALI (koda_a, koda_b, '+'/'-')
_SP_GVAOP_MAP_BS = {
    "001": "001", "00101": "002", "0010101": "003", "0010102": "010",
    "0010104": "018", "0010105": "019", "001020201": "027", "00102": "032",
    "0010205": "033", "0010201": "034", "0010203": "040", "001020202": "048",
    "0010204": "052", "00103": "053", "002": "054", "003": "055",
    "00301": "056", "003010101": "058",  # uradni DDD s.p. obrazec: začetni podjetnikov kapital je AOP 058, ne 057
    "003010201": "060a",  # II. Prenosi stvarnega premoženja - s.p.-specifična koda
    "003010202": "060b",  # III. Pritoki in odtoki denarnih sredstev - s.p.-specifična koda
    "0030106": "067", "351": "301", "003010501": "070", "003010502": "071",
    "00302": "072", "0030303": "075", "003030301": "076", "003030302": "080",
    "0030305": "085", "0030306": "086", "003030501": "087", "003030502": "091",
    "00304": "095", "004": "096",
}
_SP_GVAOP_MAP_IPI = {
    "054": "126",                    # A. KOSMATI DONOS OD POSLOVANJA <- F. Kosmati donos (GD AOP 126)
    "050": "110",                    # I. Čisti prihodki od prodaje (skupaj) <- A. Čisti prihodki od prodaje
    "051": ("121", "122", "-"),      # II. Sprememba vrednosti zalog = Povečanje - Zmanjšanje
    "052": "123",                    # III. Usredstveni lastni proizvodi in lastne storitve
    "053": ("124", "125", "+"),      # III. Drugi poslovni prihodki (skupaj s subvencijami) = D + E
    "060": "127",                    # B. POSLOVNI ODHODKI SKUPAJ
    "055": "128", "056": "139", "057": "144", "0570101": "145", "058": "148",
    "06101": ("151", "152", "-"),    # C. DOBIČEK/IZGUBA IZ POSLOVANJA (EBIT) = H - I
    "062": "153", "064": "166",
    # E. DOBIČEK/IZGUBA REDNEGA POSLOVANJA = EBIT + fin.prihodki - fin.odhodki (3-členska
    # formula - obravnavano posebej spodaj, ne prek te preproste 2-členske tabele)
    "068": "178", "069": "181",
    "071": ("182", "183", "-"),      # J. Podjetnikov dohodek / neg. poslovni izid = N - O
    "090": "188", "091": "189",
    # "073" (Davek iz dobička) in "07301" (Odloženi davki) pri s.p. namerno NISO
    # zapisana - s.p. ne plačuje davka iz dobička na tem obrazcu (obdavčitev gre prek
    # dohodnine), zato ta dva pojma v s.p. izvozu sploh ne obstajata.
}
# IPI AOP kode, ki pri s.p. POMENIJO NEKAJ DRUGEGA na CBK templatu kot pri GD - teh NE
# smemo pustiti v result.aop_data (bi jih splošni AOP-po-kodi mehanizem zapisal v NAPAČNO
# vrstico templata), ker jih namesto tega pišemo prek gvaop_data zgoraj.
_SP_IPI_COLLISION_CODES = {"110", "111", "115", "118", "119", "121", "122", "124", "125", "151", "152", "182", "183"}


def parse_native_uradni_xls(xls_path, tip_override=None, tip_subjekta_override=None):
    book = xlrd.open_workbook(str(xls_path))
    rows_by_sheet = {}
    for sheet in book.sheets():
        rows_by_sheet[sheet.name] = [tuple(sheet.cell_value(r, c) for c in range(sheet.ncols)) for r in range(sheet.nrows)]

    sheets = _find_native_uradni_xls_sheets(rows_by_sheet)
    result = ParseResult(pdf_format="URADNI_XLS")

    c = CompanyInfo()
    for name, info in sheets.items():
        cc = _native_uradni_xls_company(rows_by_sheet[name])
        if cc.name and not c.name:
            c.name = cc.name
        if cc.registration_number and not c.registration_number:
            c.registration_number = cc.registration_number
        if cc.tax_number and not c.tax_number:
            c.tax_number = cc.tax_number
        if info["kind"] == "BS" and cc.period_to:
            c.period_to, c.period_from = cc.period_to, cc.period_from
        elif info["kind"] == "IPI" and cc.period_to and not c.period_to:
            c.period_to, c.period_from = cc.period_to, cc.period_from

    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)

    if not sheets:
        result.warnings.append("Format 'naravnega' AJPES xls izvoza ni bil prepoznan v tej datoteki.")

    raw_by_code = {}  # VSE kode (digit ali lettered npr. '060a') -> current_year vrednost
    for name, info in sheets.items():
        rows = rows_by_sheet[name]
        col_aop, col_cur, col_prev = info["col_aop"], info["col_cur"], info["col_prev"]
        for row in rows[info["header_row"] + 2:]:
            if len(row) <= max(col_aop, col_cur, col_prev):
                continue
            aop_raw = row[col_aop]
            aop = aop_raw.strip() if isinstance(aop_raw, str) else None
            if not aop:
                continue
            cur = row[col_cur] if isinstance(row[col_cur], (int, float)) and not isinstance(row[col_cur], bool) else None
            prev = row[col_prev] if isinstance(row[col_prev], (int, float)) and not isinstance(row[col_prev], bool) else None
            if cur is None and prev is None:
                continue
            if re.fullmatch(r'\d{1,3}[a-zA-Z]?', aop):
                raw_by_code[aop] = cur
            if not aop.isdigit():
                continue
            code = aop.zfill(3)
            result.aop_data[code] = AopEntry(aop=code, current_year=cur, previous_year=prev)

    is_sp = _native_uradni_xls_is_sp(result.company.name, raw_by_code.keys())
    if is_sp:
        if not tip_subjekta_override:
            result.company.tip_subjekta = "2"
            result.subject_type = _subject_type_from_tip_subjekta("2")
        # Kolizijske IPI kode odstranimo iz aop_data (pri s.p. pomenijo drug pojem kot
        # pri GD na CBK templatu) - nadomestimo jih spodaj prek gvaop_data.
        for code in _SP_IPI_COLLISION_CODES:
            result.aop_data.pop(code, None)

        def _sp_formula(formula):
            if isinstance(formula, tuple):
                code_a, code_b, op = formula
                va = raw_by_code.get(code_a) or 0
                vb = raw_by_code.get(code_b) or 0
                return (va + vb) if op == "+" else (va - vb)
            return raw_by_code.get(formula)

        for gvaop, formula in _SP_GVAOP_MAP_BS.items():
            val = _sp_formula(formula)
            if val is not None:
                result.gvaop_data[gvaop] = val
        for gvaop, formula in _SP_GVAOP_MAP_IPI.items():
            val = _sp_formula(formula)
            if val is not None:
                result.gvaop_data[gvaop] = val
        # E. Dobiček/izguba rednega poslovanja = EBIT (151-152) + fin.prihodki(153) - fin.odhodki(166)
        # - 3-členska formula, ne uide v preprosto (koda_a, koda_b, op) tabelo zgoraj.
        ebit = (raw_by_code.get("151") or 0) - (raw_by_code.get("152") or 0)
        fin_prih = raw_by_code.get("153")
        fin_odh = raw_by_code.get("166")
        if any(v is not None for v in (raw_by_code.get("151"), raw_by_code.get("152"), fin_prih, fin_odh)):
            result.gvaop_data["067"] = ebit + (fin_prih or 0) - (fin_odh or 0)

    return validate(result)


def _ajpes_native_xls_sheets(rows_by_sheet):
    """Vrne (bs_sheet_name, ipiz_sheet_name) ali (None, None), če format ni prepoznan."""
    bs_name = ipiz_name = None
    for name, rows in rows_by_sheet.items():
        if len(rows) < 3:
            continue
        header = rows[1]
        if not header:
            continue
        h = [str(x).strip() if isinstance(x, str) else x for x in header]
        if len(h) >= 3 and h[0] == "Postavka" and h[2] == "Oznaka za AOP":
            bs_name = name
        elif len(h) >= 4 and h[0] == "Konto" and h[3] == "Oznaka za AOP":
            ipiz_name = name
    return bs_name, ipiz_name

def _ajpes_native_xls_period(book, sheet, row_idx, col_cur, col_prev):
    dm = book.datemode
    def _to_date(v):
        if not isinstance(v, (int, float)) or v <= 0:
            return None
        try:
            return xlrd.xldate_as_datetime(v, dm)
        except Exception:
            return None
    d_cur = _to_date(sheet.cell_value(row_idx, col_cur))
    d_prev = _to_date(sheet.cell_value(row_idx, col_prev))
    return d_cur, d_prev

def parse_ajpes_native_xls(xls_path, tip_override=None, tip_subjekta_override=None):
    book = xlrd.open_workbook(str(xls_path))
    rows_by_sheet = {}
    for sheet in book.sheets():
        rows_by_sheet[sheet.name] = [tuple(sheet.cell_value(r, c) for c in range(sheet.ncols)) for r in range(sheet.nrows)]

    bs_name, ipiz_name = _ajpes_native_xls_sheets(rows_by_sheet)
    result = ParseResult(pdf_format="AJPES_NATIVE_XLS")
    c = CompanyInfo()

    d_cur = d_prev = None
    if bs_name:
        d_cur, d_prev = _ajpes_native_xls_period(book, book.sheet_by_name(bs_name), 2, 3, 4)
    elif ipiz_name:
        d_cur, d_prev = _ajpes_native_xls_period(book, book.sheet_by_name(ipiz_name), 2, 4, 5)
    if d_cur:
        c.period_to = f"{d_cur.day}.{d_cur.month}.{d_cur.year}"
        c.period_from = f"1.1.{d_cur.year}"

    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    result.warnings.append(
        "Ta format ne vsebuje imena podjetja niti matične številke — vpiši ju ročno."
    )

    if bs_name:
        for row in rows_by_sheet[bs_name][3:]:
            if len(row) < 5:
                continue
            aop_raw = row[2]
            aop = str(aop_raw).strip() if isinstance(aop_raw, str) else None
            if not aop or not aop.isdigit():
                continue
            code = aop.zfill(3)
            cur = row[3] if isinstance(row[3], (int, float)) and not isinstance(row[3], bool) else None
            prev = row[4] if isinstance(row[4], (int, float)) and not isinstance(row[4], bool) else None
            if cur is None and prev is None:
                continue
            result.aop_data[code] = AopEntry(aop=code, current_year=cur, previous_year=prev)
    else:
        result.warnings.append("List z bilanco stanja (BS) ni bil najden v datoteki.")

    if ipiz_name:
        for row in rows_by_sheet[ipiz_name][3:]:
            if len(row) < 6:
                continue
            aop_raw = row[3]
            aop = str(aop_raw).strip() if isinstance(aop_raw, str) else None
            if not aop or not aop.isdigit():
                continue
            code = aop.zfill(3)
            cur = row[4] if isinstance(row[4], (int, float)) and not isinstance(row[4], bool) else None
            prev = row[5] if isinstance(row[5], (int, float)) and not isinstance(row[5], bool) else None
            if cur is None and prev is None:
                continue
            result.aop_data[code] = AopEntry(aop=code, current_year=cur, previous_year=prev)
    else:
        result.warnings.append("List z izkazom poslovnega izida (IPIZ) ni bil najden v datoteki.")

    return validate(result)


# ── PARSER: "DENAR | OBRAČUNI" medletni izvoz za gospodarske družbe (.xls, listi     ──
# "BS" / "IU", polni uradni besedilni opisi postavk (brez AOP kolone) z formulskimi   ──
# referencami v oklepaju npr. "(002+032+053)" - to so kode PODREJENIH postavk, ne     ──
# lastna koda vrstice, zato jih pri ujemanju odrežemo). Format nima imena podjetja,    ──
# matične številke niti obdobja - vse tri uporabnik vpiše/izbere ročno (obdobje že     ──
# obstoječe prek UI polja za medletno bilanco (tip_bilance=7), enako kot pri UPR/      ──
# ZAVOD_MEDLETNI/AJPES_NATIVE_XLS). Vrednost je vedno v zadnjem stolpcu vrstice (na    ──
# listu BS je to stolpec G, na listu IU stolpec F - številka stolpca se med izvozi     ──
# lahko razlikuje, zato ga NE hardkodiramo, ampak vzamemo zadnji neprazen stolpec iz    ──
# glave "Naziv" vrstice). Večina postavk (~85 %) se ujame kar prek splošnega           ──
# _classify_and_strip_item + _JOLP_TO_AOP_NORM mehanizma (isti kot za UPR/ZAVOD), ker  ──
# je besedilo zelo blizu uradnemu obrazcu - za preostale (dvojnice/kontekstno odvisne  ──
# postavke npr. domač/EU/izvenEU trg, ter peščica tipkarskih/besednih razlik npr.       ──
# "Finačni" namesto "Finančni", "OD" namesto "IZ" pri kosmatem donosu) imamo spodaj    ──
# neposredne alias tabele in kontekstna pravila (isti vzorec kot pri ZAVOD_MEDLETNI).  ──

# Postavke, ki se pojavijo v treh ločenih blokih (domači trg/EU/izven EU) z ENAKIM
# skrajšanim besedilom - razrešimo jih glede na kontekst zadnje prepoznane nadrejene
# postavke (I./II./III. Čisti prihodki od prodaje na ... trgu = AOP 111/115/118).
_OBRACUNI_AMBIGUOUS_IU = {
    "čisti prihodki od prodaje proizvodov in storitev": {"domaci": "112", "eu": "116", "izven_eu": "119"},
    "čisti prihodki od prodaje blaga in materiala": {"domaci": "114", "eu": "117", "izven_eu": "120"},
}
_OBRACUNI_TRIGGERS_IU = {"111": "domaci", "115": "eu", "118": "izven_eu"}

# Neposredni aliasi za postavke, ki jih splošni JOLP-mehanizem ne ujame (besedne
# razlike, tipkarske napake izvornega programa, ali dvoumnost z drugo postavko z
# enakim skrajšanim besedilom npr. "SREDSTVA" - skupni seštevek 001 proti
# "I. Sredstva (skupina za odtujitev) za prodajo" - 033). Ključi so CELOTNO
# (z oklepaji vred) besedilo postavke, z zbrisanimi večkratnimi presledki, mala črka.
_OBRACUNI_DIRECT_BS = {k.lower(): v for k, v in {
    "SREDSTVA (002+032+053)": "001",
    "6. Opredmetena OS v gradnji in izdelavi": "016",
    "1. Dolgoročne finančne naložbe razen posojil (021 do 023)": "020",
    "I. Sredstva (skupina za odtujitev) za prodajo": "033",
    "1. Kratkoročne finančne naložbe razen posojil (042 do 044)": "041",
    "1. Kratoročne poslovne terjatve do družb v skupini": "049",
    "Zabilančna sredstva": "054",
    "3. Lastne delnice in lastne poslovne deleže (kot odbitna postavka)": "064",
    "IV. Presežek iz prevrednotenja": "067",
    "1. Rezervacije": "073",  # JOLP_TO_AOP ima "1. Rezervacije" dvakrat (BS:073, IPI:149) -
                              # drugi zapis povozi prvega, zato ga tu eksplicitno popravimo za BS
    "Zabilančne obveznosti": "096",
}.items()}

_OBRACUNI_DIRECT_IU = {k.lower(): v for k, v in {
    "F. KOSMATI DONOS IZ POSLOVANJA": "126",
    "1. Nabavna vrednost prodanega blaga in materiala": "129",
    "c) povračila stroškov zaposlencem v zvezi z delom": "137",
    "2. Prevred. posl. odh. pri neopr.OS in oprdm.OS": "146",
    "3. Prevrednotovalni poslovni odhodki pri obrat. sr.": "147",
    "1. Finančni prihodki iz posojil, danim družbam v skupini": "161",
    "K. FINAČNI ODHODKI": "166",
    "Finančni odhodki od obresti (upoštevano že v II in III) del AOP 169 + AOP 170": "167",
    "II. Finačni odhodki iz finančnih obveznosti": "169",
    "3. Drugi finančni odhodki iz izdanih obveznic": "172",
    "III. Finačni odhodki iz poslovnih obveznosti": "174",
    "II. Drugi finančni prihodki in ostali prihodki": "180",
    "P. DAVEK IZ DOBIČKA": "184",
    "R. ODLOŽENI DAVKI": "185",
    "Š. ČISTI IZGUBA OBRAČUNSKEGA OBDOBJA": "187",
    "Povprečno število zaposlencev na podlagi delovnih ur v obračunskem obdobju": "188",
}.items()}


def _match_obracuni_label(label, direct_map, ambiguous_map, context):
    norm_label = re.sub(r'\s+', ' ', str(label).strip())
    raw_key = norm_label.lower()
    if raw_key in direct_map:
        return direct_map[raw_key]
    kind, text = _classify_and_strip_item(norm_label)
    opts = ambiguous_map.get(text)
    if opts:
        # Dvoumno besedilo NIMA smiselnega splošnega ujemanja (isto besedilo se
        # ponovi v več kontekstih z RAZLIČNIMI AOP kodami) - razrešimo izključno
        # prek konteksta, tudi če bi splošni JOLP mehanizem "po sreči" ujel eno
        # od variant, da ne pride do napačnega prepisovanja pri drugih variantah.
        return opts.get(context)
    return _JOLP_TO_AOP_NORM.get((kind, text))


def _is_obracuni_medletni_xls(rows_by_sheet):
    """Vrne (bs_sheet_name, iu_sheet_name) ali (None, None), če format ni prepoznan.
    Prepozna se po glavi 'DENAR | OBRAČUNI' + 'BILANCA STANJA' / 'PODATKI IZ IZKAZA
    POSLOVNEGA IZIDA' v prvih dveh vrsticah lista."""
    bs_name = iu_name = None
    for name, rows in rows_by_sheet.items():
        if len(rows) < 5:
            continue
        r0 = str(rows[0][0]).strip() if rows[0] and rows[0][0] else ""
        r1 = str(rows[1][0]).strip() if len(rows) > 1 and rows[1][0] else ""
        if r0 != "DENAR | OBRAČUNI":
            continue
        if "BILANCA STANJA" in r1:
            bs_name = name
        elif "IZKAZA POSLOVNEGA IZIDA" in r1:
            iu_name = name
    return bs_name, iu_name


def parse_obracuni_medletni_xls(xls_path, tip_override=None, tip_subjekta_override=None):
    book = xlrd.open_workbook(str(xls_path))
    rows_by_sheet = {}
    for sheet in book.sheets():
        rows_by_sheet[sheet.name] = [tuple(sheet.cell_value(r, c) for c in range(sheet.ncols)) for r in range(sheet.nrows)]

    bs_name, iu_name = _is_obracuni_medletni_xls(rows_by_sheet)
    result = ParseResult(pdf_format="OBRACUNI_MEDLETNI")

    c = CompanyInfo()
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    result.warnings.append(
        "Ta format (DENAR/OBRAČUNI medletni izvoz) ne vsebuje imena podjetja, matične "
        "številke niti obdobja — vpiši/izberi jih ročno."
    )

    def _parse_sheet(name, direct_map, ambiguous_map, triggers):
        aop_data = {}
        rows = rows_by_sheet[name]
        if len(rows) < 6:
            return aop_data
        ncols = len(rows[4]) if rows[4] else len(rows[5])
        val_col = ncols - 1
        context = None
        for row in rows[5:]:
            if not row or not row[0] or not str(row[0]).strip():
                continue
            label = str(row[0])
            val = row[val_col] if len(row) > val_col else None
            aop = _match_obracuni_label(label, direct_map, ambiguous_map, context)
            if aop and aop in triggers:
                context = triggers[aop]
            if not aop or not isinstance(val, (int, float)) or isinstance(val, bool):
                continue
            aop_data[aop] = AopEntry(aop=aop, current_year=float(val), previous_year=None)
        return aop_data

    if bs_name:
        result.aop_data.update(_parse_sheet(bs_name, _OBRACUNI_DIRECT_BS, _ZAVOD_AMBIGUOUS_BS, _ZAVOD_CONTEXT_TRIGGERS_BS))
    else:
        result.warnings.append("List z bilanco stanja (BS) ni bil najden v datoteki.")

    if iu_name:
        result.aop_data.update(_parse_sheet(iu_name, _OBRACUNI_DIRECT_IU, _OBRACUNI_AMBIGUOUS_IU, _OBRACUNI_TRIGGERS_IU))
    else:
        result.warnings.append("List z izkazom poslovnega izida (IU) ni bil najden v datoteki.")

    return validate(result)


def parse_xls_file(xls_path, tip_override=None, tip_subjekta_override=None):
    if xlrd is None:
        raise RuntimeError("Za uvoz .xls datotek manjka knjižnica xlrd. Dodaj xlrd==2.0.1 v requirements.txt in redeployaj aplikacijo.")
    book = xlrd.open_workbook(str(xls_path))
    rows_by_sheet = {}
    for sheet in book.sheets():
        rows = []
        for r in range(sheet.nrows):
            rows.append(tuple(sheet.cell_value(r, c) for c in range(sheet.ncols)))
        rows_by_sheet[sheet.name] = rows

    letinvest_sheets = _find_letinvest_xls_sheets(rows_by_sheet)
    if letinvest_sheets:
        return parse_letinvest_xls(xls_path, tip_override, tip_subjekta_override)

    native_sheets = _find_native_uradni_xls_sheets(rows_by_sheet)
    if native_sheets:
        return parse_native_uradni_xls(xls_path, tip_override, tip_subjekta_override)

    bs_name, ipiz_name = _ajpes_native_xls_sheets(rows_by_sheet)
    if bs_name or ipiz_name:
        return parse_ajpes_native_xls(xls_path, tip_override, tip_subjekta_override)

    obr_bs, obr_iu = _is_obracuni_medletni_xls(rows_by_sheet)
    if obr_bs or obr_iu:
        return parse_obracuni_medletni_xls(xls_path, tip_override, tip_subjekta_override)

    for rows in rows_by_sheet.values():
        if _is_porocilo_gd_xls(rows):
            return parse_porocilo_gd_xls(rows, tip_override, tip_subjekta_override)

    return parse_excel_rows(rows_by_sheet, tip_override, tip_subjekta_override, "XLS")


# ── PARSER: AOPCOL format (ima eksplicitno AOP kolono: "Postavka AOP Tekoče leto Preteklo leto") ──
# Pri tem formatu se pri daljših postavkah besedilo postavke VIZUALNO PREKRIVA z AOP kodo in zneski
# (napaka izvornega poročevalskega orodja - npr. 'PROIZVODNJE' in koda '121' sta narisana na isti
# poziciji). Ker ju pdfplumber združi v en sam "besedni" niz (npr. 'PROIZVODN1J2E1'), jih ločimo na
# nivoju POSAMEZNIH ZNAKOV, in sicer po VRSTNEM REDU RISANJA (ne po x-poziciji): AOP koda in zneski
# so vedno narisani kot LOČENI nizi znakov, zato x-pozicija znotraj enega niza vedno rjaste (razen
# ob meji dveh nizov, kjer x nazaduje ali naredi velik skok naprej).
AOPCOL_CODE_X_MIN = 365
AOPCOL_CODE_X_MAX = 410

def _aopcol_split_runs(rowchars, backward_eps=0.5, gap_thresh=15):
    runs, cur, prev_x1 = [], [], None
    for c in rowchars:
        if prev_x1 is not None and (c['x0'] < prev_x1 - backward_eps or (c['x0'] - prev_x1) > gap_thresh):
            runs.append(cur); cur = []
        cur.append(c)
        prev_x1 = c['x1']
    if cur: runs.append(cur)
    return runs

def _aopcol_extract_triples(chars):
    """Vrne seznam (aop_koda, [zn1_text, zn2_text]) — po eno postavko na vrstico z AOP kodo."""
    rows = defaultdict(list)
    for c in chars:
        rows[round(c['top'], 1)].append(c)
    results = []
    for top in sorted(rows.keys()):
        rowchars = rows[top]  # NAMENOMA ne sortiramo po x0 - vrstni red risanja nosi informacijo
        runs = _aopcol_split_runs(rowchars)
        code_idx = None
        for ridx, run in enumerate(runs):
            txt = ''.join(ch['text'] for ch in run).strip()
            x0 = run[0]['x0']
            if txt.isdigit() and 2 <= len(txt) <= 3 and AOPCOL_CODE_X_MIN <= x0 <= AOPCOL_CODE_X_MAX:
                code_idx = ridx
                break
        if code_idx is None:
            continue
        code_text = ''.join(ch['text'] for ch in runs[code_idx]).strip()
        remaining = runs[:code_idx] + runs[code_idx+1:]
        val_runs = remaining[-2:] if len(remaining) >= 2 else remaining
        val_texts = [''.join(ch['text'] for ch in r).strip() for r in val_runs]
        results.append((code_text, val_texts))
    return results

# ── PARSER: URADNI format (čist uraden AJPES obrazec, "Oznaka za AOP" + zaporedje zneskov) ──
# Ta obrazec ima ČISTO besedilo (brez prekrivanja/kerning napak kot pri AOPCOL). Obstajata dve
# različici zapisa vsake postavke, odvisno od izvornega orodja stranke:
#   1) "AOP_KODA [B|P]AOP_KODA ZNESEK_TEKOČE ZNESEK_PRETEKLO"  (podvojena koda, npr. '001 B001 ...')
#   2) "AOP_KODA ZNESEK_TEKOČE ZNESEK_PRETEKLO"                 (samo enkratna koda, npr. '001 7.199...')
_URADNI_ENTRY_RE_DUP    = re.compile(r'(\d{3})\s+[BP](\d{3})\s+(-?[\d.,]+)\s+(-?[\d.,]+)')
_URADNI_ENTRY_RE_SINGLE = re.compile(r'(?<!\d)(\d{3})\s+(-?[\d.]+,\d{2})\s+(-?[\d.]+,\d{2})')

def _uradni_extract_words(pages_words, code_x=(320, 383), val1_x=(383, 470), val2_x=(470, 600), tol=2.0):
    """Poišče (AOP koda -> (tekoče, preteklo)) po poziciji besed na strani - zanesljivo tudi kadar
    je prisoten samo en od dveh stolpcev zneskov, ali kadar je opis postavke v besedilnem toku
    ločen od AOP kode/zneskov (kar podre navaden regex na sklenjenem besedilu)."""
    entries = {}
    for words in pages_words:
        rows = []
        for w in sorted(words, key=lambda w: w['top']):
            placed = False
            for row in rows:
                if abs(row[0]['top'] - w['top']) <= tol:
                    row.append(w); placed = True; break
            if not placed:
                rows.append([w])
        for ws in rows:
            code_txt = None
            v1_words, v2_words = [], []
            for w in sorted(ws, key=lambda x: x['x0']):
                x0, t = w['x0'], w['text']
                if code_x[0] <= x0 < code_x[1]:
                    mm = re.fullmatch(r'[BP]?(\d{3})', t)
                    if mm and code_txt is None:
                        code_txt = mm.group(1)
                elif val1_x[0] <= x0 < val1_x[1]:
                    v1_words.append(t)
                elif val2_x[0] <= x0 < val2_x[1]:
                    v2_words.append(t)
            if code_txt and code_txt not in entries:
                v1 = ''.join(v1_words) if v1_words else None
                v2 = ''.join(v2_words) if v2_words else None
                entries[code_txt] = (v1, v2)
    return entries

def parse_uradni_format(pages_text, pages_words, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="URADNI")
    full = "\n".join(pages_text)
    c = CompanyInfo()

    m = re.search(r'Matična številka:\s*(\d+)', full)
    if m and m.group(1).strip() != '0':
        c.registration_number = m.group(1).strip()

    m = re.search(r'Ime poslovnega subjekta.*?\n(.+?)\s+(\d{7,10})\s*\n', full)
    if m:
        c.name = m.group(1).strip()
        if not c.registration_number:
            c.registration_number = m.group(2).strip()
    else:
        # Poskusi izluščiti ime iz vrstice takoj po "Ime poslovnega subjekta", tudi če se
        # dvostolpčna postavitev PDF-ja pomeša z desnim stolpcem na isti vrstici (npr.
        # "MEDITERAN PRODUKT Statusna sprememba: 0") - odreži pri znanem naslovu desnega stolpca.
        m = re.search(r'Ime poslovnega subjekta\b.*\n(.+)', full)
        if m:
            name_line = m.group(1)
            name_line = re.split(
                r'\s+(?:Statusna sprememba|Dav[cč]na [sš]tevilka|Mati[cč]na [sš]tevilka|Velikost)\b',
                name_line
            )[0]
            name_line = name_line.strip()
            # Dvostolpčna postavitev PDF-ja: vrednost polja "Velikost" (ena števka 1-4)
            # se pri nekaterih izvozih znajde na ISTI vrstici tik za imenom podjetja,
            # brez vmesne oznake "Velikost" (ta je izpisana ločeno, drugje v besedilu) -
            # npr. "24INVEST D.O.O. 1" namesto samo "24INVEST D.O.O." - odrežemo jo.
            name_line = re.sub(r'\s+[1-4]$', '', name_line).strip()
            if name_line and not re.fullmatch(r'0+', name_line):
                c.name = name_line
        if not c.name:
            m = re.search(r'\n([A-ZČŠŽ][A-Za-zČŠŽčšž\. ]*?[Dd]\.?[Oo]\.?[Oo]\.?)\s', full)
            if not m:
                m = re.search(r'\n([A-ZČŠŽ][A-Za-zČŠŽčšž\. ]*?[Dd]\.?[Dd]\.?)\s', full)
            if m: c.name = m.group(1).strip()
        if not c.registration_number:
            m2 = re.search(r'(\d{7,10})\s*\nDavčna številka', full)
            if m2: c.registration_number = m2.group(1).strip()

    m = re.search(r'Sedež poslovnega subjekta.*?\n.+?(\d{7,10})\s*\n', full)
    if m:
        c.tax_number = m.group(1).strip()
    else:
        m = re.search(r'Davčna številka:?\s*\n?(\d+)', full)
        if m and m.group(1).strip() != '0':
            c.tax_number = m.group(1).strip()

    is_ipi = "izkaza poslovnega izida" in full.lower()
    m = re.search(r'v obdobju\s*\n?((?:\d\s*){1,2}\.\s*(?:\d\s*){1,2}\.\s*(?:\d\s*){4})\s*-\s*((?:\d\s*){1,2}\.\s*(?:\d\s*){1,2}\.\s*(?:\d\s*){4})', full)
    if m:
        c.period_from = re.sub(r'\s+', '', m.group(1))
        c.period_to   = re.sub(r'\s+', '', m.group(2))
    else:
        # Nekateri izvozi namesto "X - Y" pišejo "v obdobju od X do Y"
        m = re.search(r'v obdobju\s+od\s*\n?.*?do\s*\n?((?:\d\s*){1,2}\.\s*(?:\d\s*){1,2}\.\s*(?:\d\s*){4})', full)
        if m:
            c.period_to = re.sub(r'\s+', '', m.group(1))
            leto = c.period_to.split('.')[-1]
            c.period_from = f"1.1.{leto}"
        else:
            m = re.search(r'na dan\s*\n?((?:\d\s*){1,2}\.\s*(?:\d\s*){1,2}\.\s*(?:\d\s*){4})', full)
            if m: c.period_to = re.sub(r'\s+', '', m.group(1))

    if (not c.period_from or len(c.period_from) < 6) and c.period_to:
        leto = c.period_to.split('.')[-1]
        c.period_from = f"1.1.{leto}"

    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    if not result.company.registration_number:
        result.warnings.append(
            "Matična številka v izvoru ni podana (npr. konsolidirano poročilo brez ene uradne "
            "matične številke) — vpiši jo ročno v celico D4 pred oddajo."
        )

    for code, (v1, v2) in _uradni_extract_words(pages_words).items():
        if code in result.aop_data:
            continue
        current  = parse_si(v1) if v1 else None
        previous = parse_si(v2) if v2 else None
        if current is not None or previous is not None:
            result.aop_data[code] = AopEntry(aop=code, current_year=current, previous_year=previous)

    # Regex-na dopolnitev za morebitne postavke, ki jih pozicijska metoda zgreši (npr. zaradi
    # nenavadne poravnave), samo za kode, ki jih zgornja metoda ni že našla.
    dup_matches = list(_URADNI_ENTRY_RE_DUP.finditer(full))
    for m in dup_matches:
        code, code2, val1, val2 = m.groups()
        if code != code2 or code in result.aop_data:
            continue
        current  = parse_si(val1)
        previous = parse_si(val2)
        if current is not None or previous is not None:
            result.aop_data[code] = AopEntry(aop=code, current_year=current, previous_year=previous)
    for m in _URADNI_ENTRY_RE_SINGLE.finditer(full):
        code, val1, val2 = m.groups()
        if code in result.aop_data:
            continue
        current  = parse_si(val1)
        previous = parse_si(val2)
        if current is not None or previous is not None:
            result.aop_data[code] = AopEntry(aop=code, current_year=current, previous_year=previous)
    return validate(result)

def parse_aopcol_format(pages_text, pages_chars, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="AOPCOL")
    full = "\n".join(pages_text)
    first = pages_text[0] if pages_text else ""
    c = CompanyInfo()

    m = re.search(r'\n(.+?)\nMatična številka:', full)
    if m: c.name = m.group(1).strip()
    m = re.search(r'Matična številka:\s*\n?(\d+)', full)
    if m: c.registration_number = m.group(1).strip()
    m = re.search(r'Davčna številka:\s*\n?(\d+)', full)
    if m: c.tax_number = m.group(1).strip()

    is_ipi = "IZKAZ POSLOVNEGA IZIDA" in first
    if is_ipi:
        m = re.search(r'IZKAZ POSLOVNEGA IZIDA\s*\n(\d{1,2}\.\d{1,2}\.\d{4})\s*-\s*(\d{1,2}\.\d{1,2}\.\d{4})', full)
        if m:
            c.period_from = m.group(1).strip()
            c.period_to   = m.group(2).strip()
    else:
        m = re.search(r'BILANCA STANJA\s*\n(\d{1,2}\.\d{1,2}\.\d{4})', full)
        if m: c.period_to = m.group(1).strip()

    if (not c.period_from or len(c.period_from) < 6) and c.period_to:
        leto = c.period_to.split('.')[-1]
        c.period_from = f"1.1.{leto}"

    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)

    for chars in pages_chars:
        for code, val_texts in _aopcol_extract_triples(chars):
            if code in result.aop_data:
                continue
            current  = parse_si(val_texts[0]) if len(val_texts) > 0 else None
            previous = parse_si(val_texts[1]) if len(val_texts) > 1 else None
            if current is not None or previous is not None:
                result.aop_data[code] = AopEntry(aop=code, current_year=current, previous_year=previous)
    return result

# ── PARSER: OCR fallback (za PDFje s pokvarjeno pisavo - ni ToUnicode preslikave, torej ────
# besedilnega sloja ni mogoče prebrati z NOBENO knjižnico - preverjeno s pdfplumber IN PyMuPDF).
# Tabela ima zaznavne mrežne črte (vektorska grafika), zato lahko z gotovostjo najdemo meje celic
# tudi brez berljivega besedila, nato pa vsako celico (AOP koda, znesek tekoče, znesek preteklo)
# OCR-amo LOČENO. Postavko sprejmemo SAMO če se AOP koda IN oba zneska hkrati uspešno prepoznajo
# iz iste vrstice tabele - kadarkoli en del manjka, vrstico raje izpustimo (ostane 0 v izvozu),
# kot da bi tvegali napačno št. na napačni AOP kodi (za bančno poročanje je to pomembneje kot
# popolnost). Realno pokrijemo cca 40-70% postavk odvisno od kvalitete PDF-ja; preostale je treba
# vnesti ročno - seznam manjkajočih kod je v opozorilih.
_OCR_VAL_RE = re.compile(r'^-?[\d.]*\d,\d{2}$')

def _ocr_cell_text(im, bbox, scale, config):
    x0, top, x1, bottom = bbox
    crop = im.crop((x0*scale, top*scale, x1*scale, bottom*scale))
    data = pytesseract.image_to_data(crop, config=config, output_type=pytesseract.Output.DICT)
    texts = [t for t in data['text'] if t.strip()]
    confs = [int(c) for c, t in zip(data['conf'], data['text']) if t.strip() and str(c).lstrip('-').isdigit()]
    txt = ''.join(texts).strip() if len(texts) <= 1 else ' '.join(texts).strip()
    avgconf = sum(confs)/len(confs) if confs else -1.0
    return txt, avgconf

_OCR_CFG_NUM  = r'--psm 7 -c tessedit_char_whitelist=0123456789.,-'
_OCR_CFG_CODE = r'--psm 7 -c tessedit_char_whitelist=0123456789'

_OCR_MIN_CONF = 55  # pod tem pragom OCR raje izpustimo kot tvegamo napačno števko (npr. 0<->9 zamenjava)

def _ocr_extract_page_entries(page, dpi=300):
    """Vrne dict {aop_koda: (current_year, previous_year)} za eno stran, samo za vrstice kjer
    smo hkrati zanesljivo (z dovolj visokim OCR zaupanjem) prebrali kodo in oba zneska."""
    entries = {}
    tables = page.find_tables()
    if not tables:
        return entries
    t = tables[0]
    scale = dpi / 72.0
    im = page.to_image(resolution=dpi).original
    for row in t.rows:
        cells = row.cells
        if len(cells) < 5 or cells[2] is None or cells[3] is None or cells[4] is None:
            continue
        code_txt, code_conf = _ocr_cell_text(im, cells[2], scale, _OCR_CFG_CODE)
        v1, v1_conf = _ocr_cell_text(im, cells[3], scale, _OCR_CFG_NUM)
        v2, v2_conf = _ocr_cell_text(im, cells[4], scale, _OCR_CFG_NUM)
        code_txt = code_txt.strip()
        if code_txt.isdigit():
            code_txt = code_txt.zfill(3)
        if not re.fullmatch(r'\d{3}', code_txt):
            continue
        if not (1 <= int(code_txt) <= 320):
            continue
        if code_txt in entries:
            continue
        if not (_OCR_VAL_RE.match(v1) and _OCR_VAL_RE.match(v2)):
            continue
        # Nizko zaupanje = velika verjetnost zamenjane števke (npr. 0<->9) - raje izpusti.
        if min(code_conf, v1_conf, v2_conf) < _OCR_MIN_CONF:
            continue
        entries[code_txt] = (parse_si(v1), parse_si(v2))
    return entries

# ── PARSER: OCR za SKENIRANE PDFje s čisto "Postavka | AOP | Tekoče leto | Preteklo leto" tabelo ──
# Za razliko od parse_ocr_fallback_format (ki uporablja zaznane mrežne črte tabele) ta skenirani
# PDF nima NOBENE vektorske grafike (samo rastrska slika), zato mej celic ne moremo zaznati.
# Namesto tega besede iz OCR razporedimo v vrstice po Y-poziciji (s toleranco), nato v stolpce po
# X-poziciji (kot delež širine slike, da je neodvisno od DPI/resolucije).
_OCR_SCAN_COL_FRACS = {"aop": (0.55, 0.72), "current": (0.72, 0.86), "previous": (0.86, 1.0)}

def _ocr_preprocess(im, threshold=None, denoise=False):
    """Sivinjenje + avtomatski kontrast pred OCR - pri skeniranih/fotokopiranih obrazcih (slabša
    kakovost, madeži, neenakomerna osvetlitev) to pomaga tesseractu brati številke natančneje kot
    iz surove barvne slike. Trdo binarizacijo (threshold) namenoma ne uporabljamo privzeto - v
    testih je na te obrazce slabše vplivala kot Tesseractova lastna (Otsu) binarizacija.
    denoise=True doda medianski filter PRED avtokontrastom - nekateri uradni FRP/bančni obrazci
    imajo vrstice seštevkov (skupaj/podskupaj) podložene s finim pikčastim rastrskim vzorcem
    (poudarek/highlight), ki brez tega filtra ustvari na desetine lažnih drobnih "besed" v OCR
    izhodu - te zmedejo razporejanje besedila v vrstice (več resničnih vrstic se zlepi v eno) IN
    poslabšajo branje samih številk. Medianski filter to piko odstrani, črke/številke pa ohrani."""
    gray = ImageOps.grayscale(im)
    if denoise and ImageFilter is not None:
        gray = gray.filter(ImageFilter.MedianFilter(3))
    auto = ImageOps.autocontrast(gray, cutoff=1)
    if threshold is None:
        return auto
    return auto.point(lambda p: 255 if p > threshold else 0)

def _ocr_words_from_data(data):
    words = []
    for i in range(len(data['text'])):
        t = data['text'][i].strip()
        if t:
            words.append({'text': t, 'left': data['left'][i], 'top': data['top'][i]})
    return words

def _ocr_words_to_lines(words, row_tol=8):
    """Združi posamezne OCR besede nazaj v vizualne vrstice (po Y-poziciji), besede znotraj
    vrstice pa uredi po X-poziciji - da dobimo berljivo besedilo, primerljivo s
    pytesseract.image_to_string(), namesto da bi vsaka beseda pristala v svoji vrstici."""
    rows = []
    for w in sorted(words, key=lambda x: x['top']):
        placed = False
        for row in rows:
            if abs(row[0]['top'] - w['top']) <= row_tol:
                row.append(w); placed = True; break
        if not placed:
            rows.append([w])
    return [' '.join(w['text'] for w in sorted(row, key=lambda x: x['left'])) for row in rows]

def _ocr_detect_col_bounds(words, width):
    """Poišče x-pozicije glave stolpcev ('AOP', 'Tekočega', 'Prejšnjega') na strani, ki jih ima
    (glava se navadno pojavi samo na prvi strani vsakega poročila), in iz njih izračuna meje
    stolpcev kot delež širine strani. Vrne None, če glave na tej strani ni (uporabi privzete)."""
    def _find(pattern):
        for w in words:
            if re.search(pattern, w['text'], re.I):
                return w
        return None
    w_aop = _find(r'^AOP$') or _find(r'Oznaka')
    w_cur = _find(r'Teko[cč]ega')
    w_prev = _find(r'rej.{0,4}njega') or _find(r'Prej[sš]njega')
    if not (w_aop and w_cur and w_prev):
        return None
    aop_start = (w_aop['left'] - 40) / width
    aop_end   = (w_cur['left'] - 20) / width
    cur_end   = (w_prev['left'] - 20) / width
    return {"aop": (max(0, aop_start), aop_end), "current": (aop_end, cur_end), "previous": (cur_end, 1.0)}

def _ocr_scan_page_entries(data, width, col_fracs, row_tol=8):
    words = _ocr_words_from_data(data)
    words.sort(key=lambda w: w['top'])
    rows = []
    for w in words:
        placed = False
        for row in rows:
            if abs(row[0]['top'] - w['top']) <= row_tol:
                row.append(w); placed = True; break
        if not placed:
            rows.append([w])

    aop_x = (col_fracs["aop"][0]*width, col_fracs["aop"][1]*width)
    cur_x = (col_fracs["current"][0]*width, col_fracs["current"][1]*width)
    prev_x = (col_fracs["previous"][0]*width, col_fracs["previous"][1]*width)

    entries = {}
    for ws in rows:
        aop_txt = None
        cur_words, prev_words = [], []
        for w in sorted(ws, key=lambda x: x['left']):
            x, t = w['left'], w['text']
            if aop_x[0] <= x < aop_x[1] and re.fullmatch(r'\d{3}', t):
                aop_txt = t
            elif cur_x[0] <= x < cur_x[1]:
                cur_words.append(w)
            elif prev_x[0] <= x < prev_x[1]:
                prev_words.append(w)
        if not aop_txt:
            continue
        cur_txt  = ''.join(w['text'] for w in sorted(cur_words, key=lambda x: x['left'])) if cur_words else None
        prev_txt = ''.join(w['text'] for w in sorted(prev_words, key=lambda x: x['left'])) if prev_words else None
        entries[aop_txt] = (cur_txt, prev_txt)
    return entries

def _ocr_scan_clean_number(s):
    if s is None: return None
    s = s.strip().strip(',.').strip()
    return parse_si(s)

_OCR_VAL_TOKEN_RE = re.compile(r'\d[\d.,]*\d|\d')

def _ocr_scan_page_entries_percropped(page, dpi, col_fracs, header_row_words):
    """Vrstice najprej razmeji po Y-poziciji besed iz celostranskega OCR prehoda, nato pa za
    VSAKO vrstico posebej izreže samo desni del strani (AOP + zneska), ga 3x poveča in šele
    nato OCR-a s seznamom dovoljenih znakov samo za števke. Izrezana, povečana in kontrastno
    izboljšana slika posamezne vrstice se izkaže za bistveno bolj berljivo kot ista vrstica
    znotraj OCR-a cele strani naenkrat - zato ta pristop najde občutno več postavk na slabše
    kakovostnih skenih. Vrstice se obdelajo vzporedno (klic tesseract binarnega procesa sprosti
    GIL), kar znatno pohitri obdelavo večstranskih dokumentov."""
    im_raw = page.to_image(resolution=dpi).original
    width, height = im_raw.size

    quick = pytesseract.image_to_data(_ocr_preprocess(im_raw, denoise=True), lang='slv+eng',
                                       output_type=pytesseract.Output.DICT, config='--psm 6')
    words = _ocr_words_from_data(quick)
    words = [w for w in words if w['top'] > 900]
    words.sort(key=lambda w: w['top'])
    row_tops = []
    for w in words:
        placed = False
        for rt in row_tops:
            if abs(rt[0] - w['top']) <= 12:
                rt[1] = min(rt[1], w['top']); rt[2] = max(rt[2], w['top']); placed = True; break
        if not placed:
            row_tops.append([w['top'], w['top'], w['top']])
    row_tops.sort(key=lambda rt: rt[0])
    merged = []
    for rt in row_tops:
        if merged and rt[1] - merged[-1][2] < 15:
            merged[-1][2] = max(merged[-1][2], rt[2])
        else:
            merged.append(list(rt))

    aop_frac_start = col_fracs["aop"][0] if col_fracs else 0.55
    crop_span = 1.0 - aop_frac_start
    cur_lo = (col_fracs["current"][0] - aop_frac_start) / crop_span if col_fracs else 0.3
    prev_lo = (col_fracs["previous"][0] - aop_frac_start) / crop_span if col_fracs else 0.65

    def _process(rt):
        top_min, top_max = rt[1], rt[2]
        top = max(0, top_min - 6)
        bot = min(height, top_max + 32)
        if bot - top < 8 or bot - top > 80:
            return None
        crop = im_raw.crop((int(aop_frac_start * width), top, width, bot))
        if crop.width < 10:
            return None
        crop_big = crop.resize((crop.width * 3, crop.height * 3))
        proc = _ocr_preprocess(crop_big, denoise=True)
        data = pytesseract.image_to_data(
            proc, config='--psm 11 -c tessedit_char_whitelist=0123456789.,',
            output_type=pytesseract.Output.DICT
        )
        crop_width = crop_big.size[0]
        toks = []
        for i in range(len(data['text'])):
            t = data['text'][i].strip()
            if t:
                toks.append({'text': t, 'left': data['left'][i]})
        aop_txt = None
        for tk in toks:
            if re.fullmatch(r'\d{3}', tk['text']):
                aop_txt = tk['text']
                break
        if not aop_txt or not (1 <= int(aop_txt) <= 320):
            return None
        _NUM_RE = re.compile(r'^\d{1,3}(\.\d{3})*,\d{2}$|^\d+,\d{2}$')
        cur_txt = prev_txt = None
        for tk in toks:
            if tk['text'] == aop_txt or not _NUM_RE.match(tk['text']):
                continue
            frac = tk['left'] / crop_width
            if frac >= prev_lo and prev_txt is None:
                prev_txt = tk['text']
            elif frac < prev_lo and cur_txt is None:
                cur_txt = tk['text']
        return (aop_txt, cur_txt, prev_txt)

    from concurrent.futures import ThreadPoolExecutor
    entries = {}
    with ThreadPoolExecutor(max_workers=6) as ex:
        for res in ex.map(_process, merged):
            if not res:
                continue
            aop_txt, cur_txt, prev_txt = res
            if aop_txt in entries:
                continue  # prva (zgornja) zaznava zmaga - podvojene vrstice iz merge koraka
            entries[aop_txt] = (cur_txt, prev_txt)
    return entries

def parse_ocr_scanned_aopcol_format(pdf_path, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="OCR_SCAN")
    if pytesseract is None:
        result.errors.append(
            "Ta PDF je skeniran (nima besedilnega sloja) in bi ga bilo treba brati z OCR, "
            "a knjižnica pytesseract ni nameščena na strežniku."
        )
        return result

    with pdfplumber.open(pdf_path) as pdf:
        header_ocr = ""
        header_words = []
        entries_by_code = {}
        col_fracs = None
        page_data = []  # (width, data, is_header_candidate)

        # En sam OCR prehod na stran (predobdelana slika) - iz istega rezultata izluščimo
        # tako besedilo glave (ime/matična/obdobje) kot podatke tabele, namesto da bi za
        # isto stran klicali pytesseract tri- ali večkrat.
        for pi, page in enumerate(pdf.pages):
            im_raw = page.to_image(resolution=300).original
            im = _ocr_preprocess(im_raw)
            data = pytesseract.image_to_data(im, lang='slv+eng', output_type=pytesseract.Output.DICT, config='--psm 6')
            width = im.size[0]
            page_data.append((width, data))
            if pi < 2:
                header_ocr += "\n" + "\n".join(_ocr_words_to_lines(_ocr_words_from_data(data)))
                header_words.extend(_ocr_words_from_data(data))
            if col_fracs is None:
                col_fracs = _ocr_detect_col_bounds(_ocr_words_from_data(data), width)

        if col_fracs is None:
            col_fracs = _OCR_SCAN_COL_FRACS

        # Natančnejše (vrstica-po-vrstica izrezana in povečana) branje tabelnih podatkov -
        # počasnejše od enega samega OCR prehoda čez celo stran, a bistveno bolj zanesljivo
        # na slabše kakovostnih skenih/fotokopijah, kjer so majhne 3-mestne AOP kode in zneski
        # v celostranskem OCR-u pogosto zgrešeni ali napačno prebrani.
        for page in pdf.pages:
            for code, vals in _ocr_scan_page_entries_percropped(page, 300, col_fracs, header_words).items():
                if code not in entries_by_code:
                    entries_by_code[code] = vals

    c = CompanyInfo()
    lines = header_ocr.split('\n')
    for i, line in enumerate(lines):
        if 'Naziv organizacije' in line:
            for cand in lines[i+1:i+4]:
                cand = cand.strip()
                if cand and 'tevilk' not in cand.lower():
                    c.name = re.sub(r'\s+', ' ', cand).strip()
                    break
            break
    if not c.name:
        # AJPES-jev obrazec: ime je v vrstici pod "Ime poslovnega subjekta"
        for i, line in enumerate(lines):
            if 'poslovnega subjekta' in line.lower() and 'ime' in line.lower():
                for cand in lines[i+1:i+3]:
                    cand = cand.strip()
                    if cand and 'sede' not in cand.lower() and len(cand) > 2:
                        c.name = re.sub(r'\s+', ' ', cand).strip()
                        break
                break

    # Matična/davčna sta v desnem stolpcu dvostolpčne glave - navaden OCR tekst jih pogosto
    # zgreši/premeša, zato ju poiščemo po poziciji besed ("Številka:" oznaka + število na isti vrstici).
    label_hits = sorted(
        [w for w in header_words if re.search(r'[SŠ]tevilk', w['text'], re.I)],
        key=lambda w: w['top']
    )
    found_numbers = []
    for lw in label_hits[:2]:
        for w in header_words:
            if abs(w['top'] - lw['top']) <= 10 and w['left'] > lw['left'] and re.fullmatch(r'\d{6,10}', w['text']):
                found_numbers.append(w['text'])
                break
    if len(found_numbers) >= 1: c.tax_number = found_numbers[0]
    if len(found_numbers) >= 2: c.registration_number = found_numbers[1]

    m = re.search(r'(\d{1,2}\.\d{1,2}\.\d{4})\s*-\s*(\d{1,2}\.\d{1,2}\.\d{4})', header_ocr)
    if m:
        c.period_from, c.period_to = m.group(1).strip(), m.group(2).strip()
    else:
        m = re.search(r'\n(\d{1,2}\.\d{1,2}\.\d{4})\s*\n', header_ocr)
        if m: c.period_to = m.group(1).strip()
    if (not c.period_from or len(c.period_from) < 6) and c.period_to:
        leto = c.period_to.split('.')[-1]
        c.period_from = f"1.1.{leto}"

    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    if not result.company.name or not result.company.registration_number:
        result.warnings.append("Ime podjetja in/ali matična številka niso bili zanesljivo prepoznani z OCR — prosim preveri/vpiši ročno v D3/D4.")

    for code, (cur_txt, prev_txt) in entries_by_code.items():
        if not (1 <= int(code) <= 320):
            continue
        current  = _ocr_scan_clean_number(cur_txt)
        previous = _ocr_scan_clean_number(prev_txt)
        if current is not None or previous is not None:
            result.aop_data[code] = AopEntry(aop=code, current_year=current, previous_year=previous)

    all_codes = [f"{i:03d}" for i in list(range(1, 97)) + [301] + list(range(110, 190))]
    missing = [code for code in all_codes if code not in result.aop_data]
    found_n = len(all_codes) - len(missing)
    result.warnings.append(
        f"⚠️ TA PDF JE SKENIRAN IN JE BIL PREBRAN Z OCR. Zanesljivo prepoznanih: {found_n}/{len(all_codes)} "
        f"postavk. PRED ODDAJO BANKI PREVERI ZNESKE PROTI IZVORNIKU. Manjkajoče postavke (preveri in "
        f"vpiši ročno, če veljajo za to bilanco): {', '.join(missing) if missing else '(ni jih)'}"
    )
    return validate(result)

def parse_ocr_fallback_format(pdf_path, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="OCR")
    if pytesseract is None:
        result.errors.append(
            "Ta PDF nima berljivega besedilnega sloja (pokvarjena pisava) in bi ga bilo treba "
            "brati z OCR, a knjižnica pytesseract ni nameščena na strežniku."
        )
        return result

    with pdfplumber.open(pdf_path) as pdf:
        full_text_pages = [pg.extract_text() or "" for pg in pdf.pages]
        # Za ime/matično/datume poskusimo OCR na prvih dveh straneh (naslovnica + prva podatkovna stran)
        header_ocr = ""
        for pg in pdf.pages[:2]:
            try:
                im = pg.to_image(resolution=300).original
                header_ocr += "\n" + pytesseract.image_to_string(im, lang='slv+eng')
            except Exception:
                pass

        c = CompanyInfo()
        m = re.search(r'\n([A-ZČŠŽ][A-Za-zČŠŽčšž\.\s]*?D\.?[OD]\.?O?\.?)\s+(\d{7,10})\s*\n', header_ocr)
        if not m:
            m = re.search(r'([A-ZČŠŽ][A-Za-zČŠŽčšž\. ]{2,60}?D\.?O\.?O\.?)\s*\n?\s*(\d{7,10})', header_ocr)
        if m:
            c.name = re.sub(r'\s+', ' ', m.group(1)).strip()
            c.registration_number = m.group(2).strip()
        m = re.search(r'na dan\s+(\d{1,2}\.\d{1,2}\.\d{4})', header_ocr)
        if m: c.period_to = m.group(1).strip()
        m = re.search(r'obdobju\s+(\d{1,2}\.\d{1,2}\.\d{4})\s*-\s*(\d{1,2}\.\d{1,2}\.\d{4})', header_ocr)
        if m:
            c.period_from, c.period_to = m.group(1).strip(), m.group(2).strip()
        if (not c.period_from or len(c.period_from) < 6) and c.period_to:
            leto = c.period_to.split('.')[-1]
            c.period_from = f"1.1.{leto}"

        c.tip_subjekta = "1"
        result.company = finalize_company(c, tip_override, tip_subjekta_override)
        result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
        if not result.company.name or not result.company.registration_number:
            result.warnings.append("Ime podjetja in/ali matična številka niso bili zanesljivo prepoznani z OCR — prosim preveri/vpiši ročno v D3/D4.")

        for page in pdf.pages:
            page_entries = _ocr_extract_page_entries(page)
            for code, (cur, prev) in page_entries.items():
                if code in result.aop_data:
                    continue
                result.aop_data[code] = AopEntry(aop=code, current_year=cur, previous_year=prev)

    all_codes = [f"{i:03d}" for i in list(range(1, 97)) + [301] + list(range(110, 190))]
    missing = [code for code in all_codes if code not in result.aop_data]
    found_n = len(all_codes) - len(missing)
    result.warnings.append(
        f"⚠️ TA PDF JE BIL PREBRAN Z OCR (pokvarjena pisava — besedilnega sloja ni bilo mogoče "
        f"prebrati z nobeno knjižnico). Zanesljivo prepoznanih: {found_n}/{len(all_codes)} postavk. "
        f"NUJNO PRED ODDAJO BANKI ROČNO PREVERI VSE ZNESKE PROTI IZVORNEMU PDF-JU — OCR lahko "
        f"občasno zamenja števko (npr. 0↔9). Postavke, ki jih OCR ni zanesljivo prepoznal in so "
        f"zato v izvozu 0, poglej v izvornem PDF-ju in vpiši ročno: {', '.join(missing) if missing else '(ni jih)'}"
    )
    return validate(result)

# ── PARSER: Vasco/Rave Reports medletni PDF (BS in IPI v ločenih datotekah) ──
_VASCO_AJPES_ROW_RE = re.compile(
    r'^\s*(\d{1,3})\s+(.*?)\s+(-?(?:\d{1,3}(?:\.\d{3})*|\d+),\d{2})\s*$'
)


def _vasco_ajpes_kind(pages_text):
    first = pages_text[0] if pages_text else ""
    if re.search(r'\bBS\s+za\s+družbe\s*-\s*Ajpes\b', first, re.I):
        return "BS"
    return "IPI"


def _vasco_ajpes_seq_to_aop(kind, seq):
    """Pretvori Vascovo zaporedno številko vrstice v uradno AOP kodo.

    IPI vrstice 1-78 so po vrsti AOP 110-187. Pri BS se zaporedje ujema z
    AOP 001-067, vrstica 68 je posebna AOP 301, vrstice 69-97 pa so AOP
    068-096. Vrstice 98-100 in 998/999 so kontrole oziroma nadaljevanja
    opisa ter niso AOP postavke za CBK.
    """
    if kind == "IPI":
        return f"{seq + 109:03d}" if 1 <= seq <= 78 else None
    if 1 <= seq <= 67:
        return f"{seq:03d}"
    if seq == 68:
        return "301"
    if 69 <= seq <= 97:
        return f"{seq - 1:03d}"
    return None


def parse_vasco_ajpes_1col_format(pages_text, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="VASCO_AJPES_1COL")
    full = "\n".join(pages_text)
    kind = _vasco_ajpes_kind(pages_text)

    c = CompanyInfo()
    if pages_text and pages_text[0]:
        for line in pages_text[0].split('\n'):
            candidate = line.strip()
            if not candidate or candidate.lower().startswith("stran"):
                continue
            if re.search(r'\bBS\s+za\s+družbe\b|\bIzkaz\s+poslovnega\s+izida\b', candidate, re.I):
                continue
            if candidate.startswith("Zap."):
                continue
            c.name = candidate
            break

    period = re.search(
        r'(?<!\d)(\d{1,2}\.\d{1,2}\.\d{4})\s*-\s*(\d{1,2}\.\d{1,2}\.\d{4})(?!\d)',
        full
    )
    if period:
        c.period_from, c.period_to = period.group(1), period.group(2)
    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    result.warnings.append(
        "Ta Vasco/AJPES medletni izvoz ne vsebuje matične številke — vpiši jo ročno v celico D4."
    )

    for page_text in pages_text:
        for line in page_text.split('\n'):
            match = _VASCO_AJPES_ROW_RE.match(line)
            if not match:
                continue
            seq = int(match.group(1))
            aop = _vasco_ajpes_seq_to_aop(kind, seq)
            if not aop or aop in result.aop_data:
                continue
            value = parse_si(match.group(3))
            if value is not None:
                result.aop_data[aop] = AopEntry(
                    aop=aop, current_year=value, previous_year=None
                )

    expected = 97 if kind == "BS" else 78
    if len(result.aop_data) < expected:
        result.warnings.append(
            f"Vasco/AJPES format je bil prepoznan le delno: prebranih je "
            f"{len(result.aop_data)}/{expected} pričakovanih postavk."
        )
    if kind == "IPI" and "188" not in result.aop_data:
        result.warnings.append(
            "IPI v tem izvozu ne vsebuje AOP 188 (povprečno število zaposlenih); "
            "po potrebi ga vpiši ročno."
        )
    return validate(result)


# ── PARSER: AJPES_ZAP format (varianta zgornjega, brez "tekoče leto" v glavi) ────
# BS bere postavke po BESEDILU (kot UPR/PLAIN_MST) namesto po zaporedni številki,
# ker nekateri izvozi te variante izpustijo vrstico AOP 301 kadar je 0, kar bi
# pri pozicijskem branju zamaknilo vse poznejše postavke. IPI te težave nima
# (uradna struktura AOP 110-187 nima podobnih izpustov), zato se tam obdrži
# zanesljivo, enostavnejše pozicijsko branje (_vasco_ajpes_seq_to_aop).
_AJPES_ZAP_DIRECT_BS = {k.lower(): v for k, v in {
    "I. Neopredmetena sredstva in dolg.aktivne čas.razmejitve": "003",
    "a) Dolg.premoženjske pravice": "005",
    "c) Dolg.odloženi stroški razvijanja": "007",
    "2. Dolg. aktivne časovne razmejitve": "009",
    "4. Druge naprave in oprema, DI in druga OOS": "014",
    "6. Opred.osnovna sredstva v gradnji in izdelavi": "016",
    "7. Predujmi za pridobitev opred.osnovnih sredstev": "017",
    "1. Krat.posl.terj.do družb v skupini": "049",
    "2. Krat.posl.terj. do kupcev": "050",
    "3. Krat.posl.terj.do drugih": "051",
    "C. Kratk.aktivne časovne razmejitve": "053",
    "Zabilančna sredstva": "054",
    "2. Rezerve za lastne delnice in lastne posl. deleže": "063",
    "3. Lastne delnice in lastni posl.deleži(kot odbitna p": "064",
    "IV. Presežek iz prevrednotenja": "067",   # sinonim za "IV. Revalorizacijske rezerve"
    "B. Rezervacije in dolg.pasivne čas.razmejitve": "072",
    "I. Dolg.finančne obveznosti": "076",
    "1. Dolg.fin.obveznosti do družb v skupini": "077",
    "2. Dolg.fin.obveznosti do bank": "078",
    "3. Druge dolg.fin.obveznosti": "079",
    "II. Dolg.poslovne obveznosti": "080",
    "1. Dolg.posl.obveznosti do družb v skupini": "081",
    "2. Dolg.posl.obveznosti do dobaviteljev": "082",
    "3. Druge dolg.poslovne obveznosti": "083",
    "II. Kratk.finančne obveznosti": "087",
    "1. Kratk.fin.obveznosti do družb v skupini": "088",
    "2. Kratk.fin.obveznosti do bank": "089",
    "3. Druge kratk.fin.obveznosti": "090",
    "III. Kratk.poslovne obveznosti": "091",
    "1. Kratk.posl.obveznosti do družb v skupini": "092",
    "2. Kratk.posl.obveznosti do dobaviteljev": "093",
    "3. Druge kratk.posl.obveznosti": "094",
    "D. Kratk.pasivne časovne razmejitve": "095",
    "Zabilančne obveznosti": "096",
}.items()}

def _ajpes_zap_kind(pages_text):
    first = pages_text[0] if pages_text else ""
    if re.search(r'\bIzkaz\s+poslovnega\s+izida\b', first, re.I):
        return "IPI"
    return "BS"

def parse_ajpes_zap_format(pages_text, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="AJPES_ZAP")
    full = "\n".join(pages_text)
    kind = _ajpes_zap_kind(pages_text)

    c = CompanyInfo()
    if pages_text and pages_text[0]:
        for line in pages_text[0].split('\n'):
            candidate = line.strip()
            if not candidate or candidate.lower().startswith("stran"):
                continue
            if re.search(r'\bBS\s+za\s+družbe\b|\bIzkaz\s+poslovnega\s+izida\b', candidate, re.I):
                continue
            if candidate.startswith("Zap."):
                continue
            c.name = candidate
            break

    period = re.search(
        r'(?<!\d)(\d{1,2}\.\d{1,2}\.\d{4})\s*-\s*(\d{1,2}\.\d{1,2}\.\d{4})(?!\d)',
        full
    )
    if period:
        c.period_from, c.period_to = period.group(1), period.group(2)
    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    result.warnings.append(
        "Ta izvoz ne vsebuje matične številke — vpiši jo ročno v celico D4."
    )

    context = None
    for page_text in pages_text:
        for line in page_text.split('\n'):
            match = _VASCO_AJPES_ROW_RE.match(line)
            if not match:
                continue
            seq = int(match.group(1))
            label = match.group(2)
            value = parse_si(match.group(3))
            if kind == "IPI":
                aop = _vasco_ajpes_seq_to_aop("IPI", seq)
            else:
                aop = _match_ajpes_1col_label(label, _AJPES_ZAP_DIRECT_BS, _AMBIGUOUS_ITEMS, context)
                if aop and aop in _CONTEXT_TRIGGERS:
                    context = _CONTEXT_TRIGGERS[aop]
            if not aop or aop in result.aop_data:
                continue
            if value is not None:
                result.aop_data[aop] = AopEntry(aop=aop, current_year=value, previous_year=None)

    expected = 96 if kind == "BS" else 78
    if len(result.aop_data) < expected:
        result.warnings.append(
            f"AJPES_ZAP format je bil prepoznan le delno: prebranih je "
            f"{len(result.aop_data)}/{expected} pričakovanih postavk."
        )
    return validate(result)


# ── PARSER: AJPES_1COL format ("TIGI TRADE" ipd. izvoz - polno besedilo, EN znesek/vrstico) ──
# PDF nima tabele/stolpcev - samo opisno besedilo postavke + en sam znesek na koncu
# vrstice (brez "Tekočega"/"Prejšnjega leta" primerjave - previous_year je zato vedno
# prazen, glej opombo v _is_ajpes_1col_format). BS in IPI sta v LOČENIH datotekah
# (enako kot pri BS_IPI_AJPES) - obstoječi mehanizem za več datotek ju združi.
# Matične številke format NIMA. Večina postavk (na testni BS datoteki 96/96, torej
# 100 %) se ujame kar prek splošnega _match_aop_label_exact mehanizma, ker je
# besedilo zelo blizu uradnemu AJPES obrazcu - za preostale (okrajšave npr. "AČR"
# namesto "aktivne časovne razmejitve", "DI" namesto "drobni inventar", "Zabilančna"
# namesto uradnega "Zunajbilančna", ter dvoumne/kontekstno odvisne postavke npr.
# "a) Delnice in deleži v družbah v skupini" ki se pojavi tako pri dolgoročnih kot
# kratkoročnih finančnih naložbah) imamo spodaj direktne alias tabele in ponovno
# uporabimo obstoječa kontekstna pravila _ZAVOD_AMBIGUOUS_BS/_ZAVOD_CONTEXT_TRIGGERS_BS
# (BS) oz. _OBRACUNI_AMBIGUOUS_IU/_OBRACUNI_TRIGGERS_IU (IPI) - isti vzorec kot pri
# OBRACUNI_MEDLETNI/BS_IPI_AJPES.
#
# POMEMBNA NAJDBA (ni specifična za ta format, gre za obstoječ JOLP_TO_AOP slovar):
# "1. Rezervacije" se v JOLP_TO_AOP pojavi DVAKRAT - enkrat za BS (AOP 073, znotraj
# "B. Rezervacije in dolgoročne PČR") in enkrat za IPI (AOP 149, znotraj poslovnih
# odhodkov). Ker je JOLP_TO_AOP navaden Python slovarski literal, DRUGI (149) povozi
# PRVEGA (073) že ob uvozu modula - "073" v surovem slovarju sploh ne obstaja več,
# zato ga _match_aop_label_exact za BS vrstico "1. Rezervacije" nikoli ne najde
# pravilno (vrne "149" namesto "073"). To smo tu popravili z direktnim aliasom v
# _AJPES_1COL_DIRECT_BS, a POZOR: ista kolizija verjetno prizadene tudi obstoječe
# ZAVOD_MEDLETNI/OBRACUNI_MEDLETNI/BS_IPI_AJPES formate, če ima njihova BS datoteka
# neničelno vrednost pri "1. Rezervacije" - vredno preveriti/popraviti v ločeni seji.
_AJPES_1COL_DIRECT_BS = {k.lower(): v for k, v in {
    "I. Neopredmetena sredstva in dolgoročne AČR": "003",
    "4. Druge naprave in oprema, DI in druga opred. OS": "014",
    "7. Predujmi za pridobitev opredmetenih osn. sredstev": "017",
    "Zabilančna sredstva": "054",
    "3. Lastne delnice in lastni posl. deleži (kot odbitna p.)": "064",
    "IV. Presežek iz prevrednotenja": "067",
    "B. REZERVACIJE IN DOLGOROČNE PČR": "072",
    "1. Rezervacije": "073",   # glej opombo o JOLP_TO_AOP koliziji zgoraj
    "1. Dolgoročne finančne obveznosti do družbe v skupini": "077",  # izvor ima ednino "družbe" namesto množine "družb"
    "Zabilančne obveznosti": "096",
    "POSLOVNI IZID OBDOBJA": None,  # informativna vrstica na koncu BS, ni uradna AOP postavka - namerno preskočimo
}.items()}

# Testirano na resnični IPI datoteki tega formata ("TIGI TRADE") - splošni mehanizem
# pokrije 58/80 vrstic sam, preostalih 21 (okrajšave/prirezano besedilo izvoznega
# programa, npr. "posl." namesto "poslovnih", vrstice porezane na fiksno širino kot
# "...v skupi" namesto "...v skupini") je tu, + 1 namerno preskočena vrstica
# ("Število mesecev poslovanja" - AOP 189 se itak vedno prepiše iz _months_in_period,
# glej export_excel). Skupaj 79/79 zapolnjenih AOP postavk (110-188 brez lukenj).
_AJPES_1COL_DIRECT_IPI = {k.lower(): v for k, v in {
    "1. Čisti prihodki od prodaje proiz.in storit.razen najemn": "112",
    "B. POVEČANJE VREDNOSTI ZALOG PROIZVODOV IN NEDOK.PROIZV.": "121",
    "C. ZMANJŠANJE VREDNOSTI ZALOG PROIZVOD. IN NEDOK.PROIZV.": "122",
    "Č. USREDSTVENI LASTNI PROIZVODI IN STORITVE": "123",
    "D. SUBVENCIJE, DOTACIJE, REGRESI IN DRUGI PRIHODKI": "124",  # ta je manjkala pri prejšnji pretvorbi
    "1. Nabavna vrednost prodanega blaga in materiala": "129",
    "c) povračila stroškov zaposlencem v zvezi z delom": "137",   # "zaposlencem" (dajalnik) namesto uradnega "zaposlenim"
    "2. Prevrednotovalni posl.odhodki pri neopred. in opred. O": "146",
    "3. Prevrednotovalni posl.odhodki pri obratnih sredstvih": "147",
    "Finančni prih. od obresti (upoštevano že v II. in III.": "154",   # izvor obreže zaklepaj ")" na koncu
    "1. Finančni prihodki iz posl. terjatev do družb v skupini": "164",
    "Finančni odh. za obresti (upoštevano že v II. in III.)": "167",
    "I. Finančni odhodki iz oslabitve in odpisov fin.naložb": "168",
    "1. Finančni odhodki iz posojil, prejetih od družb v skupi": "170",  # izvor obreže "-ni" na koncu (fiksna širina vrstice)
    "1. Finančni odhodki iz poslovnih obv. do družb v skupini": "175",
    "2. Finančni odhodki iz obv. do dobaviteljev in menične ob": "176",  # obrezano "obveznosti" -> "ob"
    "I. Subvencije, dotacije in podobni prih.,ki niso povezani": "179",
    "II. Drugi finančni prihodki in ostali prihodki": "180",
    "P. DAVEK IZ DOBIČKA": "184",
    "R. ODLOŽENI DAVKI": "185",
    "Povprečno število zaposlencev na podlagi delovnih ur": "188",
    "Število mesecev poslovanja": None,  # AOP 189 se vedno prepiše avtomatsko iz obdobja (_months_in_period) - namerno preskočimo
}.items()}



def _match_ajpes_1col_label(label, direct_map, ambiguous_map, context):
    """Isti troplastni vzorec kot pri _match_obracuni_label (direkten alias →
    kontekstno razrešena dvoumnost → splošni mehanizem), le da je zadnja stopnja
    _match_aop_label_exact namesto neposredno _JOLP_TO_AOP_NORM. To je pri tem
    formatu potrebno, ker je marsikatera postavka ('SREDSTVA' nasproti 'I. Sredstva
    (skupine za odtujitev) za prodajo' ...) v NORMALIZIRANI obliki dvoumna in jo
    _build_normalized_aop_map() zato namenoma izloči iz _JOLP_TO_AOP_NORM - ujameta
    pa jo prva dva (surova/skoraj-surova) koraka znotraj _match_aop_label_exact."""
    raw_key = re.sub(r'\s+', ' ', str(label).strip()).lower()
    if raw_key in direct_map:
        return direct_map[raw_key]
    kind, text = _classify_and_strip_item(label)
    opts = ambiguous_map.get(text)
    if opts:
        return opts.get(context)
    return _match_aop_label_exact(label)


def _ajpes_1col_kind(pages_text):
    """Vrne 'BS' ali 'IPI' glede na naslov v prvih vrsticah prve strani."""
    first = pages_text[0] if pages_text else ""
    if "BILANCA STANJA" in first:
        return "BS"
    if "IZKAZ POSLOVNEGA IZIDA" in first:
        return "IPI"
    return "BS"


_AJPES_1COL_TITLE_RE = re.compile(r'^(?:BILANCA STANJA|IZKAZ POSLOVNEGA IZIDA)\s*-\s*AJPES/', re.I)
_AJPES_1COL_NUM_RE = re.compile(r'(-?[\d.]+,\d{2})\s*$')


def parse_ajpes_1col_format(pages_text, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="AJPES_1COL")
    full = "\n".join(pages_text)
    kind = _ajpes_1col_kind(pages_text)
    is_bs = (kind == "BS")

    c = CompanyInfo()
    if pages_text and pages_text[0]:
        first_line = pages_text[0].split('\n')[0].strip()
        if first_line and not first_line.lower().startswith("stran"):
            c.name = first_line
    m = re.search(r'AJPES/(\d{1,2}\.\d{1,2}\.\d{4})\s*-\s*(\d{1,2}\.\d{1,2}\.\d{4})', full)
    if m:
        c.period_from, c.period_to = m.group(1), m.group(2)
    else:
        m2 = re.search(r'AJPES/(\d{1,2}\.\d{1,2}\.\d{4})', full)
        if m2:
            c.period_to = m2.group(1)
    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    result.warnings.append(
        "Ta format (AJPES_1COL - izvoz tipa \"TIGI TRADE\") ne vsebuje matične "
        "številke — vpiši jo ročno v celico D4."
    )

    direct_map = _AJPES_1COL_DIRECT_BS if is_bs else _AJPES_1COL_DIRECT_IPI
    ambiguous_map = _ZAVOD_AMBIGUOUS_BS if is_bs else _OBRACUNI_AMBIGUOUS_IU
    triggers = _ZAVOD_CONTEXT_TRIGGERS_BS if is_bs else _OBRACUNI_TRIGGERS_IU

    context = None
    for page_text in pages_text:
        for line in page_text.split('\n'):
            low = line.strip()
            if not low or low.lower().startswith("stran"):
                continue
            if _AJPES_1COL_TITLE_RE.match(low):
                continue
            mnum = _AJPES_1COL_NUM_RE.search(low)
            if not mnum:
                continue
            label = low[:mnum.start()].strip()
            if not label:
                continue
            aop = _match_ajpes_1col_label(label, direct_map, ambiguous_map, context)
            if aop and aop in triggers:
                context = triggers[aop]
            if not aop or aop in result.aop_data:
                continue
            num = parse_si(mnum.group(1))
            if num is not None:
                result.aop_data[aop] = AopEntry(aop=aop, current_year=num, previous_year=None)

    if not kind:
        result.warnings.append("Format AJPES_1COL ni bil v celoti prepoznan v tej datoteki.")
    return validate(result)


# ── PARSER: VIZIJA format ("DENAR | OBRAČUNI" računovodski program, čisto besedilo brez AOP kode) ──
_VIZIJA_MARKER_RE = re.compile(r'^[IVXLCDM]{1,6}\.\s|^[A-ZČŽŠ]\.\s|^\d{1,2}[a-zčžš]?\.\s|^[a-zčžš]\.\s', re.I)
_VIZIJA_NUM_RE = re.compile(r'(-?[\d.]+,\d{2})\s*$')
_VIZIJA_NOISE_RE = re.compile(r'\s*\d*\s*:?cezarbO.*$|Registriran uporabnik.*$|narirtsigeR.*$', re.I)

def _vizija_parse_page_lines(page_text, entries):
    lines = page_text.split('\n')
    start = 0
    for i, l in enumerate(lines):
        if re.match(r'^Naziv\s+Tekoč', l):
            start = i + 1
            break
    for line in lines[start:]:
        line = _VIZIJA_NOISE_RE.sub('', line).rstrip()
        low = line.strip()
        if not low:
            continue
        if low.startswith('Stran') or 'Registriran uporabnik' in low or 'Oseba, odgovorna' in low \
           or low.startswith('Vodja ') or low.startswith('dne:') or re.match(r'^V\s*\.{3,}', low):
            continue
        m = _VIZIJA_NUM_RE.search(line)
        if m:
            entries.append((line[:m.start()].strip(), m.group(1)))
        elif _VIZIJA_MARKER_RE.match(low):
            entries.append((low, None))
        elif entries:
            lbl, val = entries[-1]
            entries[-1] = (lbl + ' ' + low, val)

def parse_vizija_format(pages_text, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="VIZIJA")
    full = "\n".join(pages_text)
    first = pages_text[0] if pages_text else ""
    c = CompanyInfo()

    m = re.search(r'Podjetje\s+(.+?)\n', first)
    if m: c.name = m.group(1).strip()

    is_ipi = "poslovnega izida" in first.lower()
    m = re.search(r'od\s+(\d{1,2}\.\d{1,2}\.\d{4})\s+do\s+(\d{1,2}\.\d{1,2}\.\d{4})', full)
    if m:
        c.period_from, c.period_to = m.group(1).strip(), m.group(2).strip()
    if (not c.period_from or len(c.period_from) < 6) and c.period_to:
        leto = c.period_to.split('.')[-1]
        c.period_from = f"1.1.{leto}"

    c.tip_subjekta = "1"
    result.company = finalize_company(c, tip_override, tip_subjekta_override)
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)
    if not result.company.registration_number:
        result.warnings.append("Ta format ne vsebuje matične številke — vpiši jo ročno v celico D4.")

    entries = []
    for page_text in pages_text:
        _vizija_parse_page_lines(page_text, entries)

    for label, val in entries:
        if val is None:
            continue
        aop = _match_aop_label_exact(label)
        if not aop or aop in result.aop_data:
            continue
        aop_num = int(re.sub(r'\D', '', aop) or 0)
        if is_ipi and aop_num < 110:
            continue
        if not is_ipi and aop_num >= 110:
            continue
        num = parse_si(val)
        if num is not None:
            result.aop_data[aop] = AopEntry(aop=aop, current_year=num, previous_year=None)
    return validate(result)


# ── PARSER: eDavki DDD (obračun dohodka iz dejavnosti) — skeniran PDF ─────────
# Ta obrazec vsebuje Prilogo 3 (BS) in Prilogo 4 (IPI) za s.p. z ENIM stolpcem
# tekočega leta. Generični OCR parser pričakuje klasičen AJPES obrazec z dvema
# stolpcema (tekoče/preteklo leto), zato je za eDavki DDD potreben ločen parser.
_EDAVKI_SP_SOURCE_CODES = {
    # BS
    "001","002","003","004","009","010","018","019","020","024","027","032","033","034","035","036","037","038","039","040","041","045","048","052","053","054",
    "055","056","058","060a","060b","067","301","070","071","072","073","074","075","076","080","085","086","087","091","095","096",
    # IPI
    "110","111","115","118","121","122","123","124","125","126","127","128","129","130","134","139","140","141","142","143",
    "144","145","146","147","148","148a","148b","151","152","153","155","160","163","166","167","168","169","174",
    "178","179","180","181","182","183","188","189",
}

def _edavki_render_page(pdf_path, page_number, resolution=180):
    """Render ene strani s Popplerjem (pdftoppm). Pri drobnem tisku eDavki DDD je
    ta render bistveno stabilnejši za OCR kot PDF->PIL pot prek pdfplumber."""
    if Image is None:
        return None
    tmpdir = Path(tempfile.mkdtemp(prefix="edavki_ocr_"))
    prefix = tmpdir / "page"
    try:
        subprocess.run(
            ["pdftoppm", "-f", str(page_number), "-l", str(page_number),
             "-singlefile", "-png", "-r", str(resolution), str(pdf_path), str(prefix)],
            check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=45
        )
        png = prefix.with_suffix(".png")
        if not png.exists():
            return None
        im = Image.open(png).copy()
        return im
    except Exception:
        return None
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def _edavki_scan_first_page_text(pdf_path, resolution=140):
    if pytesseract is None:
        return ""
    try:
        im = _edavki_render_page(pdf_path, 1, resolution)
        if im is None:
            return ""
        return pytesseract.image_to_string(im, lang='slv+eng', config='--psm 6')
    except Exception:
        return ""

def _is_edavki_ddd_scan(pdf_path):
    txt = _edavki_scan_first_page_text(pdf_path)
    low = txt.lower()
    return ("edavki" in low and "dohodka iz dejavnosti" in low and
            ("obračun akontacije dohodnine" in low or "obracun akontacije dohodnine" in low))

def _edavki_normalize_aop_token(token):
    t = (token or "").strip().lower().strip('.,:;()[]{}')
    # Najpogostejši OCR zamenjavi na začetku AOP kode: o/O → 0, l/I → 1.
    t = t.replace('o', '0')
    if re.fullmatch(r'[0-9]{2,3}[ab]?', t):
        digits = re.match(r'\d+', t).group(0)
        suffix = t[len(digits):]
        return digits.zfill(3) + suffix
    return None

def _edavki_extract_raw_codes_from_text(text):
    """Iz OCR besedila Prilog 3/4 vrne AOP -> znesek tekočega leta.
    Vrstica brez zneska je na uradnem DDD obrazcu prazna in se obravnava kot 0."""
    raw = {}
    for line in (text or "").splitlines():
        line = re.sub(r'\s+', ' ', line).strip()
        if not line:
            continue
        first = line.split(' ', 1)[0]
        code = _edavki_normalize_aop_token(first)
        if not code or code not in _EDAVKI_SP_SOURCE_CODES:
            continue
        m = re.search(r'(-?[\d.]+,\d{2})\s*$', line)
        raw[code] = parse_si(m.group(1)) if m else 0.0
    return raw

def _edavki_sp_fill_derived(raw, seen=None):
    """Preračuna uradne kontrolne seštevke iz podrobnejših AOP vrstic.
    Tako OCR napaka v natisnjenem seštevku ne more pokvariti CBK rezultata."""
    seen = seen or set()
    def v(code):
        x = raw.get(code)
        return 0.0 if x is None else float(x)
    def missing(code):
        return code not in seen

    # BS — uradne formule Priloge 3.
    if missing("003"): raw["003"] = v("004") + v("009")
    if missing("019"): raw["019"] = v("020") + v("024")
    if missing("034"): raw["034"] = sum(v(c) for c in ("035","036","037","038","039"))
    if missing("040"): raw["040"] = v("041") + v("045")
    # AOP 048 je samostojna vrstica; OCR ga pogosto prebere kot 04g, zato ga rešimo iz AOP 032.
    if missing("048") and raw.get("032") is not None and abs(v("032")) > 0:
        raw["048"] = v("032") - v("033") - v("034") - v("040") - v("052")
    raw["002"] = v("003") + v("010") + v("018") + v("019") + v("027")
    raw["001"] = v("002") + v("032") + v("053")
    raw["056"] = v("058") + v("060a") + v("060b") + v("067") + v("301") + v("070") - v("071")
    if missing("072"): raw["072"] = v("073") + v("074")
    if missing("075"): raw["075"] = v("076") + v("080")
    if missing("085"): raw["085"] = v("086") + v("087") + v("091")
    raw["055"] = v("056") + v("072") + v("075") + v("085") + v("095")

    # IPI — uradne formule Priloge 4.
    raw["126"] = v("110") + v("121") - v("122") + v("123") + v("124") + v("125")
    # AOP 128 izračunamo iz 129+130+134. Pri DDD drobni tisk v seštevku 128 pogosto
    # zamenja 5↔6, podvrstici 130 in 134 pa sta bistveno bolj zanesljivi.
    raw["128"] = v("129") + v("130") + v("134")
    if missing("139"): raw["139"] = v("140") + v("141") + v("142") + v("143")
    if missing("144"): raw["144"] = v("145") + v("146") + v("147")

    # AOP 127 najprej kontrolno izpeljemo iz kosmatega donosa in izida poslovanja,
    # če sta 151/152 na obrazcu prepoznana. Tako ne zaupamo potencialno napačnemu OCR
    # seštevku 148 oziroma podvrstici 148a.
    if "151" in seen or "152" in seen:
        raw["127"] = v("126") - v("151") + v("152")
    # Drugi poslovni odhodki so nato rezidual do celotnih poslovnih odhodkov.
    raw["148"] = v("127") - v("128") - v("139") - v("144")
    diff = v("126") - v("127")
    raw["151"] = max(diff, 0.0)
    raw["152"] = max(-diff, 0.0)
    if missing("153"): raw["153"] = v("155") + v("160") + v("163")
    if missing("166"): raw["166"] = v("168") + v("169") + v("174")
    if missing("178"): raw["178"] = v("179") + v("180")
    total = v("151") - v("152") + v("153") - v("166") + v("178") - v("181")
    raw["182"] = max(total, 0.0)
    raw["183"] = max(-total, 0.0)
    return raw

def parse_edavki_ddd_sp_scan(pdf_path, tip_override=None, tip_subjekta_override=None):
    result = ParseResult(pdf_format="EDAVKI_DDD_SP_OCR")
    if pytesseract is None:
        result.errors.append("eDavki DDD PDF je skeniran, vendar pytesseract ni nameščen.")
        return result

    with pdfplumber.open(pdf_path) as pdf:
        # Glava: prva stran zadostuje za naziv, matično, davčno in obdobje.
        im0 = _edavki_render_page(pdf_path, 1, 180)
        header = pytesseract.image_to_string(im0, lang='slv+eng', config='--psm 6') if im0 is not None else ""

        # Priloga 3/4 je pri eDavki DDD za glavnim obračunom, za njo pa so lahko še druge
        # priloge. Zato gremo od konca nazaj in se ustavimo takoj, ko najdemo začetek BS.
        # Tako ne OCR-amo po nepotrebnem vseh strani dolgega davčnega obračuna.
        table_texts_rev = []
        for page_no in range(len(pdf.pages), 0, -1):
            im = _edavki_render_page(pdf_path, page_no, 180)
            txt = pytesseract.image_to_string(im, lang='slv+eng', config='--psm 6') if im is not None else ""
            table_texts_rev.append(txt)
            low = txt.lower()
            if "podatki iz bilance stanja" in low:
                break
            if len(table_texts_rev) >= 8:
                break

        raw = {code: 0.0 for code in _EDAVKI_SP_SOURCE_CODES}
        seen = set()
        for txt in reversed(table_texts_rev):
            found = _edavki_extract_raw_codes_from_text(txt)
            for code, val in found.items():
                if code not in seen:
                    raw[code] = val
                    seen.add(code)

    raw = _edavki_sp_fill_derived(raw, seen)

    c = CompanyInfo()
    m = re.search(r'za obdobje od\s*(\d{1,2}\.\d{1,2}\.\d{4})\s*do\s*(\d{1,2}\.\d{1,2}\.\d{4})', header, re.I)
    if m:
        c.period_from, c.period_to = m.group(1), m.group(2)
    m = re.search(r'Ime in priimek zavezanca\s+([^\n]+)', header, re.I)
    if m:
        c.name = re.sub(r'\s+', ' ', m.group(1)).strip()
    m = re.search(r'Matična številka obrata\s+(\d{7,10})', header, re.I)
    if m:
        c.registration_number = m.group(1)
    # Davčno številko raje vzamemo iz EDP številke dokumenta; OCR labela lahko zamenja 6↔8.
    m = re.search(r'EDP-(\d{8})-', header, re.I)
    if m:
        c.tax_number = m.group(1)
    else:
        m = re.search(r'Davčna številka\s+(\d{8})', header, re.I)
        if m: c.tax_number = m.group(1)

    c.tip_subjekta = "2"
    result.company = finalize_company(c, tip_override, tip_subjekta_override or "2")
    result.subject_type = _subject_type_from_tip_subjekta(result.company.tip_subjekta)

    # Ohranimo varne numerične AOP kode tudi v aop_data, da exporter lahko zapolni
    # postavke, ki imajo pri s.p. enak pomen kot pri standardnem obrazcu.
    for code, val in raw.items():
        if code.isdigit():
            z = code.zfill(3)
            result.aop_data[z] = AopEntry(aop=z, current_year=val, previous_year=None)
    for code in _SP_IPI_COLLISION_CODES:
        result.aop_data.pop(code, None)

    def sp_formula(formula):
        if isinstance(formula, tuple):
            ca, cb, op = formula
            va, vb = raw.get(ca, 0.0) or 0.0, raw.get(cb, 0.0) or 0.0
            return va + vb if op == "+" else va - vb
        return raw.get(formula, 0.0)

    for gvaop, formula in _SP_GVAOP_MAP_BS.items():
        result.gvaop_data[gvaop] = sp_formula(formula)
    for gvaop, formula in _SP_GVAOP_MAP_IPI.items():
        result.gvaop_data[gvaop] = sp_formula(formula)
    result.gvaop_data["067"] = ((raw.get("151") or 0) - (raw.get("152") or 0)
                                 + (raw.get("153") or 0) - (raw.get("166") or 0))

    result.warnings.append(
        "eDavki DDD s.p. PDF je skeniran in prebran z namenskim OCR parserjem za Prilogo 3/4. "
        "Vrstice brez izkazanega zneska so v CBK zapisane kot 0; ključni seštevki so dodatno kontrolno preračunani."
    )
    return validate(result)

def parse_pdf_file(pdf_path, tip_override=None, tip_subjekta_override=None):
    pages_text = []
    pages_words = []
    pages_chars = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            pages_text.append(page.extract_text() or "")
            pages_words.append(page.extract_words(use_text_flow=False, keep_blank_chars=False))
            pages_chars.append(page.chars)
    # Preveri ali je PDF skeniran (brez besedilnega sloja) ali CID-encoded (pokvarjen font) -
    # v obeh primerih besedila ni mogoče prebrati normalno, zato poskusimo z OCR.
    if is_scanned_no_text(pages_text):
        if str(tip_subjekta_override or "").strip() == "2":
            return parse_edavki_ddd_sp_scan(pdf_path, tip_override, tip_subjekta_override)
        if _is_edavki_ddd_scan(pdf_path):
            return parse_edavki_ddd_sp_scan(pdf_path, tip_override, tip_subjekta_override)
        return parse_ocr_scanned_aopcol_format(pdf_path, tip_override, tip_subjekta_override)
    if is_cid_encoded(pages_text):
        return parse_ocr_fallback_format(pdf_path, tip_override, tip_subjekta_override)
    fmt = detect_pdf_format(pages_text)
    if fmt == "JOLP":
        result = parse_jolp_format(pages_text, tip_override, tip_subjekta_override)
    elif fmt == "NAPOVED":
        result = parse_napoved_format(pages_text, tip_override, tip_subjekta_override)
    elif fmt == "UPR":
        result = parse_upr_format(pages_text, pages_words, tip_override, tip_subjekta_override)
    elif fmt == "BESEDILO_2COL":
        result = parse_besedilo_2col_format(pages_text, pages_words, tip_override, tip_subjekta_override)
    elif fmt == "PLAIN_MST":
        result = parse_plain_mst_format(pages_text, pages_words, tip_override, tip_subjekta_override)
    elif fmt == "AOPCOL":
        result = parse_aopcol_format(pages_text, pages_chars, tip_override, tip_subjekta_override)
    elif fmt == "URADNI":
        result = parse_uradni_format(pages_text, pages_words, tip_override, tip_subjekta_override)
    elif fmt == "VASCO_AJPES_1COL":
        result = parse_vasco_ajpes_1col_format(pages_text, tip_override, tip_subjekta_override)
    elif fmt == "AJPES_ZAP":
        result = parse_ajpes_zap_format(pages_text, tip_override, tip_subjekta_override)
    elif fmt == "AJPES_1COL":
        result = parse_ajpes_1col_format(pages_text, tip_override, tip_subjekta_override)
    elif fmt == "VIZIJA":
        result = parse_vizija_format(pages_text, tip_override, tip_subjekta_override)
    else:
        result = parse_aop_format(pages_text, tip_override, tip_subjekta_override)
    return validate(result)

# ── Exporter ──────────────────────────────────────────────────────────────────
BS_AOP  = set([str(i).zfill(3) for i in range(1,97)] + ["301"])
IPI_AOP = set([str(i).zfill(3) for i in range(110,190)])
FMT     = '#,##0.00'

def norm_aop(raw):
    if raw is None: return None
    s = str(raw).strip()
    if not s or s in ("/","—"): return None
    try: return str(int(s)).zfill(3)
    except: return None

def build_idx(ws):
    idx = {}
    for row in ws.iter_rows(min_row=12):
        k = norm_aop(row[0].value)
        if k and k not in idx: idx[k] = row[0].row
    return idx

def _to_excel_date(s):
    """Pretvori '31.12.2025' v Python date objekt za Excel."""
    from datetime import date
    if not s or not isinstance(s, str): return None
    try:
        parts = s.strip().split('.')
        if len(parts) == 3:
            return date(int(parts[2]), int(parts[1]), int(parts[0]))
    except: pass
    return s  # fallback na string če pretvorba ne uspe

def build_idx_by_gvaop(ws, max_row=200):
    """Indeksira vrstice po stolpcu B (Oznaka GVAOP, npr. '[071]' -> '071'). Uporabno za
    vrstice, ki nimajo svoje AOP kode v stolpcu A (prikazano kot '/'), a imajo unikatno
    GVAOP oznako v stolpcu B in predstavljajo seštevek/razliko dveh drugih AOP postavk."""
    idx = {}
    for row in ws.iter_rows(min_row=12, max_row=max_row):
        b = row[1].value
        if b is None:
            continue
        s = str(b).strip()
        m = re.match(r'^\[(.+)\]$', s)
        if m: s = m.group(1)
        if s and s not in idx:
            idx[s] = row[0].row
    return idx

# Vrstice brez lastne AOP kode (v templatu '/'), ki predstavljajo "dobiček/(izguba)" seštevek
# dveh ločenih AOP postavk (dobiček in izguba se v uradnem obrazcu vedno vodita ločeno, tu pa
# ju za pregled seštejemo v en skupni znesek: dobiček - izguba).
# Format: (GVAOP oznaka kombinirane vrstice, AOP koda "dobiček", AOP koda "izguba")
# Vrstice brez lastne AOP kode (v templatu '/'), ki predstavljajo izračunan seštevek/razliko
# dveh ločenih AOP postavk. Format: (GVAOP oznaka kombinirane vrstice, AOP koda A, AOP koda B, operacija)
# operacija '-' => A - B  (npr. dobiček - izguba, ki se v uradnem obrazcu vedno vodita ločeno)
# operacija '+' => A + B  (npr. dve postavki prihodkov, ki se preprosto seštejeta)
_COMBINED_ROWS_BS = [
    ("0030104", "068", "069", "-"),  # Vi. Preneseni čisti poslovni izid = VI. Preneseni čisti dobiček - VII. Prenesena čista izguba
    ("0030105", "070", "071", "-"),  # VII. Čisti poslovni izid poslovnega leta = VIII. Čisti dobiček - IX. Čista izguba poslovnega leta
]
_COMBINED_ROWS_IPI = [
    ("061", "151", "152", "-"),  # Poslovni izid iz poslovanja = H. DOBIČEK IZ POSLOVANJA - I. IZGUBA IZ POSLOVANJA
    ("071", "182", "183", "-"),  # Celotni poslovni izid = N. CELOTNI DOBIČEK - O. CELOTNA IZGUBA
    ("075", "186", "187", "-"),  # Čisti dobiček/(izguba) obračunskega obdobja = S. ČISTI DOBIČEK - Š. ČISTA IZGUBA obr. obdobja
    ("051", "121", "122", "-"),  # Sprememba vrednosti zalog = B. Povečanje - C. Zmanjšanje (enako kot v uradni formuli za AOP 126)
    ("053", "124", "125", "+"),  # Drugi poslovni prihodki (skupaj s subvencijami) = D. Subvencije... + E. Drugi poslovni prihodki
]

def _fill_combined_rows(ws, aop_data, combined_defs):
    gvaop_idx = build_idx_by_gvaop(ws)
    filled = 0
    for gvaop, code_a, code_b, op in combined_defs:
        row = gvaop_idx.get(gvaop)
        if not row:
            continue
        e1 = aop_data.get(code_a)
        e2 = aop_data.get(code_b)
        if e1 is None and e2 is None:
            continue
        v1 = e1.current_year if (e1 and e1.current_year is not None) else 0
        v2 = e2.current_year if (e2 and e2.current_year is not None) else 0
        ws.cell(row, 4).value = (v1 + v2) if op == "+" else (v1 - v2)
        filled += 1
    return filled


def _normalize_ipi_operating_loss_for_cbk(aop_data):
    """Izgubo iz poslovanja vpiše kot negativen dobiček iz poslovanja.

    Bonitetni sistem za poslovni izid uporablja AOP 151 tudi pri izgubi: vrednost
    AOP 152 se zato izvozi kot negativen AOP 151, AOP 152 pa kot 0. Drugih parov
    dobiček/izguba (182/183 in 186/187) ne spreminjamo.
    """
    normalized = dict(aop_data)
    profit_entry = aop_data.get("151")
    loss_entry = aop_data.get("152")
    profit = profit_entry.current_year if profit_entry else None
    loss = loss_entry.current_year if loss_entry else None
    if loss is not None and abs(loss) >= 0.000001 \
            and (profit is None or abs(profit) < 0.000001):
        normalized["151"] = AopEntry(
            aop="151",
            current_year=-abs(loss),
            previous_year=profit_entry.previous_year if profit_entry else None,
        )
        normalized["152"] = AopEntry(
            aop="152",
            current_year=0.0,
            previous_year=loss_entry.previous_year,
        )
    return normalized

def _months_in_period(period_from, period_to):
    """Izračuna število mesecev poslovanja iz obdobja (vključno oba robna meseca) -
    npr. 1.1.2026-30.6.2026 -> 6, 1.1.2025-31.12.2025 -> 12. Vrne None, če datumov
    ni mogoče razčleniti."""
    try:
        d1, m1, y1 = [int(x) for x in str(period_from).strip().split('.')[:3]]
        d2, m2, y2 = [int(x) for x in str(period_to).strip().split('.')[:3]]
        n = (y2 - y1) * 12 + (m2 - m1) + 1
        return n if 1 <= n <= 12 else None
    except Exception:
        return None


def export_excel(result, output_path):
    """
    Export v CBK template.

    NAMERNO NE SPREMINJAMO OBLIKOVANJA CELIC.
    Samo kopiramo originalni template in vpišemo vrednosti v že obstoječe celice.
    Datumske celice dobijo pravi Excel DATE (Python date), vendar se njihov obstoječi
    number_format/stil ne spreminja - zato ostane prikaz tak, kot je nastavljen v template-u.
    """
    # Tip subjekta 2 = s.p. → poseben template; 4 = društvo → poseben template; ostali standardni
    tip_subjekta = str(getattr(result.company, 'tip_subjekta', '1') or '1').strip()
    if tip_subjekta == "4" and TEMPLATE_DRUSTVO_PATH.exists():
        tmpl = TEMPLATE_DRUSTVO_PATH
    elif tip_subjekta == "2" and TEMPLATE_SP_PATH.exists():
        tmpl = TEMPLATE_SP_PATH
    else:
        tmpl = TEMPLATE_PATH
    if tip_subjekta == "2" and not TEMPLATE_SP_PATH.exists():
        result.warnings.append("Izbran je tip subjekta 2 (s.p.), ampak template_sp.xlsx ni najden. Uporabljen je standardni template.")
    if tip_subjekta == "4" and not TEMPLATE_DRUSTVO_PATH.exists():
        result.warnings.append("Izbran je tip subjekta 4 (društvo), ampak template_drustvo.xlsx ni najden. Uporabljen je standardni template.")

    shutil.copy2(tmpl, output_path)
    wb = openpyxl.load_workbook(output_path)
    c = result.company

    def _as_int_or_original(v):
        try:
            return int(str(v).strip())
        except Exception:
            return v

    ws = wb["BS"]
    ws.cell(3,4).value  = c.name
    # Matična mora ostati tekst, da se ohranijo vodilne ničle.
    ws.cell(4,4).value  = str(c.registration_number or "")
    ws.cell(6,4).value  = c.obdobje_bilance
    # Šifre morajo biti numerične, ne tekstovne.
    ws.cell(5,7).value  = _as_int_or_original(c.tip_bilance)
    ws.cell(6,7).value  = _as_int_or_original(c.tip_subjekta)
    # Vrednost je pravi Excel DATE; obstoječi format celice iz template-a ostane nedotaknjen.
    ws.cell(11,4).value = _to_excel_date(c.period_to)
    # C8 ("Podatki iz bilance stanja" datum) je bilo prej treba vpisati ročno — zdaj ga
    # napolnimo samodejno z istim datumom kot D11.
    ws.cell(8,3).value  = _to_excel_date(c.period_to)

    idx = build_idx(ws); bs_n = 0
    for code in BS_AOP:
        if code not in idx or code not in result.aop_data:
            continue
        val = result.aop_data[code].current_year
        if val is not None:
            ws.cell(idx[code],4).value = val
            bs_n += 1
    bs_n += _fill_combined_rows(ws, result.aop_data, _COMBINED_ROWS_BS)
    if result.gvaop_data:
        gidx = build_idx_by_gvaop(ws)
        for gvaop, val in result.gvaop_data.items():
            row = gidx.get(gvaop)
            if row and val is not None:
                ws.cell(row, 4).value = val
                bs_n += 1

    ws2 = wb["IPI"]
    ws2.cell(3,4).value = c.name
    ws2.cell(4,4).value = str(c.registration_number or "")
    if c.period_from and c.period_to:
        ws2.cell(8,3).value = f"{c.period_from} - {c.period_to}"
    ws2.cell(11,4).value = _to_excel_date(c.period_to)

    idx2 = build_idx(ws2); ipi_n = 0
    # Bonitetni sistem izgubo iz poslovanja prikaže kot negativen AOP 151;
    # AOP 152 je zato 0. Celotni in čisti dobiček/izguba ostaneta nespremenjena.
    ipi_aop_data = _normalize_ipi_operating_loss_for_cbk(result.aop_data)
    # Število mesecev poslovanja (AOP 189) izračunamo vedno sami iz dejanskega obdobja
    # (period_from/period_to) - ne zanašamo se na izvorno datoteko, ker ta podatek
    # marsikje manjka ali pa ga format sploh ne vsebuje (npr. medletni izvozi).
    meseci = _months_in_period(c.period_from, c.period_to)
    if meseci is not None:
        ipi_aop_data["189"] = AopEntry(aop="189", current_year=float(meseci), previous_year=None)

    for code in IPI_AOP:
        if code not in idx2 or code not in ipi_aop_data:
            continue
        val = ipi_aop_data[code].current_year
        if val is not None:
            ws2.cell(idx2[code],4).value = val
            ipi_n += 1
    ipi_n += _fill_combined_rows(ws2, ipi_aop_data, _COMBINED_ROWS_IPI)
    if result.gvaop_data:
        gidx2 = build_idx_by_gvaop(ws2)
        for gvaop, val in result.gvaop_data.items():
            row = gidx2.get(gvaop)
            if row and val is not None:
                ws2.cell(row, 4).value = val
                ipi_n += 1

    wb.save(output_path)
    return bs_n, ipi_n

# ── HTML ──────────────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="sl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CBK Pretvornik</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:ital,opsz,wght@0,14..32,300;0,14..32,400;0,14..32,500;0,14..32,600;0,14..32,700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root {
  --bg:        #f4f8ff;
  --surface:   #ffffff;
  --surface2:  #eef5ff;
  --border:    #dbe7f6;
  --border2:   #b7c9e2;
  --text:      #001b3d;
  --text2:     #4b607c;
  --text3:     #7d8fa8;
  --indigo:    #2f76f6;
  --indigo-lt: #e8f1ff;
  --indigo-md: #b8d2ff;
  --green:     #0f9f82;
  --green-lt:  #e9fbf7;
  --green-md:  #a8eadf;
  --amber:     #c47a00;
  --amber-lt:  #fff7e6;
  --red:       #d92d20;
  --red-lt:    #fff0f0;
  --radius-sm: 8px;
  --radius:    12px;
  --radius-lg: 16px;
  --shadow-sm: 0 1px 2px rgba(0,0,0,.05);
  --shadow:    0 1px 3px rgba(0,0,0,.07), 0 4px 16px rgba(0,0,0,.04);
  --shadow-lg: 0 4px 6px rgba(0,0,0,.04), 0 12px 32px rgba(0,0,0,.07);
}

*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

body {
  font-family: 'Inter', -apple-system, sans-serif;
  background: var(--bg);
  color: var(--text);
  min-height: 100vh;
  display: flex;
  flex-direction: column;
  align-items: center;
  justify-content: center;
  padding: 24px;
  -webkit-font-smoothing: antialiased;
}

/* Subtle noise texture on bg */
body::before {
  content: '';
  position: fixed;
  inset: 0;
  background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)' opacity='0.018'/%3E%3C/svg%3E");
  pointer-events: none;
  z-index: 0;
}

.shell {
  width: 100%;
  max-width: 460px;
  position: relative;
  z-index: 1;
}

/* ── WORDMARK ── */
.wordmark {
  display: flex;
  align-items: center;
  gap: 10px;
  margin-bottom: 20px;
  padding: 0 4px;
}

.wm-logo {
  flex-shrink: 0;
  line-height: 0;
  filter: drop-shadow(0 3px 10px rgba(47,118,246,0.22));
}

.wm-logo svg { width: 36px; height: 36px; display: block; }

.wm-name {
  display: flex;
  flex-direction: column;
  gap: 1px;
}

.wm-text {
  font-size: 15px;
  font-weight: 700;
  color: var(--text);
  letter-spacing: -0.03em;
  line-height: 1;
}

.wm-sub {
  font-size: 10.5px;
  color: var(--text3);
  font-weight: 400;
  letter-spacing: 0.02em;
}

.wm-badge {
  margin-left: auto;
  font-size: 10.5px;
  font-weight: 500;
  color: var(--text3);
  background: var(--surface2);
  border: 1px solid var(--border);
  padding: 2px 8px;
  border-radius: 20px;
  letter-spacing: 0.02em;
  flex-shrink: 0;
}

/* ── CARD ── */
.card {
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: var(--radius-lg);
  padding: 28px;
  box-shadow: var(--shadow-lg);
}

/* ── STEP HEADER ── */
.step {
  display: flex;
  align-items: center;
  gap: 8px;
  margin-top: 20px;
  margin-bottom: 9px;
}

.step:first-of-type { margin-top: 0; }

.step-num {
  width: 18px; height: 18px;
  border-radius: 50%;
  background: var(--text);
  color: white;
  font-size: 10px;
  font-weight: 700;
  display: flex; align-items: center; justify-content: center;
  flex-shrink: 0;
  letter-spacing: 0;
}

.step-label {
  font-size: 11.5px;
  font-weight: 600;
  color: var(--text2);
  text-transform: uppercase;
  letter-spacing: 0.06em;
}

/* ── SEGMENTED CONTROL (tip + obdobje) ── */
.seg {
  display: flex;
  background: var(--surface2);
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  padding: 3px;
  gap: 2px;
  margin-bottom: 4px;
}

.seg-btn {
  flex: 1;
  padding: 8px 6px;
  border-radius: 6px;
  border: none;
  background: transparent;
  cursor: pointer;
  text-align: center;
  transition: all .15s ease;
  font-family: 'Inter', sans-serif;
}

.seg-btn:hover { background: var(--border); }

.seg-btn.selected {
  background: var(--surface);
  box-shadow: var(--shadow-sm), 0 0 0 1px var(--border2);
}

.seg-code {
  display: block;
  font-family: 'JetBrains Mono', monospace;
  font-size: 18px;
  font-weight: 500;
  color: var(--text2);
  line-height: 1.1;
  margin-bottom: 3px;
  transition: color .15s;
}

.seg-btn.selected .seg-code { color: var(--text); }

.seg-name {
  font-size: 10px;
  font-weight: 500;
  color: var(--text3);
  letter-spacing: 0.02em;
  transition: color .15s;
}

.seg-btn.selected .seg-name { color: var(--text2); }

/* Obdobje segmented */
.seg-2 .seg-btn { padding: 10px 12px; text-align: left; }

.seg-main {
  display: block;
  font-family: 'JetBrains Mono', monospace;
  font-size: 12.5px;
  font-weight: 500;
  color: var(--text2);
  margin-bottom: 2px;
  transition: color .15s;
}

.seg-sub {
  display: block;
  font-size: 10.5px;
  color: var(--text3);
  transition: color .15s;
}

.seg-2 .seg-btn.selected .seg-main { color: var(--text); }
.seg-2 .seg-btn.selected .seg-sub  { color: var(--text2); }

/* ── DROP ZONE ── */
.drop-zone {
  border: 1.5px dashed var(--border2);
  border-radius: var(--radius);
  padding: 24px 20px;
  text-align: center;
  cursor: pointer;
  transition: all .2s ease;
  background: var(--surface2);
}

.drop-zone:hover, .drop-zone.drag {
  border-color: var(--indigo);
  background: var(--indigo-lt);
}

.drop-zone.has-file {
  border-color: var(--green);
  border-style: solid;
  background: var(--green-lt);
}

.drop-icon {
  width: 38px; height: 38px;
  border-radius: 9px;
  background: var(--border);
  display: flex; align-items: center; justify-content: center;
  margin: 0 auto 10px;
  transition: all .2s;
}

.drop-zone:hover .drop-icon, .drop-zone.drag .drop-icon {
  background: var(--indigo);
}

.drop-zone.has-file .drop-icon {
  background: var(--green);
}

.drop-icon svg {
  width: 17px; height: 17px;
  stroke: var(--text3); fill: none;
  stroke-width: 2; stroke-linecap: round; stroke-linejoin: round;
  transition: stroke .2s;
}

.drop-zone:hover .drop-icon svg,
.drop-zone.drag .drop-icon svg,
.drop-zone.has-file .drop-icon svg { stroke: white; }

.drop-title {
  font-size: 13.5px;
  font-weight: 600;
  color: var(--text);
  margin-bottom: 3px;
}

.drop-desc { font-size: 12px; color: var(--text3); }

.drop-hint {
  font-size: 11.5px;
  color: var(--text3);
  margin-top: 8px;
  margin-bottom: 12px;
}

.drop-hint a {
  color: var(--indigo);
  cursor: pointer;
  text-decoration: none;
  font-weight: 500;
}

.drop-hint a:hover { text-decoration: underline; }

.tags {
  display: flex;
  gap: 4px;
  justify-content: center;
  flex-wrap: wrap;
  margin-top: 9px;
}

.tag {
  padding: 2px 7px;
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: 20px;
  font-size: 10.5px;
  color: var(--text3);
  font-family: 'JetBrains Mono', monospace;
}

input[type=file] { display: none; }

/* ── FILE LIST ── */
.file-list { margin-bottom: 14px; }

.file-item {
  display: flex;
  align-items: center;
  gap: 9px;
  padding: 8px 11px;
  background: var(--surface2);
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  margin-bottom: 4px;
}

.file-item-icon {
  width: 26px; height: 26px;
  border-radius: 6px;
  background: var(--green-lt);
  border: 1px solid var(--green-md);
  display: flex; align-items: center; justify-content: center;
  flex-shrink: 0;
}

.file-item-icon svg {
  width: 12px; height: 12px;
  stroke: var(--green); fill: none;
  stroke-width: 2.2; stroke-linecap: round; stroke-linejoin: round;
}

.file-item-name {
  flex: 1;
  font-size: 12.5px;
  font-weight: 500;
  color: var(--text);
  overflow: hidden;
  text-overflow: ellipsis;
  white-space: nowrap;
}

.file-item-size {
  font-size: 11px;
  color: var(--text3);
  font-family: 'JetBrains Mono', monospace;
  flex-shrink: 0;
}

.file-item-remove {
  width: 20px; height: 20px;
  border-radius: 50%;
  border: 1px solid var(--border);
  background: transparent;
  cursor: pointer;
  display: flex; align-items: center; justify-content: center;
  flex-shrink: 0;
  transition: all .15s;
}

.file-item-remove:hover { background: var(--red-lt); border-color: #f8b4b4; }
.file-item-remove:hover svg { stroke: var(--red); }

.file-item-remove svg {
  width: 10px; height: 10px;
  stroke: var(--text3); fill: none;
  stroke-width: 2.5; stroke-linecap: round;
}

/* ── DIVIDER ── */
.rule {
  height: 1px;
  background: var(--border);
  margin: 22px 0;
}

/* ── BTN ── */
.btn {
  width: 100%;
  padding: 12px;
  background: var(--text);
  color: white;
  border: none;
  border-radius: var(--radius-sm);
  font-size: 13.5px;
  font-weight: 600;
  font-family: 'Inter', sans-serif;
  cursor: pointer;
  transition: all .15s ease;
  display: flex;
  align-items: center;
  justify-content: center;
  gap: 7px;
  letter-spacing: -0.01em;
}

.btn:hover { background: #002b5c; transform: translateY(-1px); box-shadow: 0 4px 14px rgba(47,118,246,0.22); }
.btn:active { transform: none; box-shadow: none; }
.btn:disabled { background: var(--border2); color: var(--text3); cursor: not-allowed; transform: none; box-shadow: none; }

.btn svg {
  width: 15px; height: 15px;
  stroke: currentColor; fill: none;
  stroke-width: 2; stroke-linecap: round; stroke-linejoin: round;
}

.spinner {
  width: 15px; height: 15px;
  border: 2px solid rgba(255,255,255,.25);
  border-top-color: white;
  border-radius: 50%;
  animation: spin .65s linear infinite;
  display: none;
}

@keyframes spin { to { transform: rotate(360deg); } }

/* ── LOGIN ── */
.login-eyebrow {
  font-size: 11px;
  font-weight: 600;
  color: var(--text3);
  text-transform: uppercase;
  letter-spacing: 0.08em;
  margin-bottom: 6px;
}

.login-title {
  font-size: 22px;
  font-weight: 700;
  color: var(--text);
  letter-spacing: -0.03em;
  margin-bottom: 4px;
}

.login-sub {
  font-size: 13px;
  color: var(--text3);
  margin-bottom: 22px;
  line-height: 1.5;
}

.field-label {
  font-size: 12px;
  font-weight: 600;
  color: var(--text);
  display: block;
  margin-bottom: 6px;
  letter-spacing: -0.01em;
}

input[type=password] {
  width: 100%;
  padding: 10px 13px;
  background: var(--surface);
  border: 1px solid var(--border2);
  border-radius: var(--radius-sm);
  font-size: 14px;
  font-family: 'Inter', sans-serif;
  color: var(--text);
  outline: none;
  transition: all .15s;
  margin-bottom: 14px;
  box-shadow: var(--shadow-sm);
}

input[type=password]:focus {
  border-color: var(--text);
  box-shadow: 0 0 0 3px rgba(47,118,246,.14);
}

select {
  width: 100%;
  padding: 10px 34px 10px 12px;
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  font-size: 12.5px;
  font-family: 'Inter', sans-serif;
  color: var(--text);
  outline: none;
  box-shadow: var(--shadow-sm);
  appearance: none;
  -webkit-appearance: none;
  -moz-appearance: none;
  background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%237d8fa8' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='6 9 12 15 18 9'/%3E%3C/svg%3E");
  background-repeat: no-repeat;
  background-position: right 12px center;
  background-size: 15px 15px;
}

select:focus {
  border-color: var(--text);
  box-shadow: 0 0 0 3px rgba(47,118,246,.14);
}

/* ── MONTH PICKER (medletna bilanca) ── */
.month-picker-wrap { position: relative; margin-bottom: 4px; }

.month-picker-trigger {
  width: 100%;
  padding: 10px 12px;
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  font-size: 12.5px;
  font-family: 'Inter', sans-serif;
  color: var(--text);
  outline: none;
  box-shadow: var(--shadow-sm);
  cursor: pointer;
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 8px;
  transition: all .15s ease;
}

.month-picker-trigger:hover { background: var(--surface2); }

.month-picker-trigger.open {
  border-color: var(--text);
  box-shadow: 0 0 0 3px rgba(47,118,246,.14);
}

.month-picker-value { color: var(--text3); }
.month-picker-value.is-set { color: var(--text); font-weight: 500; }

.month-picker-trigger svg {
  width: 15px; height: 15px;
  stroke: var(--text3); fill: none;
  stroke-width: 2; stroke-linecap: round; stroke-linejoin: round;
  flex-shrink: 0;
  transition: transform .15s ease;
}

.month-picker-trigger.open svg { transform: rotate(180deg); }

.month-picker-panel {
  display: none;
  position: absolute;
  top: calc(100% + 6px);
  left: 0; right: 0;
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  box-shadow: var(--shadow-lg);
  padding: 12px;
  z-index: 20;
}

.month-picker-panel.open { display: block; }

.month-picker-header {
  display: flex;
  align-items: center;
  justify-content: space-between;
  margin-bottom: 10px;
}

.month-picker-year {
  font-size: 13px;
  font-weight: 700;
  color: var(--text);
  font-family: 'JetBrains Mono', monospace;
  letter-spacing: 0.02em;
}

.month-picker-nav {
  width: 26px; height: 26px;
  border-radius: 6px;
  border: 1px solid var(--border);
  background: var(--surface2);
  color: var(--text2);
  font-size: 15px;
  line-height: 1;
  cursor: pointer;
  display: flex; align-items: center; justify-content: center;
  transition: all .15s ease;
}

.month-picker-nav:hover:not(:disabled) { background: var(--border); }
.month-picker-nav:disabled { opacity: .3; cursor: not-allowed; }

.month-picker-grid {
  display: grid;
  grid-template-columns: repeat(4, 1fr);
  gap: 5px;
}

.month-picker-cell {
  padding: 8px 4px;
  border-radius: 6px;
  border: none;
  background: var(--surface2);
  color: var(--text2);
  font-size: 12px;
  font-weight: 500;
  cursor: pointer;
  text-align: center;
  transition: all .15s ease;
  font-family: 'Inter', sans-serif;
}

.month-picker-cell:hover:not(:disabled) { background: var(--border); color: var(--text); }

.month-picker-cell.today { box-shadow: inset 0 0 0 1.5px var(--border2); }

.month-picker-cell.selected { background: var(--text); color: white; }
.month-picker-cell.selected:hover { background: var(--text); }

.month-picker-cell:disabled {
  opacity: .3;
  cursor: not-allowed;
}

/* ── MESSAGES ── */
.msg {
  padding: 10px 13px;
  border-radius: var(--radius-sm);
  font-size: 12.5px;
  margin-top: 11px;
  border: 1px solid;
  display: none;
  line-height: 1.5;
}

.msg.error { background: var(--red-lt); border-color: #f8b4b4; color: var(--red); }

/* ── RESULT ── */
.result-box {
  display: none;
  margin-top: 14px;
  border: 1px solid var(--green-md);
  border-radius: var(--radius);
  overflow: hidden;
  box-shadow: var(--shadow-sm);
}

.result-header {
  padding: 11px 15px;
  background: var(--green-lt);
  border-bottom: 1px solid var(--green-md);
  display: flex;
  align-items: center;
  gap: 7px;
}

.result-header svg {
  width: 14px; height: 14px;
  stroke: var(--green); fill: none;
  stroke-width: 2.5; stroke-linecap: round; stroke-linejoin: round;
}

.result-header-text { font-size: 12.5px; font-weight: 600; color: var(--green); }

.result-body { padding: 13px 15px; background: var(--surface); }

.result-row {
  display: flex;
  justify-content: space-between;
  align-items: baseline;
  padding: 5px 0;
  border-bottom: 1px solid var(--border);
  font-size: 12.5px;
  gap: 12px;
}

.result-row:last-child { border-bottom: none; }
.result-label { color: var(--text3); flex-shrink: 0; }

.result-val {
  font-weight: 500;
  font-family: 'JetBrains Mono', monospace;
  font-size: 11.5px;
  color: var(--text);
  text-align: right;
  word-break: break-all;
}

.warn-note {
  margin-top: 11px;
  font-size: 12px;
  color: var(--amber);
  background: var(--amber-lt);
  border: 1px solid #ffd58a;
  border-radius: var(--radius-sm);
  padding: 8px 12px;
  line-height: 1.5;
}

/* ── FOOTER ── */
.footer {
  margin-top: 16px;
  text-align: center;
  font-size: 11px;
  color: var(--text3);
  letter-spacing: 0.02em;
}

@media (max-width: 500px) {
  body { padding: 16px; }
  .card { padding: 22px 18px; }
}
</style>
</head>
<body>
<div class="shell">

  <div class="wordmark">
    <div class="wm-logo">
      <!-- CBK logo mark: stylized chart bars with upward arrow -->
      <svg viewBox="0 0 36 36" fill="none" xmlns="http://www.w3.org/2000/svg">
        <!-- Background -->
        <rect width="36" height="36" rx="9" fill="#001b3d"/>
        <!-- Bar 1 -->
        <rect x="7" y="20" width="5" height="9" rx="1.5" fill="white" opacity="0.45"/>
        <!-- Bar 2 -->
        <rect x="15.5" y="14" width="5" height="15" rx="1.5" fill="white" opacity="0.75"/>
        <!-- Bar 3 (tallest) -->
        <rect x="24" y="8" width="5" height="21" rx="1.5" fill="white"/>
        <!-- Upward tick on bar 3 -->
        <path d="M25 11 L26.5 8.5 L28 11" stroke="white" stroke-width="1.4" stroke-linecap="round" stroke-linejoin="round" fill="none"/>
      </svg>
    </div>
    <div class="wm-name">
      <span class="wm-text">CBK Pretvornik</span>
      <span class="wm-sub">Analiza bilanc</span>
    </div>
    <span class="wm-badge">Interno</span>
  </div>

  <div class="card">

    <!-- LOGIN -->
    <div id="login-section">
      <div class="login-eyebrow">CBK Pretvornik &middot; Interni dostop</div>
      <div class="login-title">Prijava</div>
      <div class="login-sub">Vpišite geslo za dostop do orodja za pretvorbo bilanc.</div>
      <label class="field-label">Geslo</label>
      <input type="password" id="password" placeholder="Vpiši geslo" onkeydown="if(event.key==='Enter')login()">
      <button class="btn" onclick="login()">
        <svg viewBox="0 0 24 24"><path d="M15 3h4a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2h-4"/>
        <polyline points="10 17 15 12 10 7"/><line x1="15" y1="12" x2="3" y2="12"/></svg>
        Prijava
      </button>
      <div class="msg error" id="login-error">Napačno geslo — poskusite znova.</div>
    </div>

    <!-- APP -->
    <div id="app-section" style="display:none;">

      <div class="step">
        <div class="step-num">1</div>
        <div class="step-label">Tip bilance</div>
      </div>
      <select id="tip-bilance-select" onchange="selectTipValue(this.value)">
        <option value="3">3 — Revidirana</option>
        <option value="4">4 — Zaključena nerevidirana</option>
        <option value="5">5 — Konsolidirana revidirana</option>
        <option value="6">6 — Konsolidirana nerevidirana</option>
        <option value="7">7 — Preliminarna</option>
        <option value="8" selected>8 — Zaključena</option>
      </select>

      <div id="medletni-datum-field" style="display:none;">
        <label class="field-label">Na kateri datum je medletna bilanca narejena?</label>
        <div class="month-picker-wrap" id="month-picker-wrap">
          <button type="button" class="month-picker-trigger" id="month-picker-trigger" onclick="toggleMonthPicker()">
            <span class="month-picker-value" id="month-picker-value">Izberi mesec</span>
            <svg viewBox="0 0 24 24"><polyline points="6 9 12 15 18 9"/></svg>
          </button>
          <div class="month-picker-panel" id="month-picker-panel">
            <div class="month-picker-header">
              <button type="button" class="month-picker-nav" id="month-picker-prev" onclick="changeMonthPickerYear(-1)">&#8249;</button>
              <span class="month-picker-year" id="month-picker-year"></span>
              <button type="button" class="month-picker-nav" id="month-picker-next" onclick="changeMonthPickerYear(1)">&#8250;</button>
            </div>
            <div class="month-picker-grid" id="month-picker-grid"></div>
          </div>
        </div>
        <div class="drop-hint">Uporabljen bo zadnji dan izbranega meseca (npr. april &rarr; 30.04.).</div>
      </div>

      <div class="step">
        <div class="step-num">2</div>
        <div class="step-label">Tip subjekta</div>
      </div>
      <select id="tip-subjekta-select" onchange="selectTipSubjekta(this.value)">
        <option value="1" selected>1 — Gospodarske družbe</option>
        <option value="2">2 — Samostojni podjetniki</option>
        <option value="3">3 — Zadruge</option>
        <option value="4">4 — Društva</option>
        <option value="5">5 — Pravne osebe javnega prava</option>
        <option value="6">6 — Pravne osebe zasebnega prava</option>
        <option value="7">7 — Banke</option>
        <option value="8">8 — Zavarovalnice</option>
        <option value="9">9 — Druge osebe javnega prava</option>
      </select>

      <div class="step">
        <div class="step-num">3</div>
        <div class="step-label">Poslovno obdobje</div>
      </div>
      <div class="seg seg-2">
        <button class="seg-btn selected" onclick="selectObdobje(this,'01.01-31.12')">
          <span class="seg-main">01.01. &#8211; 31.12.</span>
          <span class="seg-sub">Standardno leto</span>
        </button>
        <button class="seg-btn" onclick="selectObdobje(this,'01.04-31.03')">
          <span class="seg-main">01.04. &#8211; 31.03.</span>
          <span class="seg-sub">Nestandardno leto</span>
        </button>
      </div>

      <div class="step">
        <div class="step-num">4</div>
        <div class="step-label">Datoteka(-e)</div>
      </div>
      <div class="drop-zone" id="drop-zone"
           onclick="document.getElementById('file-input').click()"
           ondragover="ev(event,'drag')" ondragleave="ev(event,'')" ondrop="drop(event)">
        <div class="drop-icon">
          <svg viewBox="0 0 24 24"><polyline points="16 16 12 12 8 16"/>
          <line x1="12" y1="12" x2="12" y2="21"/>
          <path d="M20.39 18.39A5 5 0 0 0 18 9h-1.26A8 8 0 1 0 3 16.3"/></svg>
        </div>
        <div class="drop-title">Povleci datoteko sem</div>
        <div class="drop-desc">ali klikni za izbiro &mdash; PDF ali Excel</div>
        <div class="tags">
          <span class="tag">lp2025</span>
          <span class="tag">JOLP</span>
          <span class="tag">NAPOVED</span>
          <span class="tag">xlsx</span>
          <span class="tag">xls</span>
        </div>
      </div>
      <input type="file" id="file-input" accept=".pdf,.xlsx,.xls" multiple onchange="filesSelected(this.files)">
      <div class="drop-hint">Za ločena BS + IPI &mdash; <a onclick="document.getElementById('file-input').click()">dodaj oba hkrati</a></div>

      <div class="file-list" id="file-list"></div>

      <button class="btn" id="convert-btn" onclick="convert()" disabled>
        <div class="spinner" id="spinner"></div>
        <svg id="btn-icon" viewBox="0 0 24 24"><polyline points="23 4 23 10 17 10"/>
        <polyline points="1 20 1 14 7 14"/>
        <path d="M3.51 9a9 9 0 0 1 14.85-3.36L23 10M1 14l4.64 4.36A9 9 0 0 0 20.49 15"/></svg>
        <span id="btn-text">Pretvori v Excel</span>
      </button>

      <div class="msg error" id="error-msg"></div>

      <div class="result-box" id="result-box">
        <div class="result-header">
          <svg viewBox="0 0 24 24"><path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"/>
          <polyline points="22 4 12 14.01 9 11.01"/></svg>
          <span class="result-header-text">Pretvorba uspešna &mdash; prenos se začenja</span>
        </div>
        <div class="result-body">
          <div class="result-row"><span class="result-label">Podjetje</span><span class="result-val" id="r-name"></span></div>
          <div class="result-row"><span class="result-label">Matična</span><span class="result-val" id="r-reg"></span></div>
          <div class="result-row"><span class="result-label">Obdobje</span><span class="result-val" id="r-period"></span></div>
          <div class="result-row"><span class="result-label">Tip bilance</span><span class="result-val" id="r-tip"></span></div>
          <div class="result-row"><span class="result-label">Tip subjekta</span><span class="result-val" id="r-subj"></span></div>
          <div class="result-row"><span class="result-label">Format</span><span class="result-val" id="r-fmt"></span></div>
          <div class="result-row"><span class="result-label">BS vrstice</span><span class="result-val" id="r-bs"></span></div>
          <div class="result-row"><span class="result-label">IPI vrstice</span><span class="result-val" id="r-ipi"></span></div>
          <div class="warn-note" id="warn-partner">&#9888; Ročno vnesi <strong>Št. partnerja</strong> (celica D5 v BS sheetu)</div>
        <div class="warn-note" id="warn-drustvo" style="display:none;background:#fef3c7;border-color:#fcd34d;color:#92400e;">&#9888; Tip subjekta se zapiše v izhodni Excel; za s.p. se uporabi poseben template.</div>
        </div>
      </div>

    </div>
  </div>

  <div class="footer">CBK Pretvornik &middot; Samo za pooblaščene uporabnike</div>
</div>

<script>
let selectedFiles = [];
let selectedTip = '8';
let selectedTipSubjekta = '1';
let selectedObdobje = '01.01-31.12';

function login() {
  const pw = document.getElementById('password').value.trim();
  fetch('/auth',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({password:pw})})
  .then(r=>r.json()).then(d=>{
    if(d.ok){
      document.getElementById('login-section').style.display='none';
      document.getElementById('app-section').style.display='block';
    } else {
      document.getElementById('login-error').style.display='block';
    }
  }).catch(e=>{
    document.getElementById('login-error').textContent='Napaka: '+e.message;
    document.getElementById('login-error').style.display='block';
  });
}

let selectedMedletniMesec='';

const MP_MONTH_NAMES = ['januar','februar','marec','april','maj','junij','julij','avgust','september','oktober','november','december'];
const MP_MONTH_ABBR  = ['jan.','feb.','mar.','apr.','maj','jun.','jul.','avg.','sep.','okt.','nov.','dec.'];
const MP_TODAY = new Date();
let mpPanelYear = MP_TODAY.getFullYear();

function selectTip(el,val){ selectedTip=val; toggleMedletniDatumField(); }
function selectTipValue(val){ selectedTip=val; toggleMedletniDatumField(); }
function selectTipSubjekta(val){ selectedTipSubjekta=val; }

function toggleMedletniDatumField(){
  const wrap=document.getElementById('medletni-datum-field');
  if(selectedTip==='7'){
    wrap.style.display='block';
  } else {
    wrap.style.display='none';
    closeMonthPicker();
    selectedMedletniMesec='';
    const val=document.getElementById('month-picker-value');
    val.textContent='Izberi mesec';
    val.classList.remove('is-set');
  }
}

function toggleMonthPicker(){
  const panel=document.getElementById('month-picker-panel');
  const trigger=document.getElementById('month-picker-trigger');
  if(panel.classList.contains('open')){
    closeMonthPicker();
    return;
  }
  mpPanelYear = selectedMedletniMesec ? parseInt(selectedMedletniMesec.split('-')[0],10) : MP_TODAY.getFullYear();
  renderMonthPickerGrid();
  panel.classList.add('open');
  trigger.classList.add('open');
}

function closeMonthPicker(){
  document.getElementById('month-picker-panel').classList.remove('open');
  document.getElementById('month-picker-trigger').classList.remove('open');
}

function changeMonthPickerYear(delta){
  mpPanelYear += delta;
  renderMonthPickerGrid();
}

function renderMonthPickerGrid(){
  document.getElementById('month-picker-year').textContent = mpPanelYear;
  document.getElementById('month-picker-next').disabled = mpPanelYear >= MP_TODAY.getFullYear();

  const grid=document.getElementById('month-picker-grid');
  grid.innerHTML='';

  const selYear  = selectedMedletniMesec ? parseInt(selectedMedletniMesec.split('-')[0],10) : null;
  const selMonth = selectedMedletniMesec ? parseInt(selectedMedletniMesec.split('-')[1],10) : null;
  const curYear  = MP_TODAY.getFullYear();
  const curMonth = MP_TODAY.getMonth()+1;

  for(let m=1;m<=12;m++){
    const btn=document.createElement('button');
    btn.type='button';
    btn.className='month-picker-cell';
    btn.textContent=MP_MONTH_ABBR[m-1];

    const isFuture = (mpPanelYear>curYear) || (mpPanelYear===curYear && m>curMonth);
    if(isFuture){ btn.disabled=true; }
    if(mpPanelYear===curYear && m===curMonth){ btn.classList.add('today'); }
    if(selYear===mpPanelYear && selMonth===m){ btn.classList.add('selected'); }

    btn.onclick=()=>pickMonth(mpPanelYear,m);
    grid.appendChild(btn);
  }
}

function pickMonth(year,month){
  const mm=String(month).padStart(2,'0');
  selectedMedletniMesec = year+'-'+mm;

  const val=document.getElementById('month-picker-value');
  val.textContent = MP_MONTH_NAMES[month-1]+' '+year;
  val.classList.add('is-set');

  closeMonthPicker();
}

document.addEventListener('click', function(e){
  const wrap=document.getElementById('month-picker-wrap');
  if(wrap && !wrap.contains(e.target)){ closeMonthPicker(); }
});

function selectObdobje(el,val){
  selectedObdobje=val;
  document.querySelectorAll('.seg-2 .seg-btn').forEach(b=>b.classList.remove('selected'));
  el.classList.add('selected');
}

function ev(e,cls){
  e.preventDefault();
  const z=document.getElementById('drop-zone');
  z.className='drop-zone'+(cls?' '+cls:(selectedFiles.length?' has-file':''));
}

function drop(e){
  e.preventDefault();
  filesSelected(e.dataTransfer.files);
}

function filesSelected(fileList){
  for(let f of fileList){
    const ext=f.name.toLowerCase();
    if(!ext.endsWith('.pdf')&&!ext.endsWith('.xlsx')&&!ext.endsWith('.xls')){
      showError('Dovoljeni so samo PDF ali Excel (.xlsx/.xls) fajli.'); return;
    }
    if(!selectedFiles.find(x=>x.name===f.name)) selectedFiles.push(f);
  }
  renderFileList();
  document.getElementById('error-msg').style.display='none';
  document.getElementById('result-box').style.display='none';
  document.getElementById('drop-zone').className='drop-zone'+(selectedFiles.length?' has-file':'');
  document.getElementById('convert-btn').disabled=selectedFiles.length===0;
}

function removeFile(idx){
  selectedFiles.splice(idx,1);
  renderFileList();
  document.getElementById('drop-zone').className='drop-zone'+(selectedFiles.length?' has-file':'');
  document.getElementById('convert-btn').disabled=selectedFiles.length===0;
}

function renderFileList(){
  const list=document.getElementById('file-list');
  if(!selectedFiles.length){list.innerHTML='';return;}
  list.innerHTML=selectedFiles.map((f,i)=>`
    <div class="file-item">
      <div class="file-item-icon">
        <svg viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/>
        <polyline points="14 2 14 8 20 8"/></svg>
      </div>
      <span class="file-item-name" title="${f.name}">${f.name}</span>
      <span class="file-item-size">${(f.size/1024/1024).toFixed(1)}&thinsp;MB</span>
      <button class="file-item-remove" onclick="removeFile(${i})">
        <svg viewBox="0 0 24 24"><line x1="18" y1="6" x2="6" y2="18"/><line x1="6" y1="6" x2="18" y2="18"/></svg>
      </button>
    </div>`).join('');
}

function showError(msg){
  const e=document.getElementById('error-msg');
  e.textContent=msg; e.style.display='block';
}

const TIP_LABELS={'3':'3 — Revidirana (ZR)','4':'4 — Zaključena nerevidirana (ZR)','5':'5 — Konsolidirana revidirana (ZR)','6':'6 — Konsolidirana nerevidirana (ZR)','7':'7 — Preliminarna (ML)','8':'8 — Zaključena (ZR)'};
const SUBJEKT_LABELS={'1':'1 — Gospodarske družbe','2':'2 — Samostojni podjetniki','3':'3 — Zadruge','4':'4 — Društva','5':'5 — Pravne osebe javnega prava','6':'6 — Pravne osebe zasebnega prava','7':'7 — Banke','8':'8 — Zavarovalnice','9':'9 — Druge osebe javnega prava'};

async function convert(){
  if(!selectedFiles.length) return;
  if(selectedTip==='7' && !selectedMedletniMesec){
    showError('Za preliminarno (medletno) bilanco izberi mesec, na katerega je bilanca narejena.');
    return;
  }
  const btn=document.getElementById('convert-btn');
  const spinner=document.getElementById('spinner');
  const icon=document.getElementById('btn-icon');
  const btnText=document.getElementById('btn-text');
  btn.disabled=true; spinner.style.display='block'; icon.style.display='none'; btnText.textContent='Pretvarjam...';
  document.getElementById('error-msg').style.display='none';
  document.getElementById('result-box').style.display='none';
  const fd=new FormData();
  for(let f of selectedFiles) fd.append('files',f);
  fd.append('tip_bilance',selectedTip);
  fd.append('tip_subjekta',selectedTipSubjekta);
  fd.append('obdobje',selectedObdobje);
  if(selectedTip==='7' && selectedMedletniMesec) fd.append('medletni_datum',selectedMedletniMesec);
  try{
    const res=await fetch('/convert',{method:'POST',body:fd});
    if(!res.ok){
      const err=await res.json().catch(()=>({error:'Neznana napaka.'}));
      showError(err.error||'Napaka pri pretvorbi.'); return;
    }
    const downloadName=res.headers.get('X-Download-Name')||selectedFiles[0].name.replace(/\.(pdf|xlsx|xls)$/i,'')+'_CBK.xlsx';
    const name=decodeURIComponent(res.headers.get('X-Company-Name')||'');
    const reg=res.headers.get('X-Registration')||'';
    const period=res.headers.get('X-Period')||'';
    const tip=res.headers.get('X-Tip-Bilance')||'';
    const fmt=res.headers.get('X-PDF-Format')||'';
    const bs=res.headers.get('X-BS-Filled')||'';
    const ipi=res.headers.get('X-IPI-Filled')||'';
    document.getElementById('r-name').textContent=name;
    document.getElementById('r-reg').textContent=reg;
    document.getElementById('r-period').textContent=period;
    document.getElementById('r-tip').textContent=TIP_LABELS[tip]||tip;
    const tipSubj=res.headers.get('X-Tip-Subjekta')||'1';
    const subj=res.headers.get('X-Subject-Type')||'GD';
    document.getElementById('r-subj').textContent=SUBJEKT_LABELS[tipSubj]||tipSubj;
    document.getElementById('r-fmt').textContent=fmt;
    document.getElementById('r-bs').textContent=bs;
    // Pokaži opozorilo za društvo če template ni dostopen
    document.getElementById('warn-drustvo').style.display='none';
    document.getElementById('r-ipi').textContent=ipi;
    document.getElementById('result-box').style.display='block';
    const blob=await res.blob();
    const url=URL.createObjectURL(blob);
    const a=document.createElement('a');
    a.href=url; a.download=downloadName;
    document.body.appendChild(a); a.click(); document.body.removeChild(a); URL.revokeObjectURL(url);
  }catch(e){
    showError('Napaka: '+e.message);
  }finally{
    btn.disabled=false; spinner.style.display='none'; icon.style.display='block'; btnText.textContent='Pretvori v Excel';
  }
}
</script>
</body>
</html>"""


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/health")
def health():
    return jsonify({"status":"ok","template_exists":TEMPLATE_PATH.exists(),"template_sp_exists":TEMPLATE_SP_PATH.exists(),
                    "formats":["AOP","JOLP","NAPOVED","XLSX","XLS"]})

@app.route("/auth", methods=["POST"])
def auth():
    try:
        data = request.get_json(force=True, silent=True) or {}
        ok = str(data.get("password","")).strip() == ACCESS_PASSWORD
        return jsonify({"ok": ok})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500

@app.route("/convert", methods=["POST"])
def convert():
    files = request.files.getlist("files")
    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "Ni datoteke."}), 400
    if not TEMPLATE_PATH.exists():
        return jsonify({"error": "Template ni najden na strežniku."}), 500

    tip_override = request.form.get("tip_bilance", "")
    tip_subjekta_override = request.form.get("tip_subjekta", "1")
    medletni_datum_raw = request.form.get("medletni_datum", "").strip()  # format 'YYYY-MM', samo za tip_bilance == '7'
    conv_id = str(uuid.uuid4())
    saved_paths = []

    try:
        # Shrani vse naložene datoteke
        for f in files:
            if not f.filename: continue
            fname = f.filename.lower()
            if not (fname.endswith('.pdf') or fname.endswith('.xlsx') or fname.endswith('.xls')):
                return jsonify({"error": f"Nepodprt format: {f.filename}"}), 400
            ext = '.xlsx' if fname.endswith('.xlsx') else ('.xls' if fname.endswith('.xls') else '.pdf')
            p = UPLOAD_DIR / f"{conv_id}_{len(saved_paths)}{ext}"
            f.save(p)
            saved_paths.append(p)

        # Parsiraj — združi rezultate če je več fajlov
        merged = ParseResult()
        fmt_used = ""

        for path in saved_paths:
            if path.suffix == '.xlsx':
                r = parse_xlsx_file(path, tip_override, tip_subjekta_override)
            elif path.suffix == '.xls':
                r = parse_xls_file(path, tip_override, tip_subjekta_override)
            else:
                r = parse_pdf_file(path, tip_override, tip_subjekta_override)

            if r.errors:
                return jsonify({"error": "\n".join(r.errors)}), 422

            # Prva datoteka dobi podatke podjetja
            if not merged.company.name:
                merged.company = r.company
                fmt_used = r.pdf_format
                merged.subject_type = getattr(r, "subject_type", "GD")

            # Združi AOP podatke (ne prepiši obstoječih)
            for code, entry in r.aop_data.items():
                if code not in merged.aop_data:
                    merged.aop_data[code] = entry
            for gvaop, val in getattr(r, "gvaop_data", {}).items():
                if gvaop not in merged.gvaop_data:
                    merged.gvaop_data[gvaop] = val

            merged.warnings.extend(r.warnings)

        if not merged.aop_data:
            return jsonify({"error": "Ni bilo mogoče prebrati podatkov iz datotek."}), 422

        # Pri več naloženih datotekah (npr. ločena BS + IPI) se lahko isto opozorilo pojavi
        # v vsaki datoteki posebej - prikažemo ga samo enkrat.
        merged.warnings = list(dict.fromkeys(merged.warnings))

        # Medletna bilanca (tip 7): uporabnik izbere mesec, sistem uporabi zadnji dan v mesecu
        # kot datum bilance — namesto (ali poleg) datuma zaznanega iz priložene datoteke.
        if tip_override == "7" and medletni_datum_raw:
            medletni_datum = _last_day_of_month(medletni_datum_raw)
            if not medletni_datum:
                return jsonify({"error": "Neveljaven datum medletne bilance."}), 400
            leto_medletno = medletni_datum.split(".")[-1]
            merged.company.period_to   = medletni_datum
            merged.company.period_from = f"1.1.{leto_medletno}"
            merged.company.tip_bilance = "7"
            merged.company.obdobje_bilance = "ML"
        elif tip_override == "7" and not medletni_datum_raw:
            return jsonify({"error": "Za preliminarno (medletno) bilanco izberi mesec, na katerega je bilanca narejena."}), 400

        # Izvozi
        out_path = UPLOAD_DIR / f"{conv_id}_out.xlsx"
        bs_n, ipi_n = export_excel(merged, out_path)
        c = merged.company

        # Ime datoteke: NAZIV_TIP_LETO_CBK.xlsx
        # Transliteracija slovenskih znakov
        _slo = str.maketrans('ČčŠšŽžĐđ', 'CcSsZzDd')
        _name_ascii = c.name.upper().translate(_slo)
        company_slug = re.sub(r'[^A-Za-z0-9]', '_', _name_ascii)
        company_slug = re.sub(r'_+', '_', company_slug).strip('_')
        tip_label = TIP_BILANCE_LABELS.get(c.tip_bilance, "BILANCA")
        leto = c.period_to.split('.')[-1] if c.period_to else "XXXX"
        download_name = f"{company_slug}_{tip_label}_{leto}_CBK.xlsx"

        response = send_file(out_path, as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        safe_name = c.name.encode('ascii', errors='replace').decode('ascii')
        response.headers["X-Download-Name"]  = download_name
        response.headers["X-Subject-Type"]   = getattr(merged, "subject_type", "GD")
        response.headers["X-Tip-Subjekta"]   = c.tip_subjekta
        response.headers["X-Company-Name"]   = safe_name
        response.headers["X-Registration"] = c.registration_number
        response.headers["X-Period"]       = f"{c.period_from} - {c.period_to}"
        response.headers["X-Tip-Bilance"]  = c.tip_bilance
        response.headers["X-PDF-Format"]   = fmt_used
        response.headers["X-BS-Filled"]    = str(bs_n)
        response.headers["X-IPI-Filled"]   = str(ipi_n)
        response.headers["Access-Control-Expose-Headers"] = \
            "X-Download-Name,X-Subject-Type,X-Tip-Subjekta,X-Company-Name,X-Registration,X-Period,X-Tip-Bilance,X-PDF-Format,X-BS-Filled,X-IPI-Filled"
        return response

    except Exception as e:
        return jsonify({"error": f"Napaka: {str(e)}"}), 500
    finally:
        for p in saved_paths:
            try: p.unlink()
            except: pass
        try: (UPLOAD_DIR / f"{conv_id}_out.xlsx").unlink()
        except: pass

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
